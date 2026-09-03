from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..sub_models.wms_petty_cash_mod import WMSPettyCashInfo
from ..sub_models.gatein_mod import Gatein_info
from ..sub_models.credit_ledger_mod import CreditLedgerInfo
from ..models import Business_Sol_info, Location_info, ExpenseCategoryInfo, iou_info, CustomerInfo
from ..sub_forms.wms_petty_cash_form import WMSPettyCashForm
from .general_utils import get_financial_year, get_session_branch_id

def generate_wms_petty_cash_number(model_class, field_name, branch_obj=None):
    """
    Generates Voucher Number format: [BranchCode]-Wh-[MM]-[FY]-[Seq]
    Example: Maa-Wh-08-26/27-01
    """
    fy = get_financial_year() # e.g. "26-27"
    fy_slash = fy.replace('-', '/') # "26/27"
    today = datetime.now()
    month_str = today.strftime("%m") # "08"

    branch_code = "Maa"
    if branch_obj and branch_obj.loc_name:
        loc_name_upper = branch_obj.loc_name.upper()
        if "MAA" in loc_name_upper or "CHENNAI" in loc_name_upper:
            branch_code = "Maa"
        elif "BLR" in loc_name_upper or "BANGALORE" in loc_name_upper or "BENGALURU" in loc_name_upper:
            branch_code = "Blr"
        elif "PNY" in loc_name_upper or "PONDICHERRY" in loc_name_upper:
            branch_code = "Pny"
        elif "HYD" in loc_name_upper or "HYDERABAD" in loc_name_upper:
            branch_code = "Hyd"
        elif "CBE" in loc_name_upper or "COIMBATORE" in loc_name_upper:
            branch_code = "Cbe"
        else:
            branch_code = branch_obj.loc_name.split()[-1].title()

    prefix = f"{branch_code}-Wh-{month_str}-{fy_slash}-"

    # fetch latest for this prefix
    latest_obj = model_class.objects.filter(**{f"{field_name}__startswith": prefix}).order_by('-id').first()

    if latest_obj:
        latest_num_str = getattr(latest_obj, field_name)
        try:
            seq = int(latest_num_str.split('-')[-1])
            new_seq = seq + 1
        except ValueError:
            new_seq = 1
    else:
        new_seq = 1

    return f"{prefix}{str(new_seq).zfill(2)}"


def wms_petty_cash_add(request, wpc_id=0):
    initial_data = {}
    if wpc_id == 0:
        # Default Business
        bvm_storage = Business_Sol_info.objects.filter(bvm_business__icontains='Storage').first()
        if bvm_storage:
            initial_data['wpc_business'] = bvm_storage.id

        # Default Expense Category to Cash Expense
        cash_cat = ExpenseCategoryInfo.objects.filter(exp_category_name__icontains='Cash').first()
        if cash_cat:
            initial_data['wpc_category'] = cash_cat.id

        # Default Branch based on session
        branch_id = get_session_branch_id(request)
        if branch_id:
            initial_data['wpc_branch'] = branch_id

            # Default Credit Ledger based on branch name
            branch_obj = Location_info.objects.filter(id=branch_id).first()
            if branch_obj:
                branch_code = branch_obj.loc_name.split()[-1]
                ledger = CreditLedgerInfo.objects.filter(
                    ledger_name__icontains='WH Petty Cash'
                ).filter(
                    ledger_name__icontains=branch_code
                ).exclude(
                    ledger_name__icontains='Admin'
                ).first()
                if not ledger:
                    ledger = CreditLedgerInfo.objects.filter(
                        ledger_name__icontains='WH'
                    ).filter(
                        ledger_name__icontains=branch_code
                    ).first()
                if ledger:
                    initial_data['wpc_credit_ledger'] = ledger.id

        form = WMSPettyCashForm(initial=initial_data, request=request)
    else:
        wpc = get_object_or_404(WMSPettyCashInfo, pk=wpc_id)
        form = WMSPettyCashForm(instance=wpc, request=request)

    if request.method == 'POST':
        post_data = request.POST.copy()

        # Handle 'To Person' text vs User ForeignKey
        if post_data.get('wpc_to'):
            if not str(post_data['wpc_to']).isdigit():
                post_data['wpc_to_manual'] = str(post_data['wpc_to']).strip()
                post_data['wpc_to'] = ''
            else:
                post_data['wpc_to_manual'] = ''
        else:
            post_data['wpc_to_manual'] = ''

        # Auto-calculate amounts in backend
        try:
            bill_amt = float(post_data.get('wpc_bill_amount') or 0.0)
            gst_pct = float(post_data.get('wpc_gst_percentage') or 0.0)
            gst_amt = round((bill_amt * gst_pct) / 100.0, 2)
            total_amt = round(bill_amt + gst_amt, 2)
            post_data['wpc_gst_amount'] = str(gst_amt)
            post_data['wpc_total_amount'] = str(total_amt)
        except Exception:
            pass

        if wpc_id == 0:
            form = WMSPettyCashForm(post_data, request.FILES, request=request)
        else:
            wpc = get_object_or_404(WMSPettyCashInfo, pk=wpc_id)
            form = WMSPettyCashForm(post_data, request.FILES, instance=wpc, request=request)

        if form.is_valid():
            saved_wpc = form.save(commit=False)
            if wpc_id == 0:
                saved_wpc.wpc_number = generate_wms_petty_cash_number(WMSPettyCashInfo, 'wpc_number', saved_wpc.wpc_branch)
                saved_wpc.wpc_created_by = request.user
            saved_wpc.wpc_updated_by = request.user
            saved_wpc.save()
            messages.success(request, "WMS Petty Cash saved successfully.")
            return redirect('wms_petty_cash_list')
        else:
            messages.error(request, "Please correct the errors below.")

    context = {
        'form': form,
        'wpc_id': wpc_id,
    }
    return render(request, "asset_mgt_app/wms_petty_cash_add.html", context)


def wms_petty_cash_list(request):
    wpc_number = request.GET.get('wpc_number', "").strip()
    search_date = request.GET.get('search_date', "").strip()
    from_date = request.GET.get('from_date', "").strip()
    to_date = request.GET.get('to_date', "").strip()
    search_unit = request.GET.get('search_unit', "").strip()
    search_branch = request.GET.get('search_branch', "").strip()

    filters = Q()
    if wpc_number:
        filters &= Q(wpc_number__icontains=wpc_number)
    if search_date:
        filters &= Q(wpc_transaction_date=search_date)
    if from_date:
        filters &= Q(wpc_transaction_date__gte=from_date)
    if to_date:
        filters &= Q(wpc_transaction_date__lte=to_date)
    if search_unit:
        filters &= Q(wpc_unit__icontains=search_unit)

    user = request.user
    is_admin_or_supervisor = user.is_superuser
    if not is_admin_or_supervisor:
        try:
            from ..sub_models.user_ext_mod import User_extInfo
            user_ext = User_extInfo.objects.select_related('emp_designation', 'emp_role').get(user_id=user.id)
            desig = str(user_ext.emp_designation).lower() if user_ext.emp_designation else ''
            role = str(user_ext.emp_role).lower() if user_ext.emp_role else ''
            if 'supervisor' in desig or 'admin' in role:
                is_admin_or_supervisor = True
        except Exception:
            pass

    if not is_admin_or_supervisor:
        branch_id = get_session_branch_id(request)
        if branch_id:
            filters &= Q(wpc_branch_id=branch_id)
    else:
        if search_branch:
            filters &= Q(wpc_branch__loc_name__icontains=search_branch)

    wpc_list = WMSPettyCashInfo.objects.filter(filters).select_related(
        'wpc_business', 'wpc_branch', 'wpc_category', 'wpc_expense_type',
        'wpc_credit_ledger', 'wpc_to', 'wpc_customer', 'wpc_business_model',
        'wpc_created_by', 'wpc_updated_by'
    ).order_by('-id')

    paginator = Paginator(wpc_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'wpc_list': page_obj,
        'search_wpc_number': wpc_number,
        'search_date': search_date,
        'from_date': from_date,
        'to_date': to_date,
        'search_unit': search_unit,
        'search_branch': search_branch,
        'is_admin_or_supervisor': is_admin_or_supervisor,
    }
    return render(request, "asset_mgt_app/wms_petty_cash_list.html", context)


def wms_petty_cash_delete(request, wpc_id):
    wpc = get_object_or_404(WMSPettyCashInfo, pk=wpc_id)
    wpc.delete()
    messages.success(request, "WMS Petty Cash deleted successfully.")
    return redirect('wms_petty_cash_list')


def get_wms_job_details(request):
    """
    AJAX endpoint to fetch customer and business model from Job Number (Gatein_info)
    """
    job_no = request.GET.get('job_no', '').strip()
    if not job_no:
        return JsonResponse({'status': 'error', 'message': 'No job number provided'})

    gatein = Gatein_info.objects.filter(
        Q(gatein_job_no__iexact=job_no) | Q(gatein_invoice__iexact=job_no)
    ).select_related('gatein_customer', 'gatein_customer_type').first()

    if gatein:
        return JsonResponse({
            'status': 'success',
            'customer_id': gatein.gatein_customer.id if gatein.gatein_customer else None,
            'customer_name': gatein.gatein_customer.cu_name if gatein.gatein_customer else '',
            'business_model_id': gatein.gatein_customer_type.id if gatein.gatein_customer_type else None,
            'business_model_name': gatein.gatein_customer_type.tb_trbusinesstype if gatein.gatein_customer_type else '',
        })
    else:
        return JsonResponse({'status': 'not_found', 'message': 'Job number not found'})


def wms_petty_cash_export_tally(request):
    wpc_number = request.GET.get('wpc_number', "").strip()
    search_date = request.GET.get('search_date', "").strip()
    from_date = request.GET.get('from_date', "").strip()
    to_date = request.GET.get('to_date', "").strip()
    search_unit = request.GET.get('search_unit', "").strip()
    search_branch = request.GET.get('search_branch', "").strip()

    filters = Q()
    if wpc_number:
        filters &= Q(wpc_number__icontains=wpc_number)
    if search_date:
        filters &= Q(wpc_transaction_date=search_date)
    if from_date:
        filters &= Q(wpc_transaction_date__gte=from_date)
    if to_date:
        filters &= Q(wpc_transaction_date__lte=to_date)
    if search_unit:
        filters &= Q(wpc_unit__icontains=search_unit)

    user = request.user
    is_admin_or_supervisor = user.is_superuser
    if not is_admin_or_supervisor:
        try:
            from ..sub_models.user_ext_mod import User_extInfo
            user_ext = User_extInfo.objects.select_related('emp_designation', 'emp_role').get(user_id=user.id)
            desig = str(user_ext.emp_designation).lower() if user_ext.emp_designation else ''
            role = str(user_ext.emp_role).lower() if user_ext.emp_role else ''
            if 'supervisor' in desig or 'admin' in role:
                is_admin_or_supervisor = True
        except Exception:
            pass

    if not is_admin_or_supervisor:
        branch_id = get_session_branch_id(request)
        if branch_id:
            filters &= Q(wpc_branch_id=branch_id)
    else:
        if search_branch:
            filters &= Q(wpc_branch__loc_name__icontains=search_branch)

    wpc_list = WMSPettyCashInfo.objects.filter(filters).select_related(
        'wpc_business', 'wpc_branch', 'wpc_category', 'wpc_expense_type',
        'wpc_credit_ledger', 'wpc_to', 'wpc_customer', 'wpc_business_model'
    ).order_by('-id')

    wb = Workbook()
    ws = wb.active
    ws.title = "WMS Petty Cash Export"

    headers = [
        "VOUCHER NUMBER", "TRANSACTION DATE", "BUSINESS", "BRANCH", 
        "EXPENSES CATEGORY", "EXPENSES TYPE", "CREDIT LEDGER", "TO PERSON",
        "UNIT", "JOB NO", "CUSTOMER NAME", "BUSINESS MODEL",
        "BILL NO", "BILL AMOUNT", "GST %", "GST AMOUNT", "TOTAL AMOUNT", "REMARKS"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFFF00")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for wpc in wpc_list:
        date_str = wpc.wpc_transaction_date.strftime("%d-%m-%Y") if wpc.wpc_transaction_date else ""
        vou_no = wpc.wpc_number or ""
        business = wpc.wpc_business.bvm_business if wpc.wpc_business else ""
        branch = wpc.wpc_branch.loc_name if wpc.wpc_branch else ""
        category = wpc.wpc_category.exp_category_name if wpc.wpc_category else ""
        exp_type = wpc.wpc_expense_type.wms_exp_type_name if wpc.wpc_expense_type else ""
        credit_ledger = wpc.wpc_credit_ledger.ledger_name if wpc.wpc_credit_ledger else ""
        to_person = wpc.wpc_to.first_name if wpc.wpc_to else (wpc.wpc_to_manual or "")
        unit = wpc.wpc_unit or ""
        job_no = wpc.wpc_job_no or ""
        customer = wpc.wpc_customer.cu_name if wpc.wpc_customer else ""
        bus_model = wpc.wpc_business_model.tb_trbusinesstype if wpc.wpc_business_model else ""
        bill_no = wpc.wpc_bill_no or ""
        bill_amt = wpc.wpc_bill_amount or 0.0
        gst_pct = wpc.wpc_gst_percentage or 0.0
        gst_amt = wpc.wpc_gst_amount or 0.0
        total_amt = wpc.wpc_total_amount or 0.0
        remarks = wpc.wpc_remarks or ""

        ws.append([
            vou_no, date_str, business, branch,
            category, exp_type, credit_ledger, to_person,
            unit, job_no, customer, bus_model,
            bill_no, bill_amt, gst_pct, gst_amt, total_amt, remarks
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="WMS_Petty_Cash_Export.xlsx"'
    wb.save(response)
    return response
