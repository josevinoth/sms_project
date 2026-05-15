from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.core.paginator import Paginator
from django.db.models import Sum, Max, Min, Q
from django.http import HttpResponse
import json
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from ..forms import InvoiceaddForm
from ..models import VehicletypeInfo, Loadingbay_Info, TrbusinesstypeInfo, CustomerInfo, Warehouse_goods_info, \
    WhratemasterInfo, BilingInfo
from django.shortcuts import render, redirect
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font

from ..sub_models.locationmaster_mod import LocationmasterInfo


# Invoicecity
@login_required(login_url='login_page')
def invoice_add(request, invoice_id=0):
    first_name = request.session.get('first_name')
    user_id = request.session.get('ses_userID')
    if request.method == "GET":
        if invoice_id == 0:
            invoice_form = InvoiceaddForm()
            context = {
                'invoice_form': invoice_form,
                'first_name': first_name,
                'user_id': user_id,
            }
        else:
            invoice = BilingInfo.objects.get(pk=invoice_id)
            invoice_form = InvoiceaddForm(instance=invoice)
            voucher_num = invoice.bill_invoice_ref

            # goods rows linked with this invoice
            goods_qs = Warehouse_goods_info.objects.filter(
                Q(wh_voucher_id=invoice) | Q(wh_voucher_num=voucher_num)
            )
            count_stocks = goods_qs.count()

            # check whether shipper details added - block if no goods linked
            if count_stocks == 0:
                messages.warning(request, 'No Shipper Invoices linked to this invoice. Please add goods first.')
                # Build a context from the invoice instance so the form renders with saved values
                context = {
                    'invoice_form': invoice_form,
                    'first_name': first_name,
                    'user_id': user_id,
                    'shipper_invoice_list': goods_qs,
                    'weight_sum': invoice.bill_weight or 0,
                    'no_of_days': invoice.bill_no_of_days or 0,
                    'no_of_pieces': invoice.bill_no_of_pallets or 0,
                    'crane_time': invoice.bill_tot_crane_time or 0,
                    'forklift_time': invoice.bill_tot_forklift_time or 0,
                    'min_check_in_time': str(invoice.bill_start_date) if invoice.bill_start_date else '',
                    'max_check_out_time': str(invoice.bill_end_date) if invoice.bill_end_date else '',
                    'total_loading_cost': invoice.bill_loading_charge or 0,
                    'wh_storage_cost_sum': invoice.bill_wh_storage_charges or 0,
                    'crane_cost_sum': invoice.bill_tot_crane_charges or 0,
                    'forklift_cost_sum': invoice.bill_tot_forklift_charges or 0,
                    'customer_type_id': invoice.bill_customer_type.id if invoice.bill_customer_type else 0,
                }
            else:
                # Calculate Warehouse Storage Charges
                dispatch_num = goods_qs.values_list('wh_dispatch_num', flat=True).distinct()
                customer_obj = invoice.bill_customer_name
                customer_id = customer_obj.id
                customer_type_obj = customer_obj.cu_businessmodel
                customer_type_id = customer_type_obj.id
                wh_job_num = goods_qs.values_list('wh_job_no', flat=True).distinct()
                wh_job_num_count = len(wh_job_num)
                total_weight_val = goods_qs.aggregate(Sum('wh_goods_weight'))['wh_goods_weight__sum']

                # check total weight limits
                if total_weight_val is not None:
                    total_weight = total_weight_val
                else:
                    total_weight = 0
                    # Only warn if goods ARE linked but weight data is missing
                    if count_stocks > 0 and customer_type_id != 3:
                        messages.warning(request,
                                         'Unable to Calculate Total Weight — please ensure all shipment goods have weight entered.')

                # check total area limits
                total_area_val = \
                Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num).aggregate(Sum('wh_goods_area'))[
                    'wh_goods_area__sum']
                if total_area_val is not None:
                    total_area = total_area_val
                else:
                    total_area = 0
                    # Only warn if goods ARE linked but area data is missing
                    if count_stocks > 0 and customer_type_id != 3:
                        messages.warning(request,
                                         'Unable to Calculate Total Area — please ensure all shipment goods have area entered.')

                # check warehouse charges based on customer type
                if customer_type_id == 2:
                    print("Inside Exclusive Case")
                    try:
                        warehouse_charge = WhratemasterInfo.objects.get(whrm_customer_name=customer_id,
                                                                        whrm_charge_type=1).whrm_rate
                    except ObjectDoesNotExist:
                        messages.error(request,
                                       'Warehouse Storage Charges not available in master for selected Customer!')
                        return redirect(request.META['HTTP_REFERER'])
                    try:
                        max_wt = WhratemasterInfo.objects.get(whrm_customer_name=customer_id,
                                                              whrm_charge_type=1).whrm_max_wt
                        if count_stocks > 0:  # Only show weight limit messages when goods are linked
                            if total_weight <= max_wt:
                                messages.success(request, 'Total weight within customer limit!')
                            else:
                                messages.error(request, 'Total weight exceeds customer limit!')
                    except ObjectDoesNotExist:
                        max_wt = 0
                        if count_stocks > 0:
                            messages.error(request, 'Max Weight not available in master for selected Customer!')
                            return redirect(request.META['HTTP_REFERER'])
                    try:
                        warehouse_charge_1 = warehouse_charge / wh_job_num_count
                    except ZeroDivisionError:
                        warehouse_charge_1 = 0
                    storage_cost_total = round((warehouse_charge_1), 2)
                    min_check_in_time = BilingInfo.objects.get(pk=invoice_id).bill_start_date
                    max_check_out_time = BilingInfo.objects.get(pk=invoice_id).bill_end_date
                    if max_check_out_time is not None and min_check_in_time is not None:
                        max_storage_days = (max_check_out_time - min_check_in_time).days + 1
                    else:
                        # Handle the case when one or both values are None
                        max_storage_days = 0  # or set an appropriate default value

                    exclusive_invoices = list((Warehouse_goods_info.objects.filter(
                        wh_voucher_num=voucher_num).values_list('wh_gate_injob_no_id', flat=True)).distinct())
                    for inv in exclusive_invoices:
                        exclusive_goods_ids = list(Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num,
                                                                                       wh_gate_injob_no_id=inv).values_list(
                            'id', flat=True))
                        for i in range(0, len(exclusive_goods_ids)):
                            if i == 0:
                                Warehouse_goods_info.objects.filter(pk=exclusive_goods_ids[i]).update(
                                    wh_storage_cost_per_day=round(warehouse_charge_1, 2),
                                    wh_storage_cost_total=storage_cost_total, wh_voucher_id=invoice)
                            else:
                                Warehouse_goods_info.objects.filter(pk=exclusive_goods_ids[i]).update(
                                    wh_storage_cost_per_day=0, wh_storage_cost_total=0, wh_voucher_id=invoice)


                elif customer_type_id == 3:
                    print("Inside Dedicated Case")
                    try:
                        warehouse_charge = WhratemasterInfo.objects.get(whrm_customer_name=customer_id,
                                                                        whrm_charge_type=1).whrm_rate
                    except ObjectDoesNotExist:
                        messages.error(request,
                                       'Warehouse Storage Charges not available in master for selected Customer!')
                        return redirect(request.META['HTTP_REFERER'])
                    try:
                        max_area = WhratemasterInfo.objects.get(whrm_customer_name=customer_id,
                                                                whrm_charge_type=1).whrm_max_area
                        if count_stocks > 0:  # Only show area limit messages when goods are linked
                            if total_area <= max_area:
                                messages.success(request, 'Total Area within customer limit!')
                            else:
                                messages.error(request, 'Total Area exceeds customer limit!')
                    except ObjectDoesNotExist:
                        max_area = 0
                        if count_stocks > 0:
                            messages.error(request, 'Max Area not available in master for selected Customer!')
                            return redirect(request.META['HTTP_REFERER'])
                    try:
                        warehouse_charge_1 = warehouse_charge / wh_job_num_count
                    except ZeroDivisionError:
                        warehouse_charge_1 = 0
                    storage_cost_total = round((warehouse_charge_1), 2)
                    min_check_in_time = BilingInfo.objects.get(pk=invoice_id).bill_start_date
                    max_check_out_time = BilingInfo.objects.get(pk=invoice_id).bill_end_date
                    if max_check_out_time is not None and min_check_in_time is not None:
                        max_storage_days = (max_check_out_time - min_check_in_time).days + 1
                    else:
                        # Handle the case when one or both values are None
                        max_storage_days = 0  # or set an appropriate default value

                    dedicated_invoices = list((Warehouse_goods_info.objects.filter(
                        wh_voucher_num=voucher_num).values_list('wh_gate_injob_no_id', flat=True)).distinct())
                    for inv in dedicated_invoices:
                        dedicated_goods_ids = list(Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num,
                                                                                       wh_gate_injob_no_id=inv).values_list(
                            'id', flat=True))
                        for i in range(0, len(dedicated_goods_ids)):
                            if i == 0:
                                Warehouse_goods_info.objects.filter(pk=dedicated_goods_ids[i]).update(
                                    wh_storage_cost_per_day=round(warehouse_charge_1, 2),
                                    wh_storage_cost_total=storage_cost_total, wh_voucher_id=invoice)
                            else:
                                Warehouse_goods_info.objects.filter(pk=dedicated_goods_ids[i]).update(
                                    wh_storage_cost_per_day=0, wh_storage_cost_total=0, wh_voucher_id=invoice)
                else:
                    print("Inside Case To Case")
                    # Get billing_truck_type for voucher number
                    billing_truck_type = list(
                        Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num)
                        .values_list('wh_dispatch_id__dispatch_billing_truck_type', flat=True)
                        .distinct()
                    )

                    for btt in billing_truck_type:
                        if btt == 1:
                            pre_gate_in_nums = sorted(
                                Warehouse_goods_info.objects.filter(
                                    wh_voucher_num=voucher_num,
                                    wh_dispatch_id__dispatch_billing_truck_type=btt
                                ).values_list('wh_gate_injob_no_id__gatein_pre_id', flat=True)
                                .distinct()
                            )

                            for pgn in pre_gate_in_nums:
                                vehicle_type = Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num,
                                                                                   wh_gate_injob_no_id__gatein_pre_id=pgn
                                                                                   ).values_list(
                                    'wh_gate_injob_no_id__gatein_truck_type', flat=True).first()
                                print('vehicle_type', vehicle_type)
                                vehicle_type_id = VehicletypeInfo.objects.get(vt_vehicletype=vehicle_type).id
                                truck_num = sorted(
                                    Warehouse_goods_info.objects.filter(
                                        wh_gate_injob_no_id__gatein_pre_id=pgn
                                    ).values_list('wh_gate_injob_no_id__gatein_truck_number', flat=True)
                                    .distinct()
                                )

                                try:
                                    warehouse_charge = WhratemasterInfo.objects.get(
                                        whrm_customer_name=customer_id,
                                        whrm_charge_type=1,
                                        whrm_vehicle_type=vehicle_type_id
                                    ).whrm_rate
                                except ObjectDoesNotExist:
                                    messages.error(
                                        request,
                                        f'Warehouse Storage Charges not available in master for selected Customer and Vehicle Type {vehicle_type_id}!'
                                    )
                                    return redirect(request.META['HTTP_REFERER'])

                                max_storage_days = Warehouse_goods_info.objects.filter(
                                    wh_gate_injob_no_id__gatein_pre_id=pgn
                                ).values_list('wh_storage_time', flat=True).distinct().aggregate(
                                    Max('wh_storage_time'))['wh_storage_time__max']

                                storage_cost_total = round(warehouse_charge * max_storage_days, 2)

                                for tns in truck_num:
                                    ids = list(
                                        Warehouse_goods_info.objects.filter(
                                            wh_voucher_num=voucher_num,
                                            wh_gate_injob_no_id__gatein_pre_id=pgn,
                                            wh_gate_injob_no_id__gatein_truck_number=tns
                                        ).values_list('id', flat=True)
                                    )

                                    Warehouse_goods_info.objects.filter(pk=ids[0]).update(
                                        wh_storage_cost_per_day=round(warehouse_charge, 2),
                                        wh_storage_cost_total=storage_cost_total
                                    )
                                    Warehouse_goods_info.objects.filter(pk__in=ids[1:]).update(
                                        wh_storage_cost_per_day=0,
                                        wh_storage_cost_total=0
                                    )

                        elif btt == 2:
                            dispatch_ids = sorted(
                                Warehouse_goods_info.objects.filter(
                                    wh_voucher_num=voucher_num,
                                    wh_dispatch_id__dispatch_billing_truck_type=btt
                                ).values_list('wh_dispatch_id', flat=True)
                                .distinct()
                            )

                            for dis in dispatch_ids:
                                vehicle_type_id = Warehouse_goods_info.objects.filter(
                                    wh_dispatch_id=dis
                                ).values_list('wh_dispatch_id__dispatch_truck_type', flat=True).first()

                                truck_num = sorted(
                                    Warehouse_goods_info.objects.filter(
                                        wh_dispatch_id=dis
                                    ).values_list('wh_dispatch_id__dispatch_truck_number', flat=True)
                                    .distinct()
                                )

                                try:
                                    warehouse_charge = WhratemasterInfo.objects.get(
                                        whrm_customer_name=customer_id,
                                        whrm_charge_type=1,
                                        whrm_vehicle_type=vehicle_type_id
                                    ).whrm_rate
                                except ObjectDoesNotExist:
                                    messages.error(
                                        request,
                                        f'Warehouse Storage Charges not available in master for selected Customer and Vehicle Type {vehicle_type_id}!'
                                    )
                                    return redirect(request.META['HTTP_REFERER'])

                                max_storage_days = Warehouse_goods_info.objects.filter(
                                    wh_dispatch_id=dis
                                ).values_list('wh_storage_time', flat=True).distinct().aggregate(
                                    Max('wh_storage_time'))['wh_storage_time__max']

                                storage_cost_total = round(warehouse_charge * max_storage_days, 2)

                                for tns in truck_num:
                                    ids = list(
                                        Warehouse_goods_info.objects.filter(
                                            wh_voucher_num=voucher_num,
                                            wh_dispatch_id=dis,
                                            wh_dispatch_id__dispatch_truck_number=tns
                                        ).values_list('id', flat=True)
                                    )

                                    Warehouse_goods_info.objects.filter(pk=ids[0]).update(
                                        wh_storage_cost_per_day=round(warehouse_charge, 2),
                                        wh_storage_cost_total=storage_cost_total
                                    )
                                    Warehouse_goods_info.objects.filter(pk__in=ids[1:]).update(
                                        wh_storage_cost_per_day=0,
                                        wh_storage_cost_total=0
                                    )

                # check Crane and forklift charges
                for k in wh_job_num:
                    print('k', k)
                    lb = Loadingbay_Info.objects.filter(lb_job_no=k).order_by('-id').first()

                    # Calculate Loading & Unloading Charge
                    if lb and lb.lb_mh_manual:
                        manual_handling_status = lb.lb_mh_manual.id
                    else:
                        manual_handling_status = 0
                    if manual_handling_status == 1:
                        total_weight = \
                        Warehouse_goods_info.objects.filter(wh_job_no=k).aggregate(Sum('wh_goods_weight'))[
                            'wh_goods_weight__sum']
                        no_of_pieces = \
                        Warehouse_goods_info.objects.filter(wh_job_no=k).aggregate(Sum('wh_goods_pieces'))[
                            'wh_goods_pieces__sum']
                        try:
                            weight_per_piece = round((total_weight) / (no_of_pieces), 0)
                        except ZeroDivisionError:
                            weight_per_piece = float(0.0)

                        if customer_type_id == 2:
                            piece_rate_val = 0
                            total_loading_cost = piece_rate_val * no_of_pieces
                        else:
                            try:
                                piece_rate = WhratemasterInfo.objects.get(whrm_customer_name=customer_id,
                                                                          whrm_min_wt__lte=weight_per_piece,
                                                                          whrm_max_wt__gte=weight_per_piece,
                                                                          whrm_charge_type=3)
                                piece_rate_val = piece_rate.whrm_rate
                                total_loading_cost = piece_rate_val * no_of_pieces
                            except ObjectDoesNotExist:
                                messages.error(request,
                                               'Loading/Unloading Charges not available in master for selected Customer for weight! ' + str(
                                                   weight_per_piece) + str(' kg'))
                                return redirect(request.META['HTTP_REFERER'])
                            except MultipleObjectsReturned:
                                messages.error(request,
                                               'Multiple loading/unloading charge rates found in master for selected Customer and weight! Please check the master data.' + str(
                                                   weight_per_piece) + str(' kg'))
                                return redirect(request.META['HTTP_REFERER'])
                    else:
                        piece_rate_val = 0
                        total_loading_cost = 0
                    # Calculate Crane and Forklift cost
                    try:
                        if lb:
                            crane_hours = lb.lb_crane_time or 0
                            forklift_hours = lb.lb_forklift_time or 0
                            forklift_charge_l2h = lb.lb_forklift_charges_mod_l2h or 0
                            forklift_charge_g2h = lb.lb_forklift_charges_mod_g2hr or 0
                            crane_charge_l2h = lb.lb_crane_charges_mod_l2h or 0
                            crane_charge_g2h = lb.lb_crane_charges_mod_g2hr or 0
                            no_of_cranes = lb.lb_no_of_crane or 0
                            no_of_forklifts = lb.lb_no_of_forklift or 0
                        else:
                            crane_hours = 0
                            forklift_hours = 0
                            forklift_charge_l2h = 0
                            forklift_charge_g2h = 0
                            crane_charge_l2h = 0
                            crane_charge_g2h = 0
                            no_of_cranes = 0
                            no_of_forklifts = 0

                    except ObjectDoesNotExist:
                        crane_hours = 0
                        forklift_hours = 0
                        forklift_charge_l2h = 0
                        forklift_charge_g2h = 0
                        crane_charge_l2h = 0
                        crane_charge_g2h = 0
                        no_of_cranes = 0
                        no_of_forklifts = 0
                    if crane_hours <= 2 and forklift_hours <= 2:
                        print("inside Condition 1")
                        crane_cost_l2hr = round((1 * crane_charge_l2h * no_of_cranes), 2)
                        crane_cost_g2hr = 0
                        forklift_cost_l2hr = round((1 * forklift_charge_l2h * no_of_forklifts), 2)
                        forklift_cost_g2hr = 0
                        crane_cost = crane_cost_l2hr + crane_cost_g2hr
                        forklift_cost = forklift_cost_l2hr + forklift_cost_g2hr
                    elif forklift_hours <= 2 and crane_hours > 2:
                        print("inside Condition 3")
                        crane_hours_aft_2 = int(crane_hours) - 2
                        crane_cost_l2hr = round((1 * crane_charge_l2h * no_of_cranes), 2)
                        crane_cost_g2hr = round((crane_charge_g2h * crane_hours_aft_2 * no_of_cranes), 2)
                        forklift_cost_l2hr = round((1 * forklift_charge_l2h * no_of_forklifts), 2)
                        forklift_cost_g2hr = 0

                        crane_cost = crane_cost_l2hr + crane_cost_g2hr
                        forklift_cost = forklift_cost_l2hr + forklift_cost_g2hr
                    elif crane_hours <= 2 and forklift_hours > 2:
                        print("inside Condition 4")
                        forklift_hours_aft_2 = forklift_hours - 2
                        crane_cost_l2hr = round((1 * crane_charge_l2h * no_of_cranes), 2)
                        crane_cost_g2hr = 0
                        forklift_cost_l2hr = round((1 * forklift_charge_l2h * no_of_forklifts), 2)
                        forklift_cost_g2hr = round((forklift_charge_g2h * forklift_hours_aft_2 * no_of_forklifts), 2)

                        crane_cost = crane_cost_l2hr + crane_cost_g2hr
                        forklift_cost = forklift_cost_l2hr + forklift_cost_g2hr
                    else:
                        print("inside Condition 5")
                        crane_hours_aft_2 = int(crane_hours) - 2
                        forklift_hours_aft_2 = forklift_hours - 2
                        crane_cost_l2hr = round((1 * crane_charge_l2h * no_of_cranes), 2)
                        crane_cost_g2hr = round((crane_charge_g2h * crane_hours_aft_2 * no_of_cranes), 2)
                        forklift_cost_l2hr = round((1 * forklift_charge_l2h * no_of_forklifts), 2)
                        forklift_cost_g2hr = round((forklift_charge_g2h * forklift_hours_aft_2 * no_of_forklifts), 2)

                        crane_cost = crane_cost_l2hr + crane_cost_g2hr
                        forklift_cost = forklift_cost_l2hr + forklift_cost_g2hr

                    invoice_id = list(Warehouse_goods_info.objects.filter(wh_job_no=k).values_list('id', flat=True))
                    invoice_id.sort()

                    for i in range(0, len(invoice_id)):
                        if i == 0:
                            # Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_storage_cost_per_day=round(warehouse_charge_1,2))
                            # Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_storage_cost_total=storage_cost_total)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(
                                wh_crane_cost_l2h=crane_cost_l2hr)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(
                                wh_crane_cost_g2h=crane_cost_g2hr)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_crane_cost=crane_cost)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(
                                wh_forklift_cost_l2hr=forklift_cost_l2hr)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(
                                wh_forklift_cost_g2hr=forklift_cost_g2hr)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_forklift_cost=forklift_cost)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(
                                wh_loading_charge_unit=piece_rate_val)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(
                                wh_total_loading_cost=total_loading_cost)
                        else:
                            # Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_storage_cost_per_day=0)
                            # Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update( wh_storage_cost_total=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_crane_cost_l2h=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_crane_cost_g2h=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_crane_cost=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_forklift_cost_l2hr=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_forklift_cost_g2hr=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_forklift_cost=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_loading_charge_unit=0)
                            Warehouse_goods_info.objects.filter(pk=invoice_id[i]).update(wh_total_loading_cost=0)

                # Total Cost calculation
                shipper_invoice_list = goods_qs
                weight_sum = goods_qs.aggregate(Sum('wh_goods_weight'))['wh_goods_weight__sum'] or 0
                crane_cost_sum = goods_qs.aggregate(Sum('wh_crane_cost'))['wh_crane_cost__sum'] or 0
                forklift_cost_sum = goods_qs.aggregate(Sum('wh_forklift_cost'))['wh_forklift_cost__sum'] or 0
                wh_storage_cost_sum = goods_qs.aggregate(Sum('wh_storage_cost_total'))[
                                          'wh_storage_cost_total__sum'] or 0
                no_of_days = goods_qs.aggregate(Max('wh_storage_time'))['wh_storage_time__max'] or 0
                no_of_pieces = goods_qs.aggregate(Sum('wh_goods_pieces'))['wh_goods_pieces__sum'] or 0
                total_loading_cost = goods_qs.aggregate(Sum('wh_total_loading_cost'))['wh_total_loading_cost__sum'] or 0

                if customer_type_id in [2, 3]:
                    # Exclusive and Dedicated: always use full flat rate from Warehouse Rate Master
                    wh_storage_cost_sum = warehouse_charge
                else:
                    if wh_storage_cost_sum is None:
                        wh_storage_cost_sum = 0

                job_num = goods_qs.values_list('wh_job_no', flat=True).distinct()
                crane_time = 0
                forklift_time = 0
                for i in job_num:
                    lb = Loadingbay_Info.objects.filter(lb_job_no=i).order_by('-id').first()
                    if lb:
                        crane_time += lb.lb_crane_time or 0
                        forklift_time += lb.lb_forklift_time or 0

                # calculate checkin_times & checkout_times, max_storage_days for Invoice voucher
                try:
                    # Extract the list of check-in times
                    checkin_times = goods_qs.values_list('wh_checkin_time', flat=True)
                    # Find the minimum check-in time
                    min_check_in_time = min(checkin_times)
                    # Convert to datetime if necessary
                    if isinstance(min_check_in_time, datetime):
                        min_check_in_time = min_check_in_time.date()

                    # Extract the list of check-out times
                    checkout_times = goods_qs.values_list('wh_checkout_time', flat=True)
                    # Find the maximum check-out time
                    max_check_out_time = max(checkout_times)
                    # Convert to datetime if necessary
                    if isinstance(max_check_out_time, datetime):
                        max_check_out_time = max_check_out_time.date()
                    max_storage_days = ((max_check_out_time - min_check_in_time).days)
                except:
                    min_check_in_time = 0
                    max_check_out_time = 0
                    max_storage_days = ((max_check_out_time - min_check_in_time))

                context = {
                    'user_id': user_id,
                    'invoice_form': invoice_form,
                    'first_name': first_name,
                    'shipper_invoice_list': shipper_invoice_list,
                    'weight_sum': weight_sum,
                    # 'no_of_days':no_of_days,
                    'no_of_days': max_storage_days,
                    'no_of_pieces': no_of_pieces,
                    'crane_time': crane_time,
                    'forklift_time': forklift_time,
                    'min_check_in_time': str(min_check_in_time),
                    'max_check_out_time': str(max_check_out_time),
                    'total_loading_cost': total_loading_cost,
                    'wh_storage_cost_sum': round(wh_storage_cost_sum, 2),
                    'crane_cost_sum': crane_cost_sum,
                    'forklift_cost_sum': forklift_cost_sum,
                    'customer_type_id': customer_type_id,
                }
        return render(request, "asset_mgt_app/invoice_add.html", context)
    else:
        if invoice_id == 0:
            invoice_form = InvoiceaddForm(request.POST)
            if invoice_form.is_valid():
                invoice = invoice_form.save()
                voucher_num_val = invoice.bill_invoice_ref
                # Link goods by ID for reliable exports
                Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num_val).update(wh_voucher_id=invoice)
                print("Main Form Saved")
                messages.success(request, 'Record Added Successfully!')
            else:
                print("Main Form Not Saved")
                messages.error(request, 'Check all mandatory fields!')
            return redirect('/SMS/invoice_list')
        else:
            invoice = BilingInfo.objects.get(pk=invoice_id)
            invoice_form = InvoiceaddForm(request.POST, instance=invoice)
            if invoice_form.is_valid():
                invoice = invoice_form.save()  # Save the updated form data (including bill_total_post_gst)
                messages.success(request, 'Invoice Amount Updated and Saved Successfully!')
                voucher_num_val = invoice.bill_invoice_ref
                # update total invoice cost and link by ID in warehouse goods table
                Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num_val).update(wh_voucher_id=invoice)
                stock_id = list(
                    Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num_val).values_list('id', flat=True))
                total_invoice_cost = invoice.bill_total_pre_gst
                stock_id.sort()
                for i in range(0, len(stock_id)):
                    if i == 0:
                        Warehouse_goods_info.objects.filter(pk=stock_id[i]).update(
                            wh_total_invoice_cost=total_invoice_cost)
                    else:
                        Warehouse_goods_info.objects.filter(pk=stock_id[i]).update(wh_total_invoice_cost=0)
            else:
                print("Main Form Not Saved")
                messages.error(request, 'Check all mandatory fields!')
            return redirect(request.META['HTTP_REFERER'])
            # return redirect('/SMS/invoice_list')


@login_required(login_url='login_page')
def invoice_report(request):
    first_name = request.session.get('first_name')
    goods_list = Warehouse_goods_info.objects.exclude(wh_voucher_num=None)
    context = {
        'first_name': first_name,
        'goods_list': goods_list,
    }
    return render(request, "asset_mgt_app/invoice_report.html", context)


@login_required(login_url='login_page')
def invoice_list(request):
    first_name = request.session.get('first_name')
    invoice_list_val = (BilingInfo.objects.all()).order_by('-id')
    page_number = request.GET.get('page')
    paginator = Paginator(invoice_list_val, 50)
    page_obj = paginator.get_page(page_number)
    context = {
        'invoice_list_val': invoice_list_val,
        'page_obj': page_obj,
        'first_name': first_name,
    }
    return render(request, "asset_mgt_app/invoice_list.html", context)


@login_required(login_url='login_page')
def invoice_delete(request, invoice_id):
    invoice_del = BilingInfo.objects.get(pk=invoice_id)
    invoice_ref = invoice_del.bill_invoice_ref
    wh_jobs = list(Warehouse_goods_info.objects.filter(
        Q(wh_voucher_id=invoice_del) | Q(wh_voucher_num=invoice_ref)
    ).values_list('wh_job_no', flat=True))
    for i in wh_jobs:
        Warehouse_goods_info.objects.filter(wh_job_no=i).update(wh_voucher_num=None)
        Warehouse_goods_info.objects.filter(wh_job_no=i).update(wh_voucher_id=None)
    invoice_del.delete()
    return redirect('/SMS/invoice_list')


@login_required(login_url='login_page')
def shipper_invoice_list(request, voucher_id):
    first_name = request.session.get('first_name')
    invoice = BilingInfo.objects.get(pk=voucher_id)
    voucher_num_val = invoice.bill_invoice_ref
    customer_name_val = invoice.bill_customer_name
    customer_obj = customer_name_val  # invoice.bill_customer_name is already a CustomerInfo object
    customer_type_obj = customer_obj.cu_businessmodel
    customer_type_id = customer_type_obj.id
    billing_start_date = invoice.bill_start_date
    billing_end_date = invoice.bill_end_date
    request.session['ses_voucher_num_val'] = voucher_num_val
    request.session['ses_voucher_id'] = voucher_id
    shipper_invoice_list = Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num_val)
    if customer_type_id > 1:
        print("Inside Exclusive Loop")
        try:
            invoice_list_master = Warehouse_goods_info.objects.filter(wh_customer_name=customer_name_val,
                                                                      wh_checkin_time__gte=billing_start_date,
                                                                      wh_check_in_out=2,
                                                                      wh_checkin_time__lte=billing_end_date,
                                                                      wh_voucher_num=None)
            # invoice_list_master = Warehouse_goods_info.objects.filter(wh_customer_name=customer_name_val,wh_checkin_time__gte=billing_start_date,wh_checkin_time__lte=billing_end_date,wh_voucher_num=None)
        except ValueError:
            messages.error(request, 'Check Billing Start & End Date!')
            return redirect(request.META['HTTP_REFERER'])
    else:
        print("Inside Non Exclusive Loop")
        try:
            invoice_list_master = Warehouse_goods_info.objects.filter(wh_customer_name=customer_name_val,
                                                                      wh_check_in_out=2, wh_voucher_num=None)
        except ValueError:
            messages.error(request, 'Check Billing Start & End Date!')
            return redirect(request.META['HTTP_REFERER'])
    context = {
        'shipper_invoice_list': shipper_invoice_list,
        'first_name': first_name,
        'invoice_list_master': invoice_list_master,
    }
    return render(request, "asset_mgt_app/shipper_invoice_list.html", context)


@csrf_exempt
@login_required(login_url='login_page')
def shipper_invoice_goods_add(request):
    voucher_num_val = request.session.get('ses_voucher_num_val')
    voucher_id_val = request.session.get('ses_voucher_id')
    print(voucher_num_val)
    selected_stocks = request.POST.getlist('myList[]') or request.GET.getlist('myList[]')
    first_name = request.session.get('first_name')
    if selected_stocks:
        Warehouse_goods_info.objects.filter(wh_qr_rand_num__in=selected_stocks).update(
            wh_voucher_num=voucher_num_val,
            wh_voucher_id=voucher_id_val
        )
    invoice_list_1 = Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num_val)

    context = {
        'first_name': first_name,
        'shipper_invoice_list': invoice_list_1,
    }
    # return redirect(request.META['HTTP_REFERER'])
    # return redirect('/SMS/dispatch_goods_list')
    return redirect('shipperinvoice_list', voucher_id=voucher_id_val)


@csrf_exempt
@login_required(login_url='login_page')
def shipper_invoice_goods_remove(request):
    voucher_num_val = request.session.get('ses_voucher_num_val')
    voucher_id_val = request.session.get('ses_voucher_id')
    selected_stocks = request.POST.getlist('myList[]') or request.GET.getlist('myList[]')
    first_name = request.session.get('first_name')
    if selected_stocks:
        Warehouse_goods_info.objects.filter(wh_qr_rand_num__in=selected_stocks).update(
            wh_voucher_num=None,
            wh_voucher_id=None
        )
    invoice_list_1 = Warehouse_goods_info.objects.filter(wh_voucher_num=voucher_num_val)

    context = {
        'first_name': first_name,
        'shipper_invoice_list': invoice_list_1,
    }
    # return redirect(request.META['HTTP_REFERER'])
    # return redirect('/SMS/dispatch_goods_list')
    return redirect('shipperinvoice_list', voucher_id=voucher_id_val)


# Custom serialization function for date objects
def custom_serializer(obj):
    if isinstance(obj, date):
        return obj.isoformat()
    raise TypeError("Type not serializable")


@login_required(login_url='login_page')
def load_whrate_model(request):
    lm_customer_name_id = request.GET.get('lm_customer_name_id')
    customer_id = CustomerInfo.objects.get(cu_name=lm_customer_name_id).id
    # customer_type = CustomerInfo.objects.get(id=customer_id).cu_businessmodel
    # customer_type_id = TrbusinesstypeInfo.objects.get(tb_trbusinesstype=customer_type).id
    # if customer_type_id > 1:
    #     date_today = date.today()
    #     year = date_today.year
    #     month = date_today.month
    #     min_check_in_time = date(year, month, 1)
    #     if month == 12:
    #         max_check_out_time = date(year, month, 31)
    #     else:
    #         max_check_out_time = date(year, month + 1, 1) + timedelta(days=-1)
    #     max_storage_days = ((max_check_out_time - min_check_in_time).days) + 1

    customer_businessmodel = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_businessmodel')
    customer_short_name = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_nameshort')
    customer_code = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_customercode')
    customer_GST = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_gst')
    customer_person = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_customerperson')
    customer_contact = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_contactno')
    customer_address = CustomerInfo.objects.filter(cu_name=lm_customer_name_id).values('cu_address')
    customer_businessmodel_val = customer_businessmodel[0]['cu_businessmodel']  # Get value from Queryset
    customer_short_name_val = customer_short_name[0]['cu_nameshort']  # Get value from Queryset
    customer_code_val = customer_code[0]['cu_customercode']  # Get value from Queryset
    customer_GST_val = customer_GST[0]['cu_gst']  # Get value from Queryset
    customer_person_val = customer_person[0]['cu_customerperson']  # Get value from Queryset
    customer_contact_val = customer_contact[0]['cu_contactno']  # Get value from Queryset
    customer_address_val = customer_address[0]['cu_address']  # Get value from Queryset

    # Fetch whrm_rate from WhratemasterInfo ONLY for Dedicated customers (id=3)
    whrm_rate_val = 0.0
    print(
        f'DEBUG: customer_businessmodel_val = {customer_businessmodel_val}, type = {type(customer_businessmodel_val)}')
    print(f'DEBUG: customer_id = {customer_id}')
    try:
        # Only fetch whrm_rate for Dedicated customers (customer_businessmodel_val == 3)
        if customer_businessmodel_val == 3:
            print('DEBUG: Inside Dedicated customer block (businessmodel == 3)')
            # First try exact match: customer + business model (Dedicated) + charge type 1
            rate_entry = WhratemasterInfo.objects.filter(
                whrm_customer_name_id=customer_id,
                whrm_businessmodel_id=customer_businessmodel_val,
                whrm_charge_type_id=1,
            ).order_by('-id').first()

            print(f'DEBUG: Exact match query result: {rate_entry}')
            if rate_entry:
                print(f'DEBUG: Found rate entry with rate: {rate_entry.whrm_rate}')

            if not rate_entry:
                print('DEBUG: Exact match not found, trying fallback query')
                # Fallback: customer + charge type 1 (without business model)
                rate_entry = WhratemasterInfo.objects.filter(
                    whrm_customer_name_id=customer_id,
                    whrm_charge_type_id=1,
                ).order_by('-id').first()
                print(f'DEBUG: Fallback query result: {rate_entry}')

            if rate_entry:
                whrm_rate_val = float(rate_entry.whrm_rate)
                print(f'DEBUG: Final whrm_rate_val set to: {whrm_rate_val}')
        else:
            print(f'DEBUG: NOT a Dedicated customer. businessmodel_val={customer_businessmodel_val}')
    except Exception as e:
        print(f'DEBUG: Exception occurred: {e}')
        whrm_rate_val = 0.0
    print('whrm_rate_val', whrm_rate_val)
    data = {
        'customer_businessmodel_val': customer_businessmodel_val,
        'customer_short_name_val': customer_short_name_val,
        'customer_code_val': customer_code_val,
        'customer_GST_val': customer_GST_val,
        'customer_person_val': customer_person_val,
        'customer_contact_val': customer_contact_val,
        'customer_address_val': customer_address_val,
        'whrm_rate_val': whrm_rate_val,
        # 'min_check_in_time': min_check_in_time,
        # 'max_check_out_time': max_check_out_time,
        # 'max_storage_days': max_storage_days,
    }
    return HttpResponse(json.dumps(data, default=custom_serializer))


@login_required(login_url='login_page')
def case_to_case_invoice_list_open(request):
    first_name = request.session.get('first_name')
    case_to_case = str(TrbusinesstypeInfo.objects.get(id=1))
    open_invoice_list = Warehouse_goods_info.objects.filter(wh_voucher_num=None, wh_check_in_out=2, wh_customer_type=1)
    context = {
        'open_invoice_list': open_invoice_list,
        'first_name': first_name,
        'invoice_type': str('Cast-To-Case Customer Open Invoice List'),
    }
    return render(request, "asset_mgt_app/invoice_list_open.html", context)


@login_required(login_url='login_page')
def dedicated_invoice_list_open(request):
    first_name = request.session.get('first_name')
    dedicated = str(TrbusinesstypeInfo.objects.get(id=3))
    open_invoice_list = Warehouse_goods_info.objects.filter(wh_voucher_num=None, wh_check_in_out=2, wh_customer_type=3)
    context = {
        'open_invoice_list': open_invoice_list,
        'first_name': first_name,
        'invoice_type': str('Dedicated Customer  Open Invoice List'),
    }
    return render(request, "asset_mgt_app/invoice_list_open.html", context)


@login_required(login_url='login_page')
def exclusive_invoice_list_open(request):
    first_name = request.session.get('first_name')
    exlcusive = str(TrbusinesstypeInfo.objects.get(id=2))
    open_invoice_list = Warehouse_goods_info.objects.filter(wh_voucher_num=None, wh_check_in_out=1, wh_customer_type=2)
    context = {
        'open_invoice_list': open_invoice_list,
        'first_name': first_name,
        'invoice_type': str('Exclusive Customer Open Invoice List'),
    }
    return render(request, "asset_mgt_app/invoice_list_open.html", context)


@login_required(login_url='login_page')
def invoice_list_query(request):
    first_name = request.session.get('first_name')
    context = {
        'first_name': first_name,
    }
    return render(request, "asset_mgt_app/invoice_list.html", context)


import re


def extract_pincode(address):
    if not address:
        return ""
    match = re.search(r'\b\d{6}\b', address)
    return match.group() if match else ""


def excel_datetime(value):
    """
    Makes datetime safe for Excel.
    - date -> returned as-is
    - aware datetime -> converted to naive
    - naive datetime -> returned as-is
    """
    if not value:
        return ""

    # If it's a DATE (not datetime), Excel accepts it directly
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    # If it's a timezone-aware datetime
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.make_naive(value)

    return value


def extract_state(address):
    if not address:
        return ""
    states = [
        "Tamil Nadu", "Karnataka", "Kerala", "Andhra Pradesh",
        "Telangana", "Maharashtra", "Gujarat", "Delhi",
        "West Bengal", "Uttar Pradesh", "Madhya Pradesh"
    ]
    for state in states:
        if state.lower() in address.lower():
            return state
    if "bangalore" in address.lower():
        return "Karnataka"
    return ""


def extract_pincode(address):
    if not address:
        return ""
    import re
    # Look for 6 consecutive digits (standard Indian pincode)
    match = re.search(r'\b\d{6}\b', address)
    return match.group(0) if match else ""


@login_required(login_url='login_page')
def shipper_invoice_export_excel(request, invoice_id):
    # ✅ invoice = BillingInfo record
    invoice = BilingInfo.objects.select_related(
        "bill_customer_name"
    ).filter(id=invoice_id).first()

    if not invoice:
        return HttpResponse("Invoice not found", status=404)

    # ✅ goods rows linked with this invoice (Fallback to wh_voucher_num if ID link is missing)
    qs = Warehouse_goods_info.objects.select_related(
        "wh_customer_name",
        "wh_branch",
        "wh_dispatch_id"
    ).filter(
        Q(wh_voucher_id=invoice_id) | Q(wh_voucher_num=invoice.bill_invoice_ref)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "WMS_TALLY_IMPORT"

    headers = [
        "DATE", "Vch No", "JOB NO.", "SUNDRY DEBTORS", "GST No", "STATE", "PINCODE",
        "PRIMARY COST CATEGORY", "CUSTOMER",
        "Warehouse Storage Charges", "Warehouse Loading Charges", "Warehouse Unloading Charges",
        "Warehouse Handling Charges", "Crane Handling Charges", "Forklift Handling Charges",
        "Packing Charges", "TOTAL", "CGST Output @ 9%", "SGST Output @ 9%"
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Check if monthly billing (Dedicated ID=3, Exclusive ID=2)
    is_monthly = invoice.bill_customer_type and invoice.bill_customer_type.id in [2, 3]

    # Aggregated single line summary for all billing types (Monthly & Case-to-Case)
    # Use aggregate for weights/pieces (Source of Truth for physical data)
    totals = qs.aggregate(
        total_weight=Sum('wh_goods_weight'),
        total_pieces=Sum('wh_goods_pieces')
    )

    all_jobs = list(qs.values_list('wh_job_no', flat=True).distinct())

    if is_monthly and invoice.bill_invoice_date:
        month_abbr = invoice.bill_invoice_date.strftime('%b').lower()
        year_yy = invoice.bill_invoice_date.strftime('%y')
        customer_prefix = invoice.bill_customer_short_name or (
            invoice.bill_customer_name.cu_nameshort if invoice.bill_customer_name else "CUST")
        # Take first word of customer name e.g. "ROHLIG INDIA" -> "ROHLIG"
        if customer_prefix:
            customer_prefix = customer_prefix.split(' ')[0].split('(')[0].split('-')[0].strip()
        job_no_str = f"{customer_prefix}-{month_abbr}-{year_yy}"
    else:
        job_no_str = ", ".join(all_jobs) if all_jobs else ""

    customer = invoice.bill_customer_name
    address = customer.cu_address if (customer and customer.cu_address) else (invoice.bill_customer_address or "")
    state = extract_state(address)
    # Use only the dedicated pincode field from Customer Master, fallback to address extraction
    pincode = customer.cu_pincode if customer and hasattr(customer, 'cu_pincode') else ""
    if not pincode:
        pincode = extract_pincode(address)

    # Branch/Unit from first record
    first_good = qs.first()
    branch_name = first_good.wh_branch.loc_name if first_good and first_good.wh_branch else ""
    unit_no = str(first_good.wh_unit) if (first_good and first_good.wh_unit) else ""

    # Robust Branch/Unit fallback from Location Master
    lm = LocationmasterInfo.objects.filter(lm_customer_name=customer).select_related("lm_wh_location",
                                                                                     "lm_wh_unit").first()
    if not branch_name and lm and lm.lm_wh_location:
        branch_name = lm.lm_wh_location.loc_name

    if not unit_no and lm and lm.lm_wh_unit:
        unit_no = str(lm.lm_wh_unit)

    branch_name = branch_name.replace("BVM ", "").strip()
    primary_cost_category = str(unit_no) if unit_no else str(branch_name)

    # For Monetary values, match the portal screen logic:
    # 1. Use job sums for activity-based costs (Storage, Loading, Crane, Forklift)
    # 2. Use header fields for manual/fixed costs (Handling, Packing, Fumigation)
    job_totals = qs.aggregate(
        total_storage=Sum('wh_storage_cost_total'),
        total_loading=Sum('wh_total_loading_cost'),
        total_forklift=Sum('wh_forklift_cost'),
        total_crane=Sum('wh_crane_cost'),
        total_fumigation=Sum('wh_fumigation_cost')
    )

    # Prioritize job-level sum if db summation is greater than header overrides
    # (Matches detailed export calculation behavior)
    sum_storage = float(job_totals['total_storage'] or 0)
    header_storage = float(invoice.bill_wh_storage_charges if invoice.bill_wh_storage_charges is not None else 0)
    storage_base = sum_storage if sum_storage > header_storage else header_storage

    fumigation_val = float(invoice.bill_tot_fumigation_charges if (
                invoice.bill_tot_fumigation_charges is not None and invoice.bill_tot_fumigation_charges != 0) else (
                job_totals['total_fumigation'] or 0))

    # Merge storage and fumigation to avoid adding new columns while keeping totals correct
    storage_val = storage_base + fumigation_val

    loading_val = float(invoice.bill_loading_charge if (
                invoice.bill_loading_charge is not None and invoice.bill_loading_charge != 0) else (
                job_totals['total_loading'] or 0))

    # If unloading is 0, default to loading_val as per UI behavior
    unloading_val = float(invoice.bill_unloading_charge if (
                invoice.bill_unloading_charge is not None and invoice.bill_unloading_charge != 0) else loading_val)

    handling_val = float(invoice.bill_handling_charges if (
                invoice.bill_handling_charges is not None and invoice.bill_handling_charges != 0) else 0)
    crane_val = float(invoice.bill_tot_crane_charges if (
                invoice.bill_tot_crane_charges is not None and invoice.bill_tot_crane_charges != 0) else (
                job_totals['total_crane'] or 0))
    forklift_val = float(invoice.bill_tot_forklift_charges if (
                invoice.bill_tot_forklift_charges is not None and invoice.bill_tot_forklift_charges != 0) else (
                job_totals['total_forklift'] or 0))
    packing_val = float(invoice.bill_packing_charges if (
                invoice.bill_packing_charges is not None and invoice.bill_packing_charges != 0) else 0)

    # Calculate pre-gst total components for verification
    calc_pre_gst_total = round(storage_val + loading_val + unloading_val + handling_val +
                               crane_val + forklift_val + packing_val, 2)

    # Match UI Total exactly or re-calculate
    pre_gst_total = float(invoice.bill_total_pre_gst if (
                invoice.bill_total_pre_gst is not None and invoice.bill_total_pre_gst != 0) else calc_pre_gst_total)

    # Use actual taxes from header or re-calculate
    cgst_val = float(
        invoice.bill_cgst if (invoice.bill_cgst is not None and invoice.bill_cgst != 0) else round(pre_gst_total * 0.09,
                                                                                                   2))
    sgst_val = float(
        invoice.bill_sgst if (invoice.bill_sgst is not None and invoice.bill_sgst != 0) else round(pre_gst_total * 0.09,
                                                                                                   2))

    if is_monthly or not all_jobs:
        ws.append([
            invoice.bill_invoice_date.strftime('%d-%m-%Y') if invoice.bill_invoice_date else "",
            invoice.bill_invoice_ref,
            job_no_str,
            invoice.bill_customer_short_name if invoice.bill_customer_short_name else (
                customer.cu_nameshort if customer else ""),
            customer.cu_gst if customer else "",
            state,
            pincode,
            primary_cost_category,
            customer.cu_name if customer else "",
            round(storage_val, 2),
            round(loading_val, 2),
            round(unloading_val, 2),
            round(handling_val, 2),
            round(crane_val, 2),
            round(forklift_val, 2),
            round(packing_val, 2),
            pre_gst_total,
            cgst_val,
            sgst_val,
        ])
    else:
        target_s = round(storage_val, 2)
        target_l = round(loading_val, 2)
        target_u = round(unloading_val, 2)
        target_h = round(handling_val, 2)
        target_c = round(crane_val, 2)
        target_f = round(forklift_val, 2)
        target_p = round(packing_val, 2)
        target_pre = round(pre_gst_total, 2)
        target_cgst = round(cgst_val, 2)
        target_sgst = round(sgst_val, 2)

        j_db_s = float(job_totals['total_storage'] or 0) + float(job_totals['total_fumigation'] or 0)
        j_db_l = float(job_totals['total_loading'] or 0)
        j_db_c = float(job_totals['total_crane'] or 0)
        j_db_f = float(job_totals['total_forklift'] or 0)

        alloc_s = alloc_l = alloc_u = alloc_h = alloc_c = alloc_f = alloc_p = alloc_pre = alloc_cgst = alloc_sgst = 0
        num_jobs = len(all_jobs)

        for i, job_no in enumerate(all_jobs):
            is_last = (i == num_jobs - 1)

            if is_last:
                job_s = round(target_s - alloc_s, 2)
                job_l = round(target_l - alloc_l, 2)
                job_u = round(target_u - alloc_u, 2)
                job_h = round(target_h - alloc_h, 2)
                job_c = round(target_c - alloc_c, 2)
                job_f = round(target_f - alloc_f, 2)
                job_p = round(target_p - alloc_p, 2)
                job_pre = round(target_pre - alloc_pre, 2)
                job_cgst = round(target_cgst - alloc_cgst, 2)
                job_sgst = round(target_sgst - alloc_sgst, 2)
            else:
                job_qs = qs.filter(wh_job_no=job_no)
                j_totals = job_qs.aggregate(
                    s=Sum('wh_storage_cost_total'),
                    fum=Sum('wh_fumigation_cost'),
                    l=Sum('wh_total_loading_cost'),
                    c=Sum('wh_crane_cost'),
                    fl=Sum('wh_forklift_cost')
                )

                db_s = float(j_totals['s'] or 0) + float(j_totals['fum'] or 0)
                db_l = float(j_totals['l'] or 0)
                db_c = float(j_totals['c'] or 0)
                db_f = float(j_totals['fl'] or 0)

                def get_share(db_val, total_db_val, target_val):
                    if total_db_val > 0:
                        return round(target_val * (db_val / total_db_val), 2)
                    return round(target_val / num_jobs, 2)

                job_s = round(target_s / num_jobs, 2)
                job_l = get_share(db_l, j_db_l, target_l)
                job_u = get_share(db_l, j_db_l, target_u)
                job_c = get_share(db_c, j_db_c, target_c)
                job_f = get_share(db_f, j_db_f, target_f)

                job_h = round(target_h / num_jobs, 2)
                job_p = round(target_p / num_jobs, 2)

                job_pre = round(job_s + job_l + job_u + job_h + job_c + job_f + job_p, 2)

                if target_pre > 0:
                    job_cgst = round(target_cgst * (job_pre / target_pre), 2)
                    job_sgst = round(target_sgst * (job_pre / target_pre), 2)
                else:
                    job_cgst = round(target_cgst / num_jobs, 2)
                    job_sgst = round(target_sgst / num_jobs, 2)

            alloc_s += job_s
            alloc_l += job_l
            alloc_u += job_u
            alloc_h += job_h
            alloc_c += job_c
            alloc_f += job_f
            alloc_p += job_p
            alloc_pre += job_pre
            alloc_cgst += job_cgst
            alloc_sgst += job_sgst

            ws.append([
                invoice.bill_invoice_date.strftime('%d-%m-%Y') if invoice.bill_invoice_date else "",
                invoice.bill_invoice_ref,
                job_no,
                invoice.bill_customer_short_name if invoice.bill_customer_short_name else (
                    customer.cu_nameshort if customer else ""),
                customer.cu_gst if customer else "",
                state,
                pincode,
                primary_cost_category,
                customer.cu_name if customer else "",
                job_s,
                job_l,
                job_u,
                job_h,
                job_c,
                job_f,
                job_p,
                job_pre,
                job_cgst,
                job_sgst,
            ])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Filename: TALLY EXCEL-XXXX (last 4 digits of BVM invoice number)
    invoice_suffix = str(invoice.bill_invoice_ref)[-4:] if invoice.bill_invoice_ref else str(invoice_id)
    response["Content-Disposition"] = f"attachment; filename=TALLY EXCEL-{invoice_suffix}.xlsx"

    # Auto-adjust column widths for better visibility
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(response)
    return response


@login_required(login_url='login_page')
def invoice_export_excel(request, invoice_id):
    """ Standard Detailed Excel Export excluding Tally-specific fields """
    invoice = BilingInfo.objects.select_related("bill_customer_name").filter(id=invoice_id).first()
    if not invoice:
        return HttpResponse("Invoice not found", status=404)

    # Fallback to wh_voucher_num if ID link is missing
    qs = Warehouse_goods_info.objects.select_related(
        "wh_customer_name",
        "wh_branch",
        "wh_dispatch_id"
    ).filter(
        Q(wh_voucher_id=invoice_id) | Q(wh_voucher_num=invoice.bill_invoice_ref)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice_Details"

    headers = [
        "SL. NO.", "JOB NO.", "SHIPPER'S NAME", "SHIPPER'S REF. INVOICE NO.",
        "W/H CHRGS VEHICLE TYPE", "SHPT. WT (Kgs)", "SHIPMENT IN DATE", "SHIPMENT OUT DATE",
        "NO. OF DAYS", "PER DAY W/H CHARGES", "Warehouse Storage Charges",
        "NO. PALLETS/ BOXES [PCS]", "RATE PER PALLET/ BOXES", "Warehouse Loading Charges",
        "NO. PALLETS/ BOXES [PCS]", "RATE PER PALLET/ BOXES", "Warehouse UnLoading Charges",
        "Warehouse Handling Charges", "Crane Handling Charges", "Forklift Handling Charges", "Total"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Check if monthly billing (Dedicated ID=3, Exclusive ID=2)
    is_monthly = invoice.bill_customer_type and invoice.bill_customer_type.id in [2, 3]

    if is_monthly:
        # Aggregated single line summary
        # Use job-sum for physical fields (Source of Truth for weight/pieces)
        item_sums = qs.aggregate(
            total_weight=Sum('wh_goods_weight'),
            total_pieces=Sum('wh_goods_pieces')
        )

        # Use job-sum for Activity charges to match portal screen overwriting logic
        job_totals = qs.aggregate(
            total_storage=Sum('wh_storage_cost_total'),
            total_loading=Sum('wh_total_loading_cost'),
            total_forklift=Sum('wh_forklift_cost'),
            total_crane=Sum('wh_crane_cost'),
            total_fumigation=Sum('wh_fumigation_cost')
        )

        all_jobs = list(qs.values_list('wh_job_no', flat=True).distinct())
        job_no_str = ", ".join(all_jobs) if all_jobs else ""

        # Use header fields first (is not None to handle 0.0 correctly), fallback to job sums
        storage_val = float(invoice.bill_wh_storage_charges if invoice.bill_wh_storage_charges is not None else (
                    job_totals['total_storage'] or 0))
        loading_val = float(invoice.bill_loading_charge if invoice.bill_loading_charge is not None else (
                    job_totals['total_loading'] or 0))
        unloading_val = float(invoice.bill_unloading_charge if invoice.bill_unloading_charge is not None else 0)
        crane_val = float(invoice.bill_tot_crane_charges if invoice.bill_tot_crane_charges is not None else (
                    job_totals['total_crane'] or 0))
        forklift_val = float(invoice.bill_tot_forklift_charges if invoice.bill_tot_forklift_charges is not None else (
                    job_totals['total_forklift'] or 0))

        per_day_storage = invoice.bill_per_day_wh_charges or 0
        if not per_day_storage and invoice.bill_no_of_days and storage_val:
            per_day_storage = round(storage_val / invoice.bill_no_of_days, 2)

        # Prioritize wh_consigner as shown in the UI "Shipper Name" column
        shipper_list = list(qs.exclude(wh_consigner__isnull=True).exclude(wh_consigner='')
                            .values_list('wh_consigner', flat=True).distinct())

        if not shipper_list:
            shipper_list = list(qs.exclude(wh_gate_injob_no_id__gatein_shipper__isnull=True)
                                .exclude(wh_gate_injob_no_id__gatein_shipper='')
                                .values_list('wh_gate_injob_no_id__gatein_shipper', flat=True).distinct())

        shipper_str = ", ".join(shipper_list) if shipper_list else (
            invoice.bill_customer_name.cu_nameshort if invoice.bill_customer_name else "")

        # Shipper Invoice Summary
        invoice_list_vals = list(qs.exclude(wh_gate_injob_no_id__gatein_invoice__isnull=True)
                                 .exclude(wh_gate_injob_no_id__gatein_invoice='')
                                 .values_list('wh_gate_injob_no_id__gatein_invoice', flat=True).distinct())

        if not invoice_list_vals:
            invoice_list_vals = list(qs.exclude(wh_goods_invoice__isnull=True).exclude(wh_goods_invoice='')
                                     .values_list('wh_goods_invoice', flat=True).distinct())

        invoice_str = ", ".join(invoice_list_vals) if invoice_list_vals else "Summary"

        ws.append([
            1,
            job_no_str,
            shipper_str,
            invoice_str,
            "",  # Truck Type
            item_sums['total_weight'] or 0,
            excel_datetime(invoice.bill_start_date),
            excel_datetime(invoice.bill_end_date),
            invoice.bill_no_of_days or 0,
            per_day_storage,
            round(storage_val, 2),

            # Loading
            item_sums['total_pieces'] or 0,
            invoice.bill_rate_per_pallet or 0,
            round(loading_val, 2),

            # Unloading
            item_sums['total_pieces'] or 0,
            invoice.bill_rate_per_pallet or 0,
            round(unloading_val, 2),

            # Warehouse Handling/Misc (Includes Handling + Fumigation + Packing)
            round(float(invoice.bill_handling_charges or 0) + float(invoice.bill_tot_fumigation_charges or 0) + float(
                invoice.bill_packing_charges or 0), 2),
            round(crane_val, 2),
            round(forklift_val, 2),
            # Total (Sum of Storage + Loading + Unloading + Handling/Misc + Crane + Forklift)
            round(storage_val + loading_val + unloading_val +
                  (float(invoice.bill_handling_charges or 0) + float(invoice.bill_tot_fumigation_charges or 0) + float(
                      invoice.bill_packing_charges or 0)) +
                  crane_val + forklift_val, 2),
        ])
    else:
        # Pre-calculate storage split logic identically to Case-to-Case tally export
        job_totals = qs.aggregate(total_storage=Sum('wh_storage_cost_total'))
        header_storage = float(invoice.bill_wh_storage_charges if invoice.bill_wh_storage_charges is not None else 0)
        sum_storage = float(job_totals['total_storage'] or 0)

        # Prioritize job-level sum if db summation is greater than header overrides
        total_storage = sum_storage if sum_storage > header_storage else header_storage

        distinct_jobs = list(qs.values_list('wh_job_no', flat=True).distinct())
        num_distinct_jobs = len(distinct_jobs)

        # Calculate equal portion per distinct job exactly matched to standard export
        job_portion = round(total_storage / num_distinct_jobs, 2) if num_distinct_jobs > 0 else 0

        sl = 1
        for job_no in distinct_jobs:
            job_qs = qs.filter(wh_job_no=job_no)

            # Aggregate dynamic values directly to precisely matched distinct job representation
            j_agg = job_qs.aggregate(
                weight=Sum('wh_goods_weight'),
                pieces=Sum('wh_goods_pieces'),
                loading=Sum('wh_total_loading_cost'),
                crane=Sum('wh_crane_cost'),
                forklift=Sum('wh_forklift_cost'),
                min_in=Min('wh_checkin_time'),
                max_out=Max('wh_checkout_time'),
                max_storage_time=Max('wh_storage_time'),
                max_per_day_rate=Max('wh_storage_cost_per_day')
            )

            obj = job_qs.first()  # Pull meta info from the first corresponding goods instance

            # Use original max per-day rate
            row_per_day_rate = j_agg['max_per_day_rate'] or 0

            # Pull unit from job first
            unit_no = str(obj.wh_unit) if obj.wh_unit else ""
            if not unit_no:
                lm = LocationmasterInfo.objects.filter(
                    lm_customer_name=invoice.bill_customer_name
                ).select_related("lm_wh_unit").first()
                if lm and lm.lm_wh_unit:
                    unit_no = str(lm.lm_wh_unit)

            # Robust Shipper Name lookup
            shipper_name = ""
            if obj.wh_gate_injob_no_id and obj.wh_gate_injob_no_id.gatein_shipper:
                shipper_name = obj.wh_gate_injob_no_id.gatein_shipper
            else:
                shipper_name = obj.wh_consigner or ""

            if not shipper_name and invoice.bill_customer_name:
                shipper_name = invoice.bill_customer_name.cu_nameshort or ""

            # Aggregate Handling row format specifically applied to Row 1
            row_handling = round((invoice.bill_handling_charges or 0) + (invoice.bill_tot_fumigation_charges or 0) + (
                        invoice.bill_packing_charges or 0), 2) if sl == 1 else 0

            ws.append([
                sl,
                job_no,
                shipper_name,
                obj.wh_gate_injob_no_id.gatein_invoice if obj.wh_gate_injob_no_id and obj.wh_gate_injob_no_id.gatein_invoice else (
                            obj.wh_goods_invoice or "Summary"),

                # W/H CHRGS VEHICLE TYPE: Show actual type instead of status
                str(obj.wh_gate_injob_no_id.gatein_truck_type) if obj.wh_dispatch_id and obj.wh_dispatch_id.dispatch_billing_truck_type and "In" in str(
                    obj.wh_dispatch_id.dispatch_billing_truck_type) else (
                    str(obj.wh_dispatch_id.dispatch_truck_type) if obj.wh_dispatch_id and obj.wh_dispatch_id.dispatch_truck_type else str(
                        obj.wh_truck_type or "")),

                j_agg['weight'] or 0,
                j_agg['min_in'].strftime('%m-%d-%Y') if j_agg['min_in'] else "",  # Format as string to avoid ####
                j_agg['max_out'].strftime('%m-%d-%Y') if j_agg['max_out'] else "",  # Format as string to avoid ####
                j_agg['max_storage_time'] or 0,
                row_per_day_rate,
                job_portion,

                # Loading
                j_agg['pieces'] or 0,
                obj.wh_loading_charge_unit or 0,
                j_agg['loading'] or 0,

                # Unloading (defaults to wh_total_loading_cost as per UI behavior)
                j_agg['pieces'] or 0,
                obj.wh_loading_charge_unit or 0,
                j_agg['loading'] or 0,

                # Warehouse Handling/Misc (First row only)
                row_handling,
                j_agg['crane'] or 0,
                j_agg['forklift'] or 0,

                # Row Total (Sum of Storage + Loading + Unloading + Handling/Misc + Crane + Forklift)
                round(job_portion + (j_agg['loading'] or 0) + (j_agg['loading'] or 0) +
                      row_handling +
                      (j_agg['crane'] or 0) + (j_agg['forklift'] or 0), 2),
            ])
            sl += 1

    # Auto-adjust column widths for better visibility (handle None cell values)
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # Filename: INVOICE-XXXX (last 4 digits of BVM invoice number)
    invoice_suffix = str(invoice.bill_invoice_ref)[-4:] if invoice.bill_invoice_ref else str(invoice_id)
    response["Content-Disposition"] = f"attachment; filename=INVOICE-{invoice_suffix}.xlsx"
    wb.save(response)
    return response
    
