from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.db.models import Q
from django.contrib import messages

from ..sub_forms.market_bill_form import MarketBillForm
from ..sub_models.market_bill_mod import MarketBillInfo
from ..sub_models.vehiclemaster_mod import VehiclemasterInfo
from ..sub_models.tripdetail_mod import TripdetailInfo
from ..sub_models.vehicle_allotment_mod import Vehicle_allotmentInfo
from ..sub_models.haltingcharges_mod import Haltingcharges
from ..sub_models.vendorratemaster1_mod import VendorratemasterInfo1

from django.core.files.storage import default_storage
import os

# ==================================================
# ADD MARKET BILL
# ==================================================
@login_required(login_url='login_page')
def market_bill_add(request):
    if request.method == "POST":
        form = MarketBillForm(request.POST, request.FILES)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.mb_created_by = request.user
            obj.save()

            # Save per-trip costs and halting data to TripdetailInfo
            selected_trips = request.POST.get('mb_selected_trips', '')
            if selected_trips:
                trip_ids = [tid for tid in selected_trips.split(',') if tid.strip()]
                for tid in trip_ids:
                    l_cost = request.POST.get(f'loading_cost_{tid}', 0)
                    u_cost = request.POST.get(f'unloading_cost_{tid}', 0)
                    p_cost = request.POST.get(f'parking_cost_{tid}', 0)
                    h_days = request.POST.get(f'halting_days_{tid}', 0)
                    h_cost = request.POST.get(f'halting_cost_{tid}', 0)
                    t_cost = request.POST.get(f'trip_cost_{tid}', 0)

                    TripdetailInfo.objects.filter(id=tid).update(
                        tc_loadingcost=float(l_cost) if l_cost else 0.0,
                        tc_unloadingcost=float(u_cost) if u_cost else 0.0,
                        tc_parkingcost=float(p_cost) if p_cost else 0.0,
                        tc_no_of_days_halting=int(h_days) if h_days else 0,
                        tc_haltingcost=float(h_cost) if h_cost else 0.0
                    )

                    # Update Buying Price in Allotment instead of Revenue in Trip
                    trip_obj = TripdetailInfo.objects.get(id=tid)
                    Vehicle_allotmentInfo.objects.filter(
                        Q(va_enquirynumber=trip_obj.tr_enquirynumber),
                        Q(va_vehiclenumber__vm_registrationnumber__iexact=trip_obj.tr_vehiclenumber) | 
                        Q(va_vehiclenumber_mkt__iexact=trip_obj.tr_vehiclenumber)
                    ).update(va_specialbuy=float(t_cost) if t_cost else 0.0)

            messages.success(request, "Market Bill saved successfully.")
            return redirect('market_bill_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = MarketBillForm()

    return render(
        request,
        "asset_mgt_app/market_bill.html",
        {
            "form": form,
        }
    )


# ==================================================
# LIST MARKET BILLS
# ==================================================
@login_required(login_url='login_page')
def market_bill_list(request):
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    # Base queryset
    bills_qs = MarketBillInfo.objects.all().order_by('-mb_created_at')

    # If no date filters provided, return bills as usual
    if not from_date and not to_date:
        bills = bills_qs
    else:
        # The user wants to filter Market Bills by the dates of the trips INSIDE the bill
        # Collect Trip IDs that fall inside the provided date range
        trip_filter = TripdetailInfo.objects.all()
        # Use departed date when available, otherwise created_at as a fallback
        if from_date:
            trip_filter = trip_filter.filter(
                Q(tr_departeddate__gte=from_date) | Q(tr_departeddate__isnull=True, tr_created_at__gte=from_date)
            )
        if to_date:
            trip_filter = trip_filter.filter(
                Q(tr_departeddate__lte=to_date) | Q(tr_departeddate__isnull=True, tr_created_at__lte=to_date)
            )

        trip_ids = set(trip_filter.values_list('id', flat=True))

        # Evaluate bills and keep those that reference at least one matching trip id
        bills = []
        for bill in bills_qs:
            if not bill.mb_selected_trips:
                continue
            try:
                bill_trip_ids = [int(tid.strip()) for tid in bill.mb_selected_trips.split(',') if tid.strip()]
            except ValueError:
                # If non-integer values are present, skip this bill for safety
                continue
            # If any trip id in the bill matches the filtered trip ids, include the bill
            if any(tid in trip_ids for tid in bill_trip_ids):
                bills.append(bill)

    return render(
        request,
        "asset_mgt_app/market_bill_list.html",
        {
            "bills": bills,
            "from_date": from_date,
            "to_date": to_date,
        }
    )


# ==================================================
# EDIT MARKET BILL
# ==================================================
@login_required(login_url='login_page')
def market_bill_edit(request, id):
    record = get_object_or_404(MarketBillInfo, id=id)

    if request.method == "POST":
        form = MarketBillForm(request.POST, request.FILES, instance=record)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.mb_updated_by = request.user
            obj.save()

            # Save per-trip costs and halting data to TripdetailInfo
            selected_trips = request.POST.get('mb_selected_trips', '')
            if selected_trips:
                trip_ids = [tid for tid in selected_trips.split(',') if tid.strip()]
                for tid in trip_ids:
                    l_cost = request.POST.get(f'loading_cost_{tid}', 0)
                    u_cost = request.POST.get(f'unloading_cost_{tid}', 0)
                    p_cost = request.POST.get(f'parking_cost_{tid}', 0)
                    h_days = request.POST.get(f'halting_days_{tid}', 0)
                    h_cost = request.POST.get(f'halting_cost_{tid}', 0)
                    t_cost = request.POST.get(f'trip_cost_{tid}', 0)

                    TripdetailInfo.objects.filter(id=tid).update(
                        tc_loadingcost=float(l_cost) if l_cost else 0.0,
                        tc_unloadingcost=float(u_cost) if u_cost else 0.0,
                        tc_parkingcost=float(p_cost) if p_cost else 0.0,
                        tc_no_of_days_halting=int(h_days) if h_days else 0,
                        tc_haltingcost=float(h_cost) if h_cost else 0.0
                    )

                    # Update Buying Price in Allotment instead of Revenue in Trip
                    trip_obj = TripdetailInfo.objects.get(id=tid)
                    Vehicle_allotmentInfo.objects.filter(
                        Q(va_enquirynumber=trip_obj.tr_enquirynumber),
                        Q(va_vehiclenumber__vm_registrationnumber__iexact=trip_obj.tr_vehiclenumber) | 
                        Q(va_vehiclenumber_mkt__iexact=trip_obj.tr_vehiclenumber)
                    ).update(va_specialbuy=float(t_cost) if t_cost else 0.0)

            messages.success(request, "Market Bill updated successfully.")
            return redirect('market_bill_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = MarketBillForm(instance=record)

    # Fetch selected trips data for the edit page table
    selected_trips_data = []
    if record.mb_selected_trips:
        trip_ids = [int(tid) for tid in record.mb_selected_trips.split(',') if tid.strip()]        # Include trips in 'Trip Settled' (id=7) or 'Ready for Invoice' (id=9) financial status
        eligible_status_ids = [7, 9]
        selected_trips = TripdetailInfo.objects.filter(id__in=trip_ids, tc_financestatus_id__in=eligible_status_ids).select_related('tr_enquirynumber', 'tr_consignmentnumber')

        for trip in selected_trips:
            from_location = ''
            to_location = ''
            if trip.tr_enquirynumber:
                if trip.tr_enquirynumber.en_fromlocaion:
                    from_location = str(trip.tr_enquirynumber.en_fromlocaion)
                if trip.tr_enquirynumber.en_tolocation:
                    to_location = str(trip.tr_enquirynumber.en_tolocation)

            trip_date = ''
            if trip.tr_departeddate:
                trip_date = trip.tr_departeddate.strftime('%d-%m-%Y')
            elif trip.tr_created_at:
                trip_date = trip.tr_created_at.strftime('%d-%m-%Y')

            # Fetch vehicle type
            # Robust fetch for vehicle type
            v_master = VehiclemasterInfo.objects.filter(vm_registrationnumber__iexact=trip.tr_vehiclenumber).first()
            if v_master and v_master.vm_vehicletype:
                vehicle_type = str(v_master.vm_vehicletype)
            elif trip.tr_vehicletype:
                vehicle_type = str(trip.tr_vehicletype)
            else:
                vehicle_type = ''

            # Determine consignment number (cnote) and trip number separately to avoid mixing
            consignment_number = ''
            trip_no = ''
            if trip.tr_consignmentnumber and getattr(trip.tr_consignmentnumber, 'co_consignmentnumber', None):
                consignment_number = trip.tr_consignmentnumber.co_consignmentnumber
            if trip.tr_tripnumber:
                trip_no = trip.tr_tripnumber

            # Determine customer name if available
            customer_name = ''
            if trip.tr_enquirynumber and getattr(trip.tr_enquirynumber, 'en_customername', None):
                customer_name = str(trip.tr_enquirynumber.en_customername)

            # Fetch standard and special costs from allotment
            standard_cost = 0
            special_cost = 0
            allotment = Vehicle_allotmentInfo.objects.filter(
                Q(va_enquirynumber=trip.tr_enquirynumber),
                Q(va_vehiclenumber__vm_registrationnumber__iexact=trip.tr_vehiclenumber) | Q(va_vehiclenumber_mkt__iexact=trip.tr_vehiclenumber)
            ).first()
            if allotment:
                standard_cost = float(allotment.va_standardbuy or 0)
                special_cost = float(allotment.va_specialbuy or 0)

                # Refinement: Try to fetch the rate from VendorratemasterInfo1 using Cnote details if available
                if trip.tr_consignmentnumber and allotment.va_vendor:
                    cnote = trip.tr_consignmentnumber
                    v_type_id = trip.tr_vehicletype_id or trip.tr_vehicletype_placed_id
                    if v_type_id:
                        # Use Cnote locations if available, otherwise fallback to Enquiry locations
                        from_loc = cnote.co_fromlocaion or trip.tr_enquirynumber.en_fromlocaion
                        to_loc = cnote.co_tolocation or trip.tr_enquirynumber.en_tolocation

                        rate_obj = VendorratemasterInfo1.objects.filter(
                            vr1_vendor=allotment.va_vendor,
                            vr1_fromlocation=from_loc,
                            vr1_tolocation=to_loc,
                            vr1_vehicletype_id=v_type_id
                        ).first()
                        if rate_obj:
                            special_cost = float(rate_obj.vr1_rate)
                            # Always update standard_cost to match master rate if master rate is found
                            # This prevents incorrect red highlighting when the master rate is the intended standard
                            standard_cost = special_cost

            # Determine current halting rate from Master Data
            halting_rate = 0.0
            if trip.tr_enquirynumber:
                enquiry = trip.tr_enquirynumber
                try:
                    halting_obj = Haltingcharges.objects.filter(
                        hc_Customer_name=enquiry.en_customername,
                        hc_trip_type=enquiry.en_trip_type
                    ).first()
                    if halting_obj:
                        halting_rate = float(halting_obj.hc_charges)
                except:
                    pass
            
            # Fallback if not in master: use current average ONLY if it seems valid
            if halting_rate == 0.0:
                h_days = int(trip.tc_no_of_days_halting or 0)
                h_cost = float(trip.tc_haltingcost or 0)
                if h_days > 0:
                    halting_rate = h_cost / h_days
                else:
                    halting_rate = h_cost

            # Get mail attachment URL from JSON field
            mail_attachment_url = ''
            if record.mb_trip_mail_attachments and str(trip.id) in record.mb_trip_mail_attachments:
                mail_attachment_url = record.mb_trip_mail_attachments[str(trip.id)]

            selected_trips_data.append({
                'id': trip.id,
                'consignment_number': consignment_number,
                'trip_number': trip_no,
                'display_cnote': consignment_number or '',
                'vehicle_number': trip.tr_vehiclenumber or '',
                'vehicle_type': vehicle_type,
                'customer': customer_name,
                'from_location': from_location,
                'to_location': to_location,
                'trip_date': trip_date,
                'standard_cost': standard_cost,
                'special_cost': special_cost,
                'loading_cost': float(trip.tc_loadingcost or 0),
                'unloading_cost': float(trip.tc_unloadingcost or 0),
                'parking_cost': float(trip.tc_parkingcost or 0),
                'halting_days': int(trip.tc_no_of_days_halting or 0),
                'halting_cost': float(trip.tc_haltingcost or 0),
                'halting_rate': float(halting_rate),
                'mail_attachment_url': mail_attachment_url,
            })

    return render(
        request,
        "asset_mgt_app/market_bill.html",
        {
            "form": form,
            "record": record,
            "selected_trips_data": selected_trips_data,
        }
    )


# ==================================================
# DELETE MARKET BILL
# ==================================================
@login_required(login_url='login_page')
def market_bill_delete(request, id):
    record = get_object_or_404(MarketBillInfo, id=id)

    if request.method == "POST":
        record.delete()
        messages.success(request, "Market Bill deleted successfully.")
        return redirect('market_bill_list')

    return redirect('market_bill_list')


import os
from django.core.files.storage import default_storage

# ==================================================
# AJAX: GET TRIPS BY VENDOR
# ==================================================
@login_required(login_url='login_page')
def get_trips_by_vendor(request):
    vendor_id = request.GET.get('vendor_id')

    if not vendor_id:
        return JsonResponse({'trips': []})

    # Include trips in 'Trip Settled' (id=7) or 'Ready for Invoice' (id=9) financial status
    eligible_status_ids = [7, 9]
    market_ownership_id = 3

    # Option 1: Get vehicles from master (if any are assigned to this vendor)
    vendor_master_vehicles = VehiclemasterInfo.objects.filter(
        vm_vendor_id=vendor_id,
        vm_ownership_id=market_ownership_id
    ).values_list('vm_registrationnumber', flat=True)

    # Option 2: Get enquiries and vehicles allotted to this vendor
    allotted_enquiries = Vehicle_allotmentInfo.objects.filter(
        va_vendor_id=vendor_id,
        va_vehiclesource_id=market_ownership_id
    ).select_related('va_vehiclenumber')

    allotment_filters = Q()
    has_allotments = False
    for allotment in allotted_enquiries:
        reg_no = allotment.va_vehiclenumber.vm_registrationnumber if allotment.va_vehiclenumber else allotment.va_vehiclenumber_mkt
        if reg_no:
            allotment_filters |= Q(tr_enquirynumber_id=allotment.va_enquirynumber_id, tr_vehiclenumber__iexact=reg_no)
            has_allotments = True

    if not has_allotments:
        allotment_filters = Q(pk__in=[])

    # Get all already billed trip IDs from all MarketBillInfo records
    billed_trip_ids = set()
    all_bills = MarketBillInfo.objects.exclude(mb_selected_trips__isnull=True).exclude(mb_selected_trips='')
    for bill in all_bills:
        # Normalize to ints for robust filtering
        try:
            ids = [int(tid.strip()) for tid in bill.mb_selected_trips.split(',') if tid.strip()]
            billed_trip_ids.update(ids)
        except ValueError:
            # If any non-integer values slip in, fallback to string-based set
            billed_trip_ids.update([tid.strip() for tid in bill.mb_selected_trips.split(',') if tid.strip()])

    # Filter trips for ANY vehicle or allotment of this vendor that are not billed
    # Show trips in 'Trip Settled' (ID 7) or 'Ready for Invoice' (ID 9) financial status
    # Plus double check that the trip record itself is marked as 'Market' source
    trips = TripdetailInfo.objects.filter(
        (Q(tr_vehiclenumber__in=list(vendor_master_vehicles)) | allotment_filters),
        Q(tr_departeddate__gte='2026-05-01') | Q(tr_departeddate__isnull=True, tr_created_at__gte='2026-05-01 00:00:00'),
        tc_financestatus_id__in=eligible_status_ids,
        tr_vehiclesource_id=market_ownership_id
    )
    # Exclude already billed trip ids (if any)
    if billed_trip_ids:
        trips = trips.exclude(id__in=list(billed_trip_ids))
    trips = trips.select_related('tr_enquirynumber', 'tr_consignmentnumber')

    trip_list = []
    for trip in trips:
        # Get from/to locations from enquiry
        from_location = ''
        to_location = ''
        if trip.tr_enquirynumber:
            if trip.tr_enquirynumber.en_fromlocaion:
                from_location = str(trip.tr_enquirynumber.en_fromlocaion)
            if trip.tr_enquirynumber.en_tolocation:
                to_location = str(trip.tr_enquirynumber.en_tolocation)

        # Get trip date
        trip_date = ''
        if trip.tr_departeddate:
            trip_date = trip.tr_departeddate.strftime('%d-%m-%Y')
        elif trip.tr_created_at:
            trip_date = trip.tr_created_at.strftime('%d-%m-%Y')

        # --- Cost Logic ---
        # Fetch standard and special costs from allotment
        standard_cost = 0
        special_cost = 0
        allotment = Vehicle_allotmentInfo.objects.filter(
            Q(va_enquirynumber=trip.tr_enquirynumber),
            Q(va_vehiclenumber__vm_registrationnumber__iexact=trip.tr_vehiclenumber) | Q(va_vehiclenumber_mkt__iexact=trip.tr_vehiclenumber)
        ).first()
        if allotment:
            standard_cost = float(allotment.va_standardbuy or 0)
            special_cost = float(allotment.va_specialbuy or 0)

            # Refinement: Try to fetch the rate from VendorratemasterInfo1 using Cnote details if available
            if trip.tr_consignmentnumber and allotment.va_vendor:
                cnote = trip.tr_consignmentnumber
                v_type_id = trip.tr_vehicletype_id or trip.tr_vehicletype_placed_id
                if v_type_id:
                    # Use Cnote locations if available, otherwise fallback to Enquiry locations
                    from_loc = cnote.co_fromlocaion or trip.tr_enquirynumber.en_fromlocaion
                    to_loc = cnote.co_tolocation or trip.tr_enquirynumber.en_tolocation

                    rate_obj = VendorratemasterInfo1.objects.filter(
                        vr1_vendor=allotment.va_vendor,
                        vr1_fromlocation=from_loc,
                        vr1_tolocation=to_loc,
                        vr1_vehicletype_id=v_type_id
                    ).first()
                    if rate_obj:
                        special_cost = float(rate_obj.vr1_rate)
                        # Always update standard_cost to match master rate if master rate is found
                        # This prevents incorrect red highlighting when the master rate is the intended standard
                        standard_cost = special_cost

        # --- Halting Cost Logic ---
        # Fetch halting rate based on customer and trip type
        halting_days = int(trip.tc_no_of_days_halting or 0)
        halting_rate = 0
        if trip.tr_enquirynumber:
            enquiry = trip.tr_enquirynumber
            try:
                halting_obj = Haltingcharges.objects.filter(
                    hc_Customer_name=enquiry.en_customername,
                    hc_trip_type=enquiry.en_trip_type
                ).first()
                if halting_obj:
                    halting_rate = halting_obj.hc_charges
            except:
                pass
        
        halting_cost = halting_rate * halting_days

        # Fetch vehicle type for this trip
        vehicle_type = ''
        if trip.tr_vehicletype:
             vehicle_type = str(trip.tr_vehicletype)
        else:
            vehicle = VehiclemasterInfo.objects.filter(vm_registrationnumber=trip.tr_vehiclenumber).first()
            vehicle_type = str(vehicle.vm_vehicletype) if vehicle and vehicle.vm_vehicletype else ''

        # Determine consignment and trip numbers separately
        consignment_number = ''
        trip_no = ''
        if trip.tr_consignmentnumber and getattr(trip.tr_consignmentnumber, 'co_consignmentnumber', None):
            consignment_number = trip.tr_consignmentnumber.co_consignmentnumber
        if trip.tr_tripnumber:
            trip_no = trip.tr_tripnumber

        # Determine customer
        customer_name = ''
        if trip.tr_enquirynumber and getattr(trip.tr_enquirynumber, 'en_customername', None):
            customer_name = str(trip.tr_enquirynumber.en_customername)

        trip_list.append({
            'id': trip.id,
            'consignment_number': consignment_number,
            'trip_number': trip_no,
            # Only provide consignment in the Cnote column; do not show trip numbers
            'trip_number_display': consignment_number or '',
            'vehicle_number': trip.tr_vehiclenumber or '',
            'vehicle_type': vehicle_type,
            'customer': customer_name,
            'from_location': from_location,
            'to_location': to_location,
            'trip_date': trip_date,
            'standard_cost': float(standard_cost),
            'special_cost': float(special_cost),
            'loading_cost': float(trip.tc_loadingcost or 0),
            'unloading_cost': float(trip.tc_unloadingcost or 0),
            'parking_cost': float(trip.tc_parkingcost or 0),
            'halting_days': halting_days,
            'halting_cost': float(halting_cost),
            'halting_rate': float(halting_rate),
            'mail_attachment_url': '', # New trips won't have attachments yet
        })

    return JsonResponse({'trips': trip_list})

# ==================================================
# UPLOAD BILL ATTACHMENT
# ==================================================
@login_required(login_url='login_page')
def market_bill_upload(request, id):
    record = get_object_or_404(MarketBillInfo, id=id)

    if request.method == "POST" and request.FILES.get('mb_attachment'):
        record.mb_attachment = request.FILES['mb_attachment']
        record.save()
        messages.success(request, "Attachment uploaded successfully")
    else:
        messages.error(request, "Failed to upload attachment. Please select a file.")

    return redirect('market_bill_list')

# ==================================================
# UPLOAD MAIL ATTACHMENT
# ==================================================
@login_required(login_url='login_page')
def market_mail_upload(request, id, trip_id):
    record = get_object_or_404(MarketBillInfo, id=id)

    if request.method == "POST" and request.FILES.get('mb_mail_attachment'):
        uploaded_file = request.FILES['mb_mail_attachment']
        
        # Determine path: MarketMailAttachments/bill_id/trip_id/filename
        file_name = uploaded_file.name
        safe_file_name = "".join([c for c in file_name if c.isalnum() or c in ('.', '_')]).strip()
        relative_path = os.path.join('MarketMailAttachments', str(record.id), str(trip_id), safe_file_name)
        
        # Save file using default_storage
        path = default_storage.save(relative_path, uploaded_file)
        file_url = default_storage.url(path)

        # Update JSON field
        if not record.mb_trip_mail_attachments:
            record.mb_trip_mail_attachments = {}
        
        record.mb_trip_mail_attachments[str(trip_id)] = file_url
        record.save()
        
        messages.success(request, "Attachment uploaded successfully")
    else:
        messages.error(request, "Failed to upload mail attachment. Please select a file.")
    return redirect('market_bill_edit', id=id)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required(login_url='login_page')
def market_bill_export_tally(request):
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    bills = MarketBillInfo.objects.all().order_by('mb_bill_date')
    if from_date:
        bills = bills.filter(mb_bill_date__gte=from_date)
    if to_date:
        bills = bills.filter(mb_bill_date__lte=to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Export"

    headers = [
        "VOUCHER NUMBER", "DATE", "REF NO.", "SUNDRY CREDITORS", "TOTAL AMT",
        "EXPENSES LEDGER", "AMOUNT", "Primary Cost Category", "Job No", 
        "VEH. NO.", "Customer", "TDS LEDGER", "TDS AMOUNT", "NARRATION"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="B2FFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    expense_fill = PatternFill("solid", fgColor="FFE5CC")
    
    for bill in bills:
        trip_ids = []
        if bill.mb_selected_trips:
            try:
                trip_ids = [int(tid) for tid in bill.mb_selected_trips.split(',') if tid.strip()]
            except ValueError:
                pass
        
        trips = TripdetailInfo.objects.filter(id__in=trip_ids).select_related('tr_enquirynumber')
        
        # Calculate Financial Year and Month for Voucher Number
        if bill.mb_bill_date:
            year = bill.mb_bill_date.year
            month = bill.mb_bill_date.month
            if month >= 4:
                fy_str = f"{str(year)[-2:]}-{str(year+1)[-2:]}"
            else:
                fy_str = f"{str(year-1)[-2:]}-{str(year)[-2:]}"
            month_str = f"{month:02d}"
        else:
            fy_str = "00-00"
            month_str = "00"
            
        voucher_number = f"MAA_MKT_{fy_str}_{month_str}_{bill.id:03d}"
        bill_date = bill.mb_bill_date.strftime("%d-%m-%Y") if bill.mb_bill_date else ""
        ref_no = bill.mb_bill_no or ""
        vendor_name = bill.mb_vendor.vend_name if bill.mb_vendor else ""
        total_amt = bill.mb_payable_amount if bill.mb_payable_amount else bill.mb_total_cost
        
        tds_amount = bill.mb_tds_amount or 0.0
        # Set TDS ledger text based on TDS type selection on the bill
        tds_ledger = ""
        try:
            tds_type = (bill.mb_tds_type or '').strip()
            if tds_amount > 0:
                if tds_type == 'Company':
                    tds_ledger = "TDS Payable 194C (Company)"
                else:
                    # Default/Non company
                    tds_ledger = "TDS Payable 194C (Non Company)"
        except Exception:
            tds_ledger = "TDS Payable 194C (Non Company)" if tds_amount > 0 else ""
        
        month_str = bill.mb_bill_date.strftime('%b%y') if bill.mb_bill_date else ""
        rec_date_str = bill.mb_created_at.strftime('%d-%b-%y') if bill.mb_created_at else ""
        narration = f"Being oncall Vehicle hire charges for the month of {month_str} (Bill Received on {rec_date_str})"
        
        is_first_row = True
        
        if not trips:
            row = [
                voucher_number, bill_date, ref_no, vendor_name, total_amt,
                "Transportation", total_amt, "Maa-Mkt", "",
                # Per requirement: do not include actual vehicle number in export; show 'mkt' instead
                "mkt", "", tds_ledger, tds_amount if tds_amount else "", narration
            ]
            ws.append(row)
            continue
            
        for trip in trips:
            # Use ONLY the Consignment number (Cnote) as Job No for Tally export — no fallbacks
            job_no = ""
            try:
                if getattr(trip, 'tr_consignmentnumber', None) and getattr(trip.tr_consignmentnumber, 'co_consignmentnumber', None):
                    job_no = trip.tr_consignmentnumber.co_consignmentnumber
                else:
                    job_no = ""
            except Exception:
                job_no = ""
            customer_name = ""
            if trip.tr_enquirynumber and trip.tr_enquirynumber.en_customername:
                customer_name = trip.tr_enquirynumber.en_customername.customer_name if hasattr(trip.tr_enquirynumber.en_customername, 'customer_name') else str(trip.tr_enquirynumber.en_customername)
            
            transport_cost = 0.0
            allotment = Vehicle_allotmentInfo.objects.filter(
                Q(va_enquirynumber=trip.tr_enquirynumber),
                Q(va_vehiclenumber__vm_registrationnumber__iexact=trip.tr_vehiclenumber) | Q(va_vehiclenumber_mkt__iexact=trip.tr_vehiclenumber)
            ).first()
            if allotment and allotment.va_specialbuy:
                transport_cost = float(allotment.va_specialbuy)
                
            expenses = []
            if transport_cost > 0: expenses.append(("Transportation", transport_cost))
            if trip.tc_loadingcost and float(trip.tc_loadingcost) > 0: expenses.append(("Loading", float(trip.tc_loadingcost)))
            if trip.tc_unloadingcost and float(trip.tc_unloadingcost) > 0: expenses.append(("Unloading", float(trip.tc_unloadingcost)))
            if trip.tc_parkingcost and float(trip.tc_parkingcost) > 0: expenses.append(("Parking", float(trip.tc_parkingcost)))
            if trip.tc_haltingcost and float(trip.tc_haltingcost) > 0: expenses.append(("Halting", float(trip.tc_haltingcost)))
            
            for exp_name, amt in expenses:
                row = [
                    voucher_number, bill_date, ref_no,
                    vendor_name if is_first_row else "",
                    total_amt if is_first_row else "",
                    exp_name, amt,
                    "Maa-Mkt", job_no,
                    # Replace actual vehicle number with constant 'mkt' per spec
                    "mkt", customer_name,
                    tds_ledger if is_first_row else "",
                    tds_amount if is_first_row and tds_amount > 0 else "",
                    narration
                ]
                ws.append(row)
                ws.cell(row=ws.max_row, column=6).fill = expense_fill
                is_first_row = False

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Market_Bill_Tally_Export.xlsx"'
    wb.save(response)
    return response
