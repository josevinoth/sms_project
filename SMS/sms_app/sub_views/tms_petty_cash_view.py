from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from datetime import datetime
from ..sub_models.tms_petty_cash_mod import TMSPettyCashInfo
from ..sub_models.tripdetail_mod import TripdetailInfo
from ..sub_models.credit_ledger_mod import CreditLedgerInfo
from ..models import Business_Sol_info, Location_info
from ..sub_forms.tms_petty_cash_form import TMSPettyCashForm
from .general_utils import get_financial_year, get_session_branch_id

def generate_tms_petty_cash_number(model_class, field_name, branch_obj=None):
    """
    Generates [BranchCode]_T_26-27_yy_xxx
    """
    fy = get_financial_year() # returns e.g. "24-25"
    today = datetime.now()
    month_str = today.strftime("%m")
    
    branch_prefix = "M"
    if branch_obj and branch_obj.loc_name:
        loc_name_lower = branch_obj.loc_name.lower()
        if "blr" in loc_name_lower or "bengaluru" in loc_name_lower or "bangalore" in loc_name_lower:
            branch_prefix = "B"
            
    prefix = f"{branch_prefix}_T_{fy}_{month_str}_"
    
    # fetch latest for this prefix
    latest_obj = model_class.objects.filter(**{f"{field_name}__startswith": prefix}).order_by('-id').first()
    
    if latest_obj:
        latest_num_str = getattr(latest_obj, field_name)
        # extract xxx
        try:
            seq = int(latest_num_str.split('_')[-1])
            new_seq = seq + 1
        except ValueError:
            new_seq = 1
    else:
        new_seq = 1
        
    return f"{prefix}{str(new_seq).zfill(3)}"

def tms_petty_cash_add(request, tpc_id=0):
    initial_data = {}
    if tpc_id == 0:
        # Default Business
        bvm_trans = Business_Sol_info.objects.filter(bvm_business__icontains='bvm trans solutions').first()
        if bvm_trans:
            initial_data['tpc_business'] = bvm_trans.id
            
        # Default Expense Category to Cash Expense
        from ..models import ExpenseCategoryInfo
        cash_cat = ExpenseCategoryInfo.objects.filter(exp_category_name__icontains='Cash').first()
        if cash_cat:
            initial_data['tpc_category'] = cash_cat.id

        # Default Branch based on session
        branch_id = get_session_branch_id(request)
        if branch_id:
            initial_data['tpc_branch'] = branch_id
            
            # Default Credit Ledger based on branch name
            branch_obj = Location_info.objects.filter(id=branch_id).first()
            if branch_obj:
                branch_code = branch_obj.loc_name.split()[-1] # e.g. "MAA" or "BLR"
                ledger = CreditLedgerInfo.objects.filter(
                    ledger_name__icontains='Trans Petty Cash'
                ).filter(
                    ledger_name__icontains=branch_code
                ).exclude(
                    ledger_name__icontains='Admin'
                ).first()
                if not ledger:
                    ledger = CreditLedgerInfo.objects.filter(
                        ledger_name__icontains='Trans'
                    ).filter(
                        ledger_name__icontains=branch_code
                    ).first()
                if ledger:
                    initial_data['tpc_credit_ledger'] = ledger.id
                    
        form = TMSPettyCashForm(initial=initial_data, request=request)
    else:
        tpc = get_object_or_404(TMSPettyCashInfo, pk=tpc_id)
        form = TMSPettyCashForm(instance=tpc, request=request)
        
    if request.method == 'POST':
        post_data = request.POST.copy()
        
        # If vehicle number or driver name are raw text from Select2 (because they were missing from masters), 
        # create them dynamically and update POST data with the new ForeignKey ID.
        if post_data.get('tpc_vehicle_number') and not str(post_data['tpc_vehicle_number']).isdigit():
            from ..models import VehiclemasterInfo
            veh, _ = VehiclemasterInfo.objects.get_or_create(
                vm_registrationnumber=str(post_data['tpc_vehicle_number']).strip(),
                defaults={
                    'vm_numberoftyres': None, 
                    'vm_primarydrivermob': None, 
                    'vm_secondarydrivermob': None
                }
            )
            post_data['tpc_vehicle_number'] = str(veh.id)
            
        if post_data.get('tpc_driver_name') and not str(post_data['tpc_driver_name']).isdigit():
            from ..models import DrivermasterInfo
            drv, _ = DrivermasterInfo.objects.get_or_create(
                dm_name=str(post_data['tpc_driver_name']).strip(),
                defaults={
                    'dm_vehiclesource_id': post_data.get('tpc_vehicle_source') or 1
                }
            )
            post_data['tpc_driver_name'] = str(drv.id)

        if post_data.get('tpc_to'):
            if not str(post_data['tpc_to']).isdigit():
                post_data['tpc_to_manual'] = str(post_data['tpc_to']).strip()
                post_data['tpc_to'] = ''
            else:
                post_data['tpc_to_manual'] = ''
        else:
            post_data['tpc_to_manual'] = ''

        if tpc_id == 0:
            form = TMSPettyCashForm(post_data, request=request)
        else:
            tpc = get_object_or_404(TMSPettyCashInfo, pk=tpc_id)
            form = TMSPettyCashForm(post_data, instance=tpc, request=request)
            
        if form.is_valid():
            saved_tpc = form.save(commit=False)
            if tpc_id == 0:
                saved_tpc.tpc_number = generate_tms_petty_cash_number(TMSPettyCashInfo, 'tpc_number', saved_tpc.tpc_branch)
            saved_tpc.save()
            messages.success(request, "Petty Cash saved successfully.")
            return redirect('tms_petty_cash_list')
        else:
            messages.error(request, "Please correct the errors below.")
            
    from_location = ""
    to_location = ""
    if tpc_id != 0:
        trip = TripdetailInfo.objects.filter(
            Q(tr_consignmentnumber__co_consignmentnumber=tpc.tpc_job_no) | 
            Q(tr_tripnumber=tpc.tpc_job_no)
        ).first()
        if trip:
            from_location = trip.tr_departedlocation.place_name if trip.tr_departedlocation else ""
            to_location = trip.tr_reportedlocation.place_name if trip.tr_reportedlocation else ""

    context = {
        'form': form,
        'tpc_id': tpc_id,
        'from_location': from_location,
        'to_location': to_location,
    }
    return render(request, "asset_mgt_app/tms_petty_cash_add.html", context)


def tms_petty_cash_list(request):
    tpc_number = request.GET.get('tpc_number', "").strip()
    search_date = request.GET.get('search_date', "").strip()
    from_date = request.GET.get('from_date', "").strip()
    to_date = request.GET.get('to_date', "").strip()
    search_vehicle = request.GET.get('search_vehicle', "").strip()
    search_branch = request.GET.get('search_branch', "").strip()
    
    filters = Q()
    if tpc_number:
        filters &= Q(tpc_number__icontains=tpc_number)
    if search_date:
        filters &= Q(tpc_trip_date=search_date)
    if from_date:
        filters &= Q(tpc_transaction_date__gte=from_date)
    if to_date:
        filters &= Q(tpc_transaction_date__lte=to_date)
    if search_vehicle:
        filters &= Q(tpc_vehicle_number__vm_registrationnumber__icontains=search_vehicle)
        
    user = request.user
    is_admin_or_supervisor = user.is_superuser
    if not is_admin_or_supervisor:
        try:
            from ..sub_models.user_ext_mod import User_extInfo
            user_ext = User_extInfo.objects.select_related('emp_designation', 'emp_role').get(user_id=user.id)
            desig = str(user_ext.emp_designation).lower() if user_ext.emp_designation else ''
            role = str(user_ext.emp_role).lower() if user_ext.emp_role else ''
            if 'supervisor' in desig or 'admin' in role:
                is_admin_or_supervisor = True
        except Exception:
            pass

    if not is_admin_or_supervisor:
        branch_id = get_session_branch_id(request)
        if branch_id:
            filters &= Q(tpc_branch_id=branch_id)
    else:
        if search_branch:
            filters &= Q(tpc_branch__loc_name__icontains=search_branch)
            
            
    tpc_list = TMSPettyCashInfo.objects.filter(filters).order_by('-id')
    paginator = Paginator(tpc_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Dynamically add from_to_location to each object in the page
    for tpc in page_obj.object_list:
        tpc.from_to_location = ""
        if tpc.tpc_job_no:
            try:
                trip = TripdetailInfo.objects.filter(
                    Q(tr_consignmentnumber__co_consignmentnumber=tpc.tpc_job_no) | 
                    Q(tr_tripnumber=tpc.tpc_job_no)
                ).first()
                if trip:
                    fr = trip.tr_departedlocation.place_name if trip.tr_departedlocation else ""
                    to = trip.tr_reportedlocation.place_name if trip.tr_reportedlocation else ""
                    if fr or to:
                        tpc.from_to_location = f"{fr} - {to}"
            except Exception:
                pass
                
    context = {
        'tpc_list': page_obj,
        'search_tpc_number': tpc_number,
        'search_date': search_date,
        'from_date': from_date,
        'to_date': to_date,
        'search_vehicle': search_vehicle,
        'search_branch': search_branch,
        'is_admin_or_supervisor': is_admin_or_supervisor,
    }
    return render(request, "asset_mgt_app/tms_petty_cash_list.html", context)

def tms_petty_cash_delete(request, tpc_id):
    tpc = get_object_or_404(TMSPettyCashInfo, pk=tpc_id)
    tpc.delete()
    messages.success(request, "Petty Cash deleted successfully.")
    return redirect('tms_petty_cash_list')

def get_tms_trips_by_date(request):
    trip_date = request.GET.get('trip_date')
    if not trip_date:
        return JsonResponse({'status': 'error', 'message': 'No date provided'})
        
    try:
        from datetime import datetime
        # Check if format is DD-MM-YYYY
        if len(trip_date.split('-')[0]) == 2:
            parsed_date = datetime.strptime(trip_date, "%d-%m-%Y").date()
            trip_date = parsed_date.strftime("%Y-%m-%d")
    except Exception:
        pass
        
    trips = TripdetailInfo.objects.filter(tr_departeddate__date=trip_date).exclude(tr_category__category__icontains='empty')
    
    trip_data = []
    for t in trips:
        customer_name = ""
        cnote_no = ""
        customer_id = ""
        
        if t.tr_consignmentnumber:
            cnote_no = t.tr_consignmentnumber.co_consignmentnumber or ""
            if t.tr_consignmentnumber.co_customer:
                customer_name = t.tr_consignmentnumber.co_customer.cu_name or ""
                customer_id = t.tr_consignmentnumber.co_customer.id
                
        trip_data.append({
            'trip_id': t.id,
            'cnote_no': cnote_no,
            'customer_id': customer_id,
            'customer_name': customer_name,
            'vehicle_source_id': t.tr_vehiclesource.id if t.tr_vehiclesource else "",
            'vehicle_source_name': t.tr_vehiclesource.ow_ownership if t.tr_vehiclesource else "",
            'vehicle_number': t.tr_vehiclenumber or "",
            'driver_name': t.tr_drivername or "",
            'from_location': t.tr_departedlocation.place_name if t.tr_departedlocation else "",
            'to_location': t.tr_reportedlocation.place_name if t.tr_reportedlocation else ""
        })
        
    return JsonResponse({'status': 'success', 'data': trip_data})

def tms_petty_cash_export_tally(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    tpc_number = request.GET.get('tpc_number', "").strip()
    search_date = request.GET.get('search_date', "").strip()
    from_date = request.GET.get('from_date', "").strip()
    to_date = request.GET.get('to_date', "").strip()
    search_vehicle = request.GET.get('search_vehicle', "").strip()
    search_branch = request.GET.get('search_branch', "").strip()
    
    filters = Q()
    if tpc_number:
        filters &= Q(tpc_number__icontains=tpc_number)
    if search_date:
        filters &= Q(tpc_trip_date=search_date)
    if from_date:
        filters &= Q(tpc_transaction_date__gte=from_date)
    if to_date:
        filters &= Q(tpc_transaction_date__lte=to_date)
    if search_vehicle:
        filters &= Q(tpc_vehicle_number__vm_registrationnumber__icontains=search_vehicle)

    user = request.user
    is_admin_or_supervisor = user.is_superuser
    if not is_admin_or_supervisor:
        try:
            from ..sub_models.user_ext_mod import User_extInfo
            user_ext = User_extInfo.objects.select_related('emp_designation', 'emp_role').get(user_id=user.id)
            desig = str(user_ext.emp_designation).lower() if user_ext.emp_designation else ''
            role = str(user_ext.emp_role).lower() if user_ext.emp_role else ''
            if 'supervisor' in desig or 'admin' in role:
                is_admin_or_supervisor = True
        except Exception:
            pass

    if not is_admin_or_supervisor:
        branch_id = get_session_branch_id(request)
        if branch_id:
            filters &= Q(tpc_branch_id=branch_id)
    else:
        if search_branch:
            filters &= Q(tpc_branch__loc_name__icontains=search_branch)
        
    tpc_list = TMSPettyCashInfo.objects.filter(filters).order_by('-id').select_related(
        'tpc_expense_type', 'tpc_credit_ledger', 'tpc_branch', 'tpc_customer', 
        'tpc_vehicle_source', 'tpc_vehicle_number', 'tpc_driver_name'
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Export"
    
    headers = [
        "DATE", "VOU. NO.", "DEBIT", "CREDIT", "PRIMARY COST CATEGORY", 
        "CUSTOMER", "JOB NO", "VEH.NO.", "AMOUNT", "DRIVER NAME", 
        "TRN DATE", "REMARKS"
    ]
    ws.append(headers)
    
    # Style Header: Yellow Background (#FFFF00), Bold, Center alignment
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFFF00")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Process and append data rows
    for tpc in tpc_list:
        date_str = tpc.tpc_transaction_date.strftime("%d-%m-%Y") if tpc.tpc_transaction_date else ""
        vou_no = tpc.tpc_number or ""
        debit = tpc.tpc_expense_type.tms_exp_type_name if tpc.tpc_expense_type else ""
        credit = tpc.tpc_credit_ledger.ledger_name if tpc.tpc_credit_ledger else ""
        
        # Primary Cost Category
        branch_code = ""
        if tpc.tpc_branch:
            loc_name_upper = tpc.tpc_branch.loc_name.upper()
            if "MAA" in loc_name_upper:
                branch_code = "Maa"
            elif "BLR" in loc_name_upper:
                branch_code = "Blr"
            else:
                branch_code = tpc.tpc_branch.loc_name.split()[-1].title()
                
        source_code = ""
        if tpc.tpc_vehicle_source:
            source_name = tpc.tpc_vehicle_source.ow_ownership.lower()
            if "market" in source_name:
                source_code = "Mkt"
            elif "attached" in source_name:
                source_code = "Att"
            elif "own" in source_name:
                source_code = "Own"
            else:
                source_code = tpc.tpc_vehicle_source.ow_ownership[:3].title()
        elif tpc.tpc_unit:
            source_code = tpc.tpc_unit
            
        primary_cost_cat = f"{branch_code}-{source_code}" if (branch_code and source_code) else (branch_code or source_code or "")
        
        # Customer
        customer = tpc.tpc_customer.cu_name if tpc.tpc_customer else ""
        
        # Job No
        job_no = tpc.tpc_job_no or ""
        
        # Veh. No: "Mkt" constant for market vehicles, else the registration number
        veh_no = ""
        if tpc.tpc_vehicle_source and "market" in tpc.tpc_vehicle_source.ow_ownership.lower():
            veh_no = "Mkt"
        elif tpc.tpc_vehicle_number:
            reg_num = tpc.tpc_vehicle_number.vm_registrationnumber or ""
            if tpc.tpc_vehicle_source and "attached" in tpc.tpc_vehicle_source.ow_ownership.lower():
                veh_no = f"{reg_num}(A)"
            else:
                veh_no = reg_num
            
        # Amount
        amount = tpc.tpc_amount or 0.0
        
        # Driver Name
        driver_name = tpc.tpc_driver_name.dm_name if tpc.tpc_driver_name else ""
        
        # Trn Date (DD-MMM, e.g. 28-Jul)
        trn_date = tpc.tpc_trip_date.strftime("%d-%b") if tpc.tpc_trip_date else ""
        
        # Remarks: Route string from trip if available, else tpc_remarks
        remarks = tpc.tpc_remarks or ""
        if job_no:
            trip = TripdetailInfo.objects.filter(
                Q(tr_consignmentnumber__co_consignmentnumber=job_no) | 
                Q(tr_tripnumber=job_no)
            ).first()
            if trip:
                from_loc = trip.tr_departedlocation.place_name if trip.tr_departedlocation else ""
                to_loc = trip.tr_reportedlocation.place_name if trip.tr_reportedlocation else ""
                if from_loc or to_loc:
                    remarks = f"{from_loc}-{to_loc}"
                    
        ws.append([
            date_str, vou_no, debit, credit, primary_cost_cat,
            customer, job_no, veh_no, amount, driver_name,
            trn_date, remarks
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="TMS_Petty_Cash_Tally_Export.xlsx"'
    wb.save(response)
    return response


def get_ledger_balance(request):
    ledger_id = request.GET.get('ledger_id')
    if not ledger_id:
        return JsonResponse({'status': 'error', 'message': 'No ledger ID provided'})
    
    from django.db.models import Sum
    from ..models import iou_info
    
    total_iou = iou_info.objects.filter(iou_credit_ledger_id=ledger_id).aggregate(Sum('amount'))['amount__sum'] or 0.0
    total_spent = TMSPettyCashInfo.objects.filter(tpc_credit_ledger_id=ledger_id).aggregate(Sum('tpc_amount'))['tpc_amount__sum'] or 0.0
    balance = total_iou - total_spent
    
    return JsonResponse({
        'status': 'success',
        'balance': f"{balance:.2f}"
    })
