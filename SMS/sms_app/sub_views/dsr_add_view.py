from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.core.paginator import Paginator
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from .send_department_email import send_department_email
from ..forms import DsrForm
from ..models import CustomerInfo,Warehouse_goods_info
from django.shortcuts import redirect
import openpyxl
from io import BytesIO
from datetime import datetime
from ..sub_models.damagereport_mod import DamagereportInfo
from ..sub_models.gatein_mod import Gatein_info


@login_required(login_url='login_page')
def dsr_reports(request):
    first_name = request.session.get('first_name')
    form = DsrForm(request.POST or None)
    customer_name = request.POST.get('ds_customer', '')
    goods_list = Warehouse_goods_info.objects.all()
    if customer_name:
        goods_list = goods_list.filter(wh_customer_name=customer_name)
        print(f"Filtering by customer name: {customer_name}")
    paginator = Paginator(goods_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'first_name': first_name,
        'form': form,
        'page_obj': page_obj,
        'customer_name': customer_name,
    }
    return render(request, "asset_mgt_app/dsr_report.html", context)
@login_required(login_url='login_page')
def dsr_send_email_view(request, pre_gatein_id=None, customer_name=None, subject=None):
    print('Entering dsr_send_email_view')
    header_font = Font(name="Arial", bold=True, size=11)  # Make the header row bold
    cell_font = Font(name="Arial", bold=False, size=10)  # settings fro cells
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    if request.method == 'POST' or pre_gatein_id:
        recipient = request.POST.get('recipient')
        message = request.POST.get('message')
        customer_name_1 = customer_name
        if customer_name_1 is None:
            customer_name = request.POST.get('ds_customer')
        else:
            customer_name = customer_name

        recipient_list = [email.strip() for email in recipient.split(',')]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DSR Report"

        customer_obj = CustomerInfo.objects.get(id=customer_name)
        customer_name_str = customer_obj.cu_name.upper()

        # Step 1: Choose headers based on customer
        if "DHL" in customer_name_str:
            headers = ["S.No", "Date & Time Of Arrival", "Un-Loading Start Date & Time", "Un-Loading End Date & Time", "Transporter", "Truck Number","Consigner","Consignee",
                "Docs Received", "HAWB", "Destination", "Invoice Number", "PO Number",
                "Invoice Qty", "Invoice Weight (kg)", "Checkin Weight (kg)", "UOM", "Length",
                "Width", "Height", "Dims Qty",
                "Invoice Value", "Invoice Currency", "Invoice (INR)", "E-Way Bill#", "E-Way Bill Validity",
                "Fumigation Status", "Branch", "Storage Days","Job Number", "Stock Number", "Remarks","Fumigation"
                ]
        elif "EIPL" in customer_name_str:
            headers = ["S:NO", "DATE & TIME OF ARRIVAL","DATE & MAIL RECEIVED","UNLODING START DATE & TIME", "UNLODING END DATE & TIME","TRANSPORTERS", "TRUCK NUM","SHIPPER NAME","CONSIGNEE",
                "DOCS RECEIVED","CUSTOMER SERVICE NAME","CHA","HAWB", "DESTINATION", "INVOICE NUM", "CASE NO",
                "TOTAL NO OF PCS", "INVOICE WEIGHT(kg)", "ACTUAL WEIGHT(kg)","L CMS",
                "B CMS", "H CMS", "Carton in DIM","VOLUME WEIGHT","CBM",
                "INVOICE VALUE", "CURRENCY TYPE", "VALUE IN INR", "E-WAY BILL NO", "VALID TILL",
                "FUMIGATION STATUS", "LOCATION", "DAYS IN WAREHOUSE","Job Number", "Stock Number", "REMARKS"]
        elif "DBS" in customer_name_str:
            headers = ["S:NO", "DATE OF ARRIVAL", "UNLODING START DATE & TIME", "UNLODING END DATE & TIME","TRUCK NUM","SHIPPER NAME",
                "DESTINATION","VALUE IN","INVOICE VALUE","ORDER NO", "INVOICE NUM","INVOICE WEIGHT",
                "TOTAL NO OF PCS", "TOTAL WEIGHT", "L CMS",
                "B CMS", "H CMS", "Carton in DIM","VOLUME WEIGHT","DOCS RECD","LOCATION","DAYS IN WAREHOUSE",
                "E-WAY BILL NO", "E-WAY BILL VALID DATE","Job Number", "Stock Number", "REMARKS"]

        elif  "JEENA" in customer_name_str:
            headers = ["S:NO", "DATE", "Dock Intime at unloading bay", "Dock Outime from unloading bay", "TRUCK NUM", "SHIPPER NAME", "CONSIGNEE", "DESTINATION",
                "VALUE IN","Invoice Value", "Invoice Number", "INVOICE WEIGHT", "Total No of Cartons",
                "Gross Weight","L CMS",
                "B CMS", "H CMS", "Carton in DIM", "Volume weight", "LOCATION", "Documents received with goods",
                "Job Number", "Stock Number", "REMARKS"]
        elif "MAERSK" in customer_name_str:
            headers = ["S:NO","IN DATE", "Dock Intime at unloading bay",
                       "Dock Outime from unloading bay", "TRUCK NUM", "SHIPPER NAME", "Invoice Number", "PO NO","DESTINATION","Total No of Cartons",
                       "QTY AS PER INV","CBM",
                       "INVOICE WEIGHT", "GROSS WEIGHT", "L CMS",
                       "B CMS", "H CMS", "Cartons", "Volume weight", "LOCATION","DOCS RECEIVED WITH GOODS",
                       "OTL NO",
                       "Job Number", "Stock Number", "REMARKS"]
        elif "DSV" in customer_name_str:
            headers = ["S:NO","DATE","Dock Intime at unloading bay","Dock Outime from unloading bay","TRUCK NUM","SHIPPER NAME","CONSIGNEE","DESTINATION",
                "VALUE IN","VALUE","Invoice Number","PO NUMBER/BATCH NO","INVOICE WEIGHT","Total No of Cartons",
                "WH WEIGHT","L CMS",
                "B CMS", "H CMS", "Cartons","Volume weight","LOCATION","Documents received with goods","E-Way Bill", "E-Way Bill Validity",
                "Job Number", "Stock Number", "REMARKS"]
        else:
            headers = [
                "Job Number", "Stock Number", "Customer", "Date Of Arrival", "Unloading Start Time",
                "Unloading End Time", "Transporter", "Truck Number", "Consignor", "Consignee",
                "Docs Received", "HAWB", "Destination", "Invoice Number", "Case Number",
                "Invoice Qty", "Invoice Weight (kg)", "Checkin Weight (kg)", "UOM", "Length",
                "Width", "Height", "Dims Qty", "Package Type", "Volume Weight", "CBM",
                "Invoice Value", "Invoice Currency", "Invoice (INR)", "E-Way Bill#", "E-Way Bill Validity",
                "Fumigation Status", "Check In-Out?", "Branch", "Unit", "Bay", "Storage Days",
                "Damage/Deviation?", "GRN Number", "Damages", "Deviations", "Remarks"
            ]
        ws.append(headers)
        gate_in_ids = Gatein_info.objects.filter(gatein_pre_id=pre_gatein_id).values_list('id', flat=True)
        stock_values = None

        if customer_name and gate_in_ids.exists():
            stock_values = Warehouse_goods_info.objects.filter(
                wh_customer_name=customer_name,
                wh_gate_injob_no_id__in=list(gate_in_ids)
            )
        elif customer_name:
            stock_values = Warehouse_goods_info.objects.filter(wh_customer_name=customer_name)
        elif gate_in_ids.exists():
            stock_values = Warehouse_goods_info.objects.filter(wh_gate_injob_no_id__in=list(gate_in_ids))

        if stock_values is not None:
            stock_values = stock_values.filter(wh_check_in_out="1").order_by('-wh_gate_injob_no_id__gatein_arrival_date')

            for i, stock_value in enumerate(stock_values):

                if "DHL" in customer_name_str:
                    date_of_arrival = stock_value.wh_gate_injob_no_id.gatein_arrival_date
                    if date_of_arrival:
                        date_of_arrival = date_of_arrival.replace(tzinfo=None)
                    damage_report = DamagereportInfo.objects.filter(dam_wh_job_num=stock_value.wh_job_no).first()

                    remarks = damage_report.dam_comments if damage_report else ""

                    row = [
                        i + 1,
                        str(getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_transporter", ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_truck_number", ""),
                        stock_value.wh_consigner,
                        stock_value.wh_consignee,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_hawb', ''),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),
                        stock_value.wh_po_num,
                        stock_value.wh_total_qty,
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_gross_weight,
                        str(stock_value.wh_uom),
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        stock_value.wh_invoice_value,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_stock_invoice_currency', '')),
                        stock_value.wh_invoice_amount_inr,
                        getattr(stock_value.wh_lb_job_no_id, 'lb_eway_bill', ''),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', "") or ""),
                        str(stock_value.wh_fumigation_process or ''),
                        str(stock_value.wh_branch),
                        stock_value.wh_storage_time,
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        stock_value.wh_comments,
                        str(getattr(stock_value.wh_fumigation_action, 'action_taken_by', '') or '')

                    ]
                elif "EIPL" in customer_name_str:
                    row = [
                        i + 1,
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_date_mail_received", ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_transporter", ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_truck_number", ""),
                        stock_value.wh_consigner,
                        stock_value.wh_consignee,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_customer_service_name", ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_CHA", ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_hawb", ""),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_invoice", ""),
                        stock_value.wh_po_num,
                        stock_value.wh_total_qty,
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_gross_weight,
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        stock_value.wh_chargeable_weight,
                        stock_value.wh_cbm,
                        stock_value.wh_invoice_value,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_stock_invoice_currency', '')),
                        stock_value.wh_invoice_amount_inr,
                        getattr(stock_value.wh_lb_job_no_id, 'lb_eway_bill', ''),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', "") or ""),
                        str(stock_value.wh_fumigation_process or ''),
                        str(stock_value.wh_branch),
                        stock_value.wh_storage_time,
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        stock_value.wh_comments,
                    ]

                elif "DBS" in customer_name_str:
                    row = [
                        i + 1,
                        str(getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_truck_number", ""),
                        stock_value.wh_consigner,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        stock_value.wh_invoice_amount_inr,
                        stock_value.wh_invoice_value,
                        stock_value.wh_po_num,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_total_qty,
                        stock_value.wh_gross_weight,
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        stock_value.wh_chargeable_weight,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        str(stock_value.wh_branch),
                        stock_value.wh_storage_time,
                        getattr(stock_value.wh_lb_job_no_id, 'lb_eway_bill', ''),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', "") or ""),
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        stock_value.wh_comments,

                    ]
                elif "JEENA" in customer_name_str:
                    row = [
                        i + 1,
                        str(getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_truck_number", ""),
                        stock_value.wh_consigner,
                        stock_value.wh_consignee,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        stock_value.wh_invoice_amount_inr,
                        stock_value.wh_invoice_value,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_total_qty,
                        stock_value.wh_gross_weight,
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        stock_value.wh_chargeable_weight,
                        str(stock_value.wh_branch),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        stock_value.wh_comments,
                    ]

                elif "MAERSK" in customer_name_str:
                    row = [
                        i + 1,
                        str(getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_truck_number", ""),
                        stock_value.wh_consigner,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),
                        stock_value.wh_po_num,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        stock_value.wh_goods_pieces,
                        stock_value.wh_total_qty,
                        stock_value.wh_cbm,
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_gross_weight,
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        stock_value.wh_chargeable_weight,
                        str(stock_value.wh_branch),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_otl', ''),
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        stock_value.wh_comments,
                    ]

                elif "DSV" in customer_name_str:
                    row = [
                        i + 1,
                        str(getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, "gatein_truck_number", ""),
                        stock_value.wh_consigner,
                        stock_value.wh_consignee,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        stock_value.wh_invoice_value,
                        stock_value.wh_invoice_amount_inr,
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),
                        stock_value.wh_po_num,
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_total_qty,
                        stock_value.wh_gross_weight,
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        stock_value.wh_chargeable_weight,
                        str(stock_value.wh_branch),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        getattr(stock_value.wh_lb_job_no_id, 'lb_eway_bill', ''),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', "") or ""),
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        stock_value.wh_comments,
                    ]
                else:
                    damage_report = DamagereportInfo.objects.filter(dam_wh_job_num=stock_value.wh_job_no).first()
                    damage_names = ", ".join(
                        damage_report.dam_damages1.values_list('damage_name', flat=True)) if damage_report else ""
                    deviation_names = ", ".join(
                        damage_report.dam_deviation1.values_list('deviation_name', flat=True)) if damage_report else ""
                    grn_number = damage_report.dam_GRN_num if damage_report else ""
                    remarks = damage_report.dam_comments if damage_report else ""
                    damage_check_flag = stock_value.wh_damage_check_id == 1

                    row = [
                        stock_value.wh_job_no,
                        stock_value.wh_qr_rand_num,
                        str(stock_value.wh_customer_name or ""),
                        str(getattr(stock_value.wh_gate_injob_no_id, "gatein_arrival_date", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_start_time", "") or ""),
                        str(getattr(stock_value.wh_lb_job_no_id, "lb_stock_unloading_end_time", "") or ""),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_transporter', ''),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_truck_number', ''),
                        stock_value.wh_consigner,
                        stock_value.wh_consignee,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_hawb', ''),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                        getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),
                        stock_value.wh_po_num,
                        stock_value.wh_total_qty,
                        stock_value.wh_invoice_weight_unit,
                        stock_value.wh_gross_weight,
                        str(stock_value.wh_uom),
                        stock_value.wh_goods_length,
                        stock_value.wh_goods_width,
                        stock_value.wh_goods_height,
                        stock_value.wh_goods_pieces,
                        str(stock_value.wh_goods_package_type),
                        stock_value.wh_chargeable_weight,
                        stock_value.wh_cbm,
                        stock_value.wh_invoice_value,
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_stock_invoice_currency', '')),
                        stock_value.wh_invoice_amount_inr,
                        getattr(stock_value.wh_lb_job_no_id, 'lb_eway_bill', ''),
                        str(getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', "") or ""),
                        str(stock_value.wh_fumigation_process or ''),
                        "Stock on Hand" if str(stock_value.wh_check_in_out) == "Checked-In" else "Checked-In",
                        str(stock_value.wh_branch),
                        str(stock_value.wh_unit),
                        str(stock_value.wh_bay),
                        stock_value.wh_storage_time,
                        "Yes" if damage_check_flag else "No",
                        grn_number,
                        damage_names,
                        deviation_names,
                        remarks,
                    ]
                from datetime import datetime
                row = [
                    value.replace(tzinfo=None) if isinstance(value, datetime) and value.tzinfo else value
                    for value in row
                ]
                ws.append(row)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = yellow_fill
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = (max_length + 2)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        wb.close()
        attachment = excel_file
        attachment_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        file_name = f"{customer_obj.cu_name}_DSR_report.xlsx"
        if subject is None:
            subject = f"{customer_obj.cu_name}_DSR Report"

        message = message.replace('\n', '<br>')
        pre_gatein_id = request.session.get('ses_pre_gatein_id')
        send_department_email('warehouse', subject, message, recipient_list, attachment, attachment_type, file_name)
        messages.success(request, "E-mail sent successfully")
        return redirect(request.META['HTTP_REFERER'])

    else:
        messages.error(request, 'Invalid input in the email form.')
    return redirect(request.META['HTTP_REFERER'])

