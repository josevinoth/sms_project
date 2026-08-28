import base64
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from xhtml2pdf import pisa

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from django.contrib import messages
from django.shortcuts import redirect
from itertools import groupby
from operator import itemgetter
from .general_utils import get_base64_image
from ..forms import PregateintruckForm
from ..models import Pregateintruckinfo,Gatein_pre_info,HighvalueInfo,Gatein_info,Warehouse_goods_info
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .send_department_email import send_department_email
from ..sub_models.damagereport_mod import DamagereportInfo
from ..sub_models.loadingbay_mod import Loadingbay_Info
from ..sub_models.transporter_mod import Transporter_name
from django.utils import timezone

@login_required(login_url='login_page')
def pregateintruck_add(request, pregateintruck_id=0):
    first_name = request.session.get('first_name')
    user_id = request.session.get('ses_userID')
    gatein_num_id = request.session['gatein_num_id']
    print(gatein_num_id, 'gatein')

    high_list = HighvalueInfo.objects.filter(hc_pregatein_number_id=gatein_num_id)

    # 👉 Get the latest High Value Checklist for this Gatein (if exists)
    highvalue_instance = high_list.order_by('-id').first()
    checklist = highvalue_instance  # alias for template
    approval_status_id = None
    if highvalue_instance:
        approval_status_id = highvalue_instance.hc_approval_status_id

    if request.method == "GET":
        if pregateintruck_id == 0:
            form = PregateintruckForm(initial={'pregatein_number': gatein_num_id})
            truck = None
            high_value_check = None
            email_enable = False
            invoice_count = 0
            job_count = 0
        else:
            pregateintruck = get_object_or_404(Pregateintruckinfo, pk=pregateintruck_id)
            truck = pregateintruck
            form = PregateintruckForm(instance=pregateintruck)
            request.session['ses_pregateintruck_id'] = pregateintruck_id
            high_value_check = getattr(pregateintruck, 'pregatein_high_value_id', None)

            invoice_count = int(truck.pregatein_invoice_ref or 0)
            job_count = Gatein_info.objects.filter(gatein_truck_number_n=truck).count()
            email_enable = (invoice_count == job_count)

        context = {
            'form': form,
            'first_name': first_name,
            'user_id': user_id,
            'gatein_num_id': gatein_num_id,
            'truck': truck,
            'high_list': high_list,
            'highvalue_instance': highvalue_instance,
            'checklist': checklist,
            'high_value_check': high_value_check,
            'approval_status_id': approval_status_id,  # ✅ add this
            'email_enable': email_enable,
            'invoice_count': invoice_count,
            'job_count': job_count,
        }
        return render(request, "asset_mgt_app/pregateintruck_add.html", context)

    else:
        if pregateintruck_id == 0:
            form = PregateintruckForm(request.POST)
            if form.is_valid():
                pregateintruck = form.save(commit=False)

                # Process driver signature
                driver_data = request.POST.get('driver_signature_data')
                if driver_data:
                    format, imgstr = driver_data.split(';base64,')
                    ext = format.split('/')[-1]
                    data = ContentFile(base64.b64decode(imgstr), name=f'driver_signature.{ext}')
                    pregateintruck.pregatein_driver_signature = data

                # Process supervisor signature
                supervisor_data = request.POST.get('supervisor_signature_data')
                if supervisor_data:
                    format, imgstr = supervisor_data.split(';base64,')
                    ext = format.split('/')[-1]
                    data = ContentFile(base64.b64decode(imgstr), name=f'supervisor_signature.{ext}')
                    pregateintruck.pregatein_supervisor_signature = data

                pregateintruck.save()
                messages.success(request, 'Record Updated Successfully')

                # Automatic Email Logic
                try:
                    invoice_count = int(pregateintruck.pregatein_invoice_ref or 0)
                    job_count = Gatein_info.objects.filter(gatein_truck_number_n=pregateintruck).count()
                    if invoice_count == job_count and invoice_count > 0:
                        gateins = Gatein_info.objects.filter(gatein_truck_number_n=pregateintruck)
                        for g in gateins:
                            customer = g.gatein_customer
                            if customer.cu_automatic_email == 'YES':
                                recipient_list = [customer.cu_email]
                                subject = f"{customer.cu_name} - Truck Wise Report (Auto)"
                                message = "Dear Customer,\n\nPlease find the Gate-In Truck report attached.\n\nRegards,\nBVM Warehouse Team"
                                send_truck_wise_report_logic(gatein_num_id, recipient_list, subject, message)
                                break
                except Exception as e:
                    print(f"Auto-Email Error: {e}")

                last_id = pregateintruck.id
                pregateintruckdetails_list(request, gatein_num_id)
                return redirect('/SMS/pregateintruck_update/' + str(last_id))
            else:
                messages.error(request, 'Record Not Updated Successfully')
                return redirect(request.META['HTTP_REFERER'])

        else:
            pregateintruck = get_object_or_404(Pregateintruckinfo, pk=pregateintruck_id)
            form = PregateintruckForm(request.POST, instance=pregateintruck)
            if form.is_valid():
                pregateintruck = form.save(commit=False)

                # Process driver signature
                driver_data = request.POST.get('driver_signature_data')
                if driver_data:
                    format, imgstr = driver_data.split(';base64,')
                    ext = format.split('/')[-1]
                    data = ContentFile(base64.b64decode(imgstr), name=f'driver_signature.{ext}')
                    pregateintruck.pregatein_driver_signature = data

                # Process supervisor signature
                supervisor_data = request.POST.get('supervisor_signature_data')
                if supervisor_data:
                    format, imgstr = supervisor_data.split(';base64,')
                    ext = format.split('/')[-1]
                    data = ContentFile(base64.b64decode(imgstr), name=f'supervisor_signature.{ext}')
                    pregateintruck.pregatein_supervisor_signature = data

                pregateintruck.save()
                messages.success(request, 'Record Updated Successfully')

                # Automatic Email Logic
                try:
                    invoice_count = int(pregateintruck.pregatein_invoice_ref or 0)
                    job_count = Gatein_info.objects.filter(gatein_truck_number_n=pregateintruck).count()
                    if invoice_count == job_count and invoice_count > 0:
                        gateins = Gatein_info.objects.filter(gatein_truck_number_n=pregateintruck)
                        for g in gateins:
                            customer = g.gatein_customer
                            if customer.cu_automatic_email == 'YES':
                                recipient_list = [customer.cu_email]
                                subject = f"{customer.cu_name} - Truck Wise Report (Auto)"
                                message = "Dear Customer,\n\nPlease find the Gate-In Truck report attached.\n\nRegards,\nBVM Warehouse Team"
                                send_truck_wise_report_logic(gatein_num_id, recipient_list, subject, message)
                                break
                except Exception as e:
                    print(f"Auto-Email Error: {e}")

                pregateintruckdetails_list(request, gatein_num_id)
                return redirect(request.META['HTTP_REFERER'])
            else:
                messages.error(request, 'Record Not Updated Successfully')
                return redirect(request.META['HTTP_REFERER'])
def pregateintruckdetails_list(request,gatein_num_id):
    turck_numbers=list(Pregateintruckinfo.objects.filter(pregatein_number=gatein_num_id).values_list('pregatein_truck_number',flat=True))
    driver_names=list(Pregateintruckinfo.objects.filter(pregatein_number=gatein_num_id).values_list('pregatein_driver',flat=True))
    pre_gatein_num=Gatein_pre_info.objects.get(id=gatein_num_id).gatein_pre_number
    Gatein_pre_info.objects.filter(gatein_pre_number=pre_gatein_num).update(gatein_pre_truck_number=turck_numbers)
    Gatein_pre_info.objects.filter(gatein_pre_number=pre_gatein_num).update(gatein_pre_driver_name=driver_names)
    return (turck_numbers,driver_names)

# List pregateintruck
@login_required(login_url='login_page')
def pregateintruck_list(request):
    first_name = request.session.get('first_name')
    Gatein_pre_list=Pregateintruckinfo.objects.all()
    page_number = request.GET.get('page')
    paginator = Paginator(Gatein_pre_list, 50)
    page_obj = paginator.get_page(page_number)
    context = {
            'page_obj' :page_obj ,
            'first_name': first_name
        }
    return render(request,"asset_mgt_app/gatein_pre_list.html",context)

#Delete pregateintruck
@login_required(login_url='login_page')
def pregateintruck_delete(request,pregateintruck_id):
    pregateintruck = Pregateintruckinfo.objects.get(pk=pregateintruck_id)
    pregateintruck.delete()
    gatein_num_id = request.session['gatein_num_id']
    pregateintruckdetails_list(request,gatein_num_id)
    return (redirect(request.META['HTTP_REFERER'])

# Cancel pregateintruck
@login_required(login_url='login_page'))
def pregateintruck_cancel(request):
    gatein_num_id = request.session['gatein_num_id']
    return redirect('/SMS/gatein_pre_update/' + str(gatein_num_id))
@csrf_exempt
def add_transporter(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if not name:
            return JsonResponse({"success": False, "error": "Transporter name cannot be empty."})

        existing = Transporter_name.objects.filter(transporter_name__iexact=name).first()
        if existing:
            return JsonResponse({
                "success": False,
                "id": existing.id,
                "name": existing.transporter_name,
                "error": "This transporter already exists."
            })

        new = Transporter_name.objects.create(transporter_name=name)
        return JsonResponse({"success": True, "id": new.id, "name": new.transporter_name})

    return JsonResponse({"success": False, "error": "Invalid request"})
@login_required(login_url='login_page')
def pregatein_gatepass_pdf(request, pregatein_id=0, download=False):
    try:
        truck = get_object_or_404(Pregateintruckinfo, id=pregatein_id)

        if truck.pregatein_job_category_id == 1:
            return JsonResponse({"success": False, "error": "Gatepass PDF only available for Job Category 3."}, status=403)

        # Convert signatures to base64 strings
        truck.driver_signature_base64 = get_base64_image(truck.pregatein_driver_signature)
        truck.supervisor_signature_base64 = get_base64_image(truck.pregatein_supervisor_signature)

        context = {
            'truck': truck,
        }

        template = get_template('asset_mgt_app/pregatein_gate_pass.html')
        html = template.render(context)

        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)

        if pisa_status.err:
            raise ValueError('Error generating PDF')

        pdf_buffer.seek(0)
        pdf_data = pdf_buffer.read()
        pdf_buffer.close()

        if download:
            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Pregatein_Gatepass_{truck.pregatein_number}.pdf"'
            return response

        return pdf_data  # For email attachment use

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@login_required(login_url='login_page')
def pregatein_gatepass_pdf_download(request, pregatein_id):
    return pregatein_gatepass_pdf(request, pregatein_id, download=True)


from datetime import datetime

def send_truck_wise_report_logic(pre_gatein_id, recipient_list, subject, message):
    # --- Excel setup ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Truck Wise Details"

    headers = [
        "Vehicle No", "Shipper", "Shipper Value", "Invoice No",
        "Loading Start Time", "Loading End Time", "No. of Pieces",
        "Damage (Yes/No)", "WH Job No", "Stock No", "GRN Number",
        "Damage Names", "Deviation Names", "Remarks",
    ]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Step 1: Get all trucks for this pre-gate-in ---
    trucks = Pregateintruckinfo.objects.filter(pregatein_number_id=pre_gatein_id)

    # --- Step 2: Collect detailed goods data for each truck ---
    for truck in trucks:
        gateins = Gatein_info.objects.filter(gatein_truck_number_n=truck)
        for g in gateins:
            goods = Warehouse_goods_info.objects.filter(wh_gate_injob_no_id=g)
            for item in goods:
                shipper = item.wh_consigner or ""
                shipper_value = item.wh_invoice_value or ""
                invoice = getattr(item.wh_gate_injob_no_id, 'gatein_invoice', '')
                loading_start = getattr(item.wh_lb_job_no_id, "lb_stock_unloading_start_time", "")
                loading_end = getattr(item.wh_lb_job_no_id, "lb_stock_unloading_end_time", "")
                pieces = item.wh_goods_pieces or 0
                damage = "Yes" if getattr(item, 'wh_damage_check_id', 0) == 1 else "No"
                wh_job_no = getattr(item, 'wh_job_no', "")
                stock_no = getattr(item, 'wh_qr_rand_num', "")

                damage_report = DamagereportInfo.objects.filter(dam_wh_job_num=wh_job_no).first()
                if damage_report:
                    grn_number = damage_report.dam_GRN_num or ""
                    remarks = damage_report.dam_comments or ""
                    damage_names = ", ".join(damage_report.dam_damages1.values_list('damage_name', flat=True))
                    deviation_names = ", ".join(damage_report.dam_deviation1.values_list('deviation_name', flat=True))
                else:
                    grn_number = remarks = damage_names = deviation_names = ""

                ws.append([
                    truck.pregatein_truck_number, shipper, shipper_value, invoice,
                    timezone.localtime(loading_start).strftime("%d-%b-%Y %H:%M") if loading_start else "",
                    timezone.localtime(loading_end).strftime("%d-%b-%Y %H:%M") if loading_end else "",
                    pieces, damage, wh_job_no, stock_no, grn_number, damage_names, deviation_names, remarks,
                ])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    excel_file = BytesIO()
    wb.save(excel_file)
    wb.close()
    excel_file.seek(0)

    send_department_email(
        department='warehouse',
        subject=subject,
        message=message.replace('\n', '<br>'),
        recipient_list=recipient_list,
        attachment=excel_file,
        attachment_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        file_name='Truck_Wise_Report.xlsx'
    )

@login_required(login_url='login_page')
def truck_send_email_view(request, pre_gatein_id=None):
    if request.method == 'POST':
        pre_gatein_id = pre_gatein_id or request.POST.get('pre_gatein_id') or request.session.get('gatein_num_id')
        if not pre_gatein_id:
            messages.error(request, "Missing Pre-Gate-In ID.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        recipient = request.POST.get('recipient')
        subject = request.POST.get('subject', 'Truck-wise Report')
        message = request.POST.get('message', '')
        recipient_list = [r.strip() for r in recipient.split(',') if r.strip()]

        send_truck_wise_report_logic(pre_gatein_id, recipient_list, subject, message)
        messages.success(request, "Truck-wise E-mail sent successfully.")
        return redirect(request.META.get('HTTP_REFERER', '/'))
    return redirect(request.META.get('HTTP_REFERER', '/'))



from django.http import JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from sms_app.models import Pregateintruckinfo

def pregateintruck_list_ajax(request):
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    pregatein_number = request.GET.get('pregatein_number', '')

    queryset = Pregateintruckinfo.objects.select_related(
        'pregatein_number', 'pregatein_updated_by'
    ).all()

    if pregatein_number:
        queryset = queryset.filter(pregatein_number__gatein_pre_number__icontains=pregatein_number)

    if search_value:
        queryset = queryset.filter(
            Q(id__icontains=search_value) |
            Q(pregatein_number__gatein_pre_number__icontains=search_value) |
            Q(pregatein_truck_number__icontains=search_value) |
            Q(pregatein_driver__icontains=search_value)
        )

    total_records = Pregateintruckinfo.objects.count()
    filtered_records = queryset.count()

    # Ordering
    order_column_index = request.GET.get('order[0][column]', 1)
    order_dir = request.GET.get('order[0][dir]', 'desc')

    columns = [
        'id', 'id', 'pregatein_created_at', 'pregatein_number__gatein_pre_number',
        'pregatein_number__gatein_pre_branch__loc_name',
        'pregatein_truck_number', 'pregatein_driver', 'pregatein_updated_at',
        'pregatein_updated_by__username'
    ]

    if int(order_column_index) < len(columns):
        order_by = columns[int(order_column_index)]
        if order_by:
            if order_dir == 'desc':
                order_by = f"-{order_by}"
            queryset = queryset.order_by(order_by)
    else:
        queryset = queryset.order_by('-id')

    data = []
    for item in queryset[start:start+length]:
        edit_btn = f'''
        <div class="d-flex justify-content-center gap-1">
            <a class="btn btn-submit" style="background: linear-gradient(135deg, #fbbf24, #f59e0b); border: none; color: white;" href="/SMS/pregateintruck_update/{item.id}" >
                <i class="far fa-edit"></i>
            </a>
        </div>'''
        delete_btn = f'''
        <div class="d-flex justify-content-center gap-1">
            <form action="/SMS/pregateintruck_delete/{item.id}" method="post" onclick="return confirm('Are you sure?');">
                <input type="hidden" name="csrfmiddlewaretoken" value="{request.META.get('CSRF_COOKIE', '')}">
                <button type="submit" class="btn shadow-sm" style="background: white; color: #dc3545; border-radius: 20px; border: 1px solid #f1f3f5; padding: 6px 16px; display: inline-flex; align-items: center; justify-content: center;">
                    <i class="fas fa-trash-alt"></i>
                </button>
            </form>
        </div>'''

        pre_gatein = item.pregatein_number  # FK to Gatein_pre_info
        branch_name = ''
        if pre_gatein and hasattr(pre_gatein, 'gatein_pre_branch') and pre_gatein.gatein_pre_branch:
            branch_name = str(pre_gatein.gatein_pre_branch.loc_name)

        data.append({
            'edit': edit_btn,
            'id': item.id,
            'gatein_pre_created_at': item.pregatein_created_at.strftime('%d-%m-%Y %I:%M %p') if item.pregatein_created_at else '',
            'gatein_pre_number': str(pre_gatein.gatein_pre_number) if pre_gatein else '',
            'gatein_pre_branch': branch_name,
            'gatein_pre_truck_number': item.pregatein_truck_number or '',
            'gatein_pre_driver_name': item.pregatein_driver or '',
            'gatein_pre_updated_at': item.pregatein_updated_at.strftime('%d-%m-%Y %I:%M %p') if item.pregatein_updated_at else '',
            'gatein_pre_updated_by': str(item.pregatein_updated_by) if item.pregatein_updated_by else '',
            'delete': delete_btn
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


