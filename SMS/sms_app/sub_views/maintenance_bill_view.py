from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Q
from ..sub_models.maintenance_bill_mod import MaintenanceBillInfo
from ..sub_forms.maintenance_bill_form import MaintenanceBillForm
from ..sub_models.maintenance_mod import MaintenanceInfo
from ..sub_models.vehiclemaster_mod import VehiclemasterInfo

@login_required(login_url='login_page')
def maintenance_bill_add(request, id=None):
    instance = get_object_or_404(MaintenanceBillInfo, id=id) if id else None
    
    if request.method == "POST":
        form = MaintenanceBillForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            bill = form.save(commit=False)
            if not id:
                bill.mnb_created_by = request.user
            bill.mnb_updated_by = request.user
            bill.save()
            msg = "Maintenance Bill updated successfully." if id else "Maintenance Bill added successfully."
            messages.success(request, msg)
            return redirect('maintenance_bill_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = MaintenanceBillForm(instance=instance)
    
    if instance:
        # In edit mode, we want the table to show ONLY the linked record initially (as per user request "only shown the bill edit")
        # BUT we must ensure the vehicles dropdown contains the current vehicle.
        # We also include other unbilled vehicles so the user can still use the dropdown if needed.
        pending_maintenance = MaintenanceInfo.objects.filter(id=instance.mnb_maintenance_id)
        
        vehicles = VehiclemasterInfo.objects.filter(
            Q(id=instance.mnb_maintenance.mi_vehicle_id) |
            Q(maintenance_records__mi_approval_status_id=3, maintenance_records__bills_v1__isnull=True)
        ).distinct().order_by('vm_registrationnumber')
    else:
        # Fetch only maintenance records that are "Finance Approved" (3) 
        # and have NOT been billed yet (bills_v1__isnull=True)
        pending_maintenance = MaintenanceInfo.objects.filter(
            mi_approval_status_id=3,
            bills_v1__isnull=True
        ).order_by('-mi_created_at')

        # Fetch vehicles that have records in the pending_maintenance above (unbilled ones)
        vehicles = VehiclemasterInfo.objects.filter(
            maintenance_records__mi_approval_status_id=3,
            maintenance_records__bills_v1__isnull=True
        ).distinct().order_by('vm_registrationnumber')
    
    return render(request, "asset_mgt_app/maintenance_bill_add.html", {
        "form": form, 
        "pending_maintenance": pending_maintenance,
        "vehicles": vehicles,
        "is_edit": True if id else False,
        "instance": instance
    })

@login_required(login_url='login_page')
def maintenance_bill_edit(request, id):
    return maintenance_bill_add(request, id=id)

@login_required(login_url='login_page')
def maintenance_bill_delete(request, id):
    if request.session.get('ses_role') not in ['Admin', 'Super User']:
        messages.error(request, "You do not have permission to delete this record.")
        return redirect('maintenance_bill_list')

    bill = get_object_or_404(MaintenanceBillInfo, id=id)
    bill.delete()
    messages.success(request, "Maintenance Bill deleted successfully.")
    return redirect('maintenance_bill_list')

@login_required(login_url='login_page')
def maintenance_bill_list(request):
    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()

    bills = MaintenanceBillInfo.objects.all().select_related(
        'mnb_maintenance__mi_vehicle',
        'mnb_maintenance__mi_location',
        'mnb_created_by',
        'mnb_updated_by'
    ).order_by('-mnb_bill_date')

    if from_date:
        bills = bills.filter(mnb_bill_date__gte=from_date)
    if to_date:
        bills = bills.filter(mnb_bill_date__lte=to_date)

    return render(request, "asset_mgt_app/maintenance_bill_list.html", {
        "bills": bills,
        "from_date": from_date,
        "to_date": to_date,
    })

@login_required(login_url='login_page')
def fetch_maintenance_bill_details(request):
    maintenance_id = request.GET.get('maintenance_id')
    try:
        maintenance = MaintenanceInfo.objects.get(id=maintenance_id)
        advance_val = 0
        try:
            advance_val = float(maintenance.mi_advance) if maintenance.mi_advance else 0
        except (ValueError, TypeError):
            advance_val = 0
            
        tds_type = ""
        pan = ""
        if maintenance.mi_technician:
            from ..sub_models.vendor_info_mod import Vendor_info
            vendor = Vendor_info.objects.filter(vend_name__iexact=maintenance.mi_technician).first()
            if vendor and vendor.vend_pan:
                pan = vendor.vend_pan.strip().upper()
                if len(pan) >= 4 and pan[3] in ('C', 'F'):
                    tds_type = 'Company'
                else:
                    tds_type = 'Non company'
                    
        data = {
            "vehicle_no": maintenance.mi_vehicle.vm_registrationnumber,
            "vehicle_type": maintenance.mi_vehicle.vm_vehicletype.vt_vehicletype if maintenance.mi_vehicle.vm_vehicletype else "N/A",
            "service_type": maintenance.mi_service_type,
            "estimated_amount": float(maintenance.mi_estimated_amount) if maintenance.mi_estimated_amount else 0,
            "vendor_name": maintenance.mi_technician if maintenance.mi_technician else "N/A",
            "technician": maintenance.mi_technician,
            "advance": advance_val,
            "tds_type": tds_type,
            "pan": pan,
        }
        return JsonResponse(data)
    except MaintenanceInfo.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

@login_required(login_url='login_page')
def get_maintenance_records_by_vehicle(request):
    vehicle_id = request.GET.get('vehicle_id')
    current_mi_id = request.GET.get('current_mi_id') # Optional: to keep the currently selected even if billed
    
    # Match records with status 3 (Finance Approved)
    # Include those that are unbilled OR the one currently being edited
    if current_mi_id:
        records = MaintenanceInfo.objects.filter(
            Q(mi_vehicle_id=vehicle_id, mi_approval_status_id=3, bills_v1__isnull=True) |
            Q(id=current_mi_id)
        ).distinct().order_by('-mi_created_at')
    else:
        records = MaintenanceInfo.objects.filter(
            mi_vehicle_id=vehicle_id, 
            mi_approval_status_id=3,
            bills_v1__isnull=True
        ).order_by('-mi_created_at')
    
    data = []
    for r in records:
        data.append({
            "id": r.id,
            "job_card_no": r.mi_job_card_no or f"JC-{r.id}",
            "service_type": r.mi_service_type,
            "estimated_amount": str(r.mi_estimated_amount),
            "created_at": r.mi_created_at.strftime('%Y-%m-%d'),
            "status": str(r.mi_approval_status) if r.mi_approval_status else "N/A"
        })
    
    return JsonResponse({"records": data})

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required(login_url='login_page')
def maintenance_bill_export_tally(request):
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    bills = MaintenanceBillInfo.objects.all().select_related(
        'mnb_maintenance__mi_vehicle',
        'mnb_maintenance__mi_location',
    ).order_by('mnb_bill_date')
    if from_date:
        bills = bills.filter(mnb_bill_date__gte=from_date)
    if to_date:
        bills = bills.filter(mnb_bill_date__lte=to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Export"

    headers = [
        "VOUCHER NUMBER", "DATE", "REF NO.", "SUNDRY CREDITOR", "TOTAL AMT",
        "EXPENSES LEDGER", "AMOUNT", "Primary Cost Category", "Job No", 
        "Vehicle No", "Customer", "TDS LEDGER", "TDS AMOUNT", "NARRATION"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="B2FFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    expense_fill = PatternFill("solid", fgColor="FFE5CC")
    
    for bill in bills:
        maintenance = bill.mnb_maintenance
        
        voucher_number = bill.get_voucher_number
        bill_date = bill.mnb_bill_date.strftime("%d-%m-%Y") if bill.mnb_bill_date else ""
        ref_no = bill.mnb_bill_no or ""
        vendor_name = maintenance.mi_technician or ""
        total_amt = float(bill.mnb_amount_payable or bill.mnb_total_amount or 0)
        expense_amt = float(bill.mnb_bill_amount_taxable or 0)
        expense_ledger = bill.mnb_expenses_type or "Vehicle Maintenance"
        
        tds_amount = float(bill.mnb_tds_amount or 0)
        tds_ledger = ""
        try:
            tds_type = (bill.mnb_tds_type or '').strip()
            if tds_amount > 0:
                if tds_type == 'Company':
                    tds_ledger = "TDS Payable 194C (Company)"
                else:
                    tds_ledger = "TDS Payable 194C (Non Company)"
        except Exception:
            tds_ledger = "TDS Payable 194C (Non Company)" if tds_amount > 0 else ""
        
        service_type = maintenance.mi_service_type or "service"
        month_str_narration = bill.mnb_bill_date.strftime('%b%y') if bill.mnb_bill_date else ""
        rec_date_str = bill.mnb_created_at.strftime('%d-%b-%y') if bill.mnb_created_at else ""
        narration = f"being {service_type} done for the month of {month_str_narration} (Bill received on -{rec_date_str})"
        
        job_no = "NA(J)"
        vehicle_no = maintenance.mi_vehicle.vm_registrationnumber if maintenance.mi_vehicle else ""
        customer_name = "NA(C)"
        
        primary_cost_cat = "MAA - OWN"
        if vehicle_no and vehicle_no.upper().startswith("KA"):
            primary_cost_cat = "BLR - OWN"
        elif vehicle_no and vehicle_no.upper().startswith("TN"):
            primary_cost_cat = "MAA - OWN"
        
        row = [
            voucher_number, bill_date, ref_no, vendor_name, total_amt,
            expense_ledger, expense_amt, primary_cost_cat, job_no,
            vehicle_no, customer_name, tds_ledger, tds_amount if tds_amount else "", narration
        ]
        ws.append(row)
        ws.cell(row=ws.max_row, column=6).fill = expense_fill

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Maintenance_Bill_Tally_Export.xlsx"'
    wb.save(response)
    return response
