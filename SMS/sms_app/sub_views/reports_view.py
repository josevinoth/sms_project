import csv
import json
from itertools import chain
from io import BytesIO

from django.core.exceptions import ObjectDoesNotExist
from django.http import StreamingHttpResponse
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.db.models import Q, ExpressionWrapper, fields, F, DurationField
from django.db.models.functions import Cast, Extract, TruncMonth
from django.shortcuts import render
from django.template.loader import get_template
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import make_naive
from xhtml2pdf import pisa
from ..models import CustomerInfo,ExpenseInfo,Gatein_info,LocationmasterInfo,Loadingbay_Info,DamagereportInfo,Warehouse_goods_info,ExpenseExtinfo,GoodsPartialDispatchInfo,Location_info
from datetime import datetime, timedelta, time
from django.utils.dateparse import parse_date
from django.db.models import Count, Sum
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from .send_department_email import send_department_email
from django.shortcuts import redirect
from django.contrib import messages
from ..forms import DsrForm
from django.utils.timezone import make_aware, now
from itertools import zip_longest

@login_required(login_url='login_page')
def reports(request):
    first_name = request.session.get('first_name')
    context = {
               'first_name': first_name
               }
    return render(request,"asset_mgt_app/reports.html",context)

@login_required(login_url='login_page')
def warehouse_reports(request):
    first_name = request.session.get('first_name')
    context = {
               'first_name': first_name
               }
    return render(request,"asset_mgt_app/warehouse_reports.html",context)

@login_required(login_url='login_page')
def space_availability_reports(request):
    first_name = request.session.get('first_name')
    context = {
                'space_utilization_list': LocationmasterInfo.objects.all(),
                'first_name': first_name,
                }
    return render(request,"asset_mgt_app/space_availability_report.html",context)


@login_required(login_url='login_page')
def space_utilization_reports(request):
    first_name = request.session.get('first_name')
    branches = Location_info.objects.all()

    selected_branch = request.GET.get('branch', '')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    from_date = parse_date(from_date_str) if from_date_str else None
    to_date = parse_date(to_date_str) if to_date_str else None

    # Base filter
    goods_filter = {'wh_check_in_out': 1}
    if selected_branch:
        goods_filter['wh_branch__loc_name'] = selected_branch
    if from_date:
        goods_filter['wh_checkin_time__gte'] = datetime.combine(from_date, time.min)
    if to_date:
        goods_filter['wh_checkin_time__lte'] = datetime.combine(to_date, time.max)

    space_utilization_list = []

    # Fast bulk aggregation
    aggregated_goods = Warehouse_goods_info.objects.filter(**goods_filter).values(
        'wh_branch_id', 'wh_unit_id', 'wh_bay_id'
    ).annotate(
        total_area=Sum('wh_goods_area'),
        total_volume=Sum('wh_goods_volume_weight')
    )

    # Create lookup dict O(1)
    goods_lookup = {}
    for agg in aggregated_goods:
        key = (agg['wh_branch_id'], agg['wh_unit_id'], agg['wh_bay_id'])
        goods_lookup[key] = {
            'area': agg['total_area'] or 0,
            'volume': agg['total_volume'] or 0
        }

    # select_related to prevent N+1 queries for string representations
    locations = LocationmasterInfo.objects.select_related(
        'lm_wh_location', 'lm_wh_unit', 'lm_areaside', 'lm_customer_name', 'lm_customer_model'
    ).all()

    for loc in locations:
        key = (loc.lm_wh_location_id, loc.lm_wh_unit_id, loc.lm_areaside_id)
        agg_data = goods_lookup.get(key, {'area': 0, 'volume': 0})

        occupied_area = agg_data['area']
        occupied_volume = agg_data['volume']

        available_area = loc.lm_size - occupied_area
        available_volume = loc.lm_total_volume - occupied_volume

        space_utilization_list.append({
            'lm_wh_location': loc.lm_wh_location,
            'lm_wh_unit': loc.lm_wh_unit,
            'lm_areaside': loc.lm_areaside,
            'lm_customer_name': loc.lm_customer_name,
            'lm_customer_model': loc.lm_customer_model,
            'lm_size': loc.lm_size,
            'lm_area_occupied': round(occupied_area, 2),
            'lm_available_area': round(available_area, 2),
            'lm_total_volume': loc.lm_total_volume,
            'lm_volume_occupied': round(occupied_volume, 2),
            'lm_available_volume': round(available_volume, 2),
        })

    if request.GET.get('draw'):
        from django.http import JsonResponse
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))

        # Paginate the built list
        total_records = len(space_utilization_list)
        paginated_list = space_utilization_list[start:start+length]

        data = []
        for item in paginated_list:
            data.append([
                str(item.get('lm_wh_location', '')),
                str(item.get('lm_wh_unit', '')),
                str(item.get('lm_areaside', '')),
                str(item.get('lm_size', '')),
                str(item.get('lm_area_occupied', '')),
                str(item.get('lm_available_area', ''))
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })

    context = {
        'first_name': first_name,
        'selected_branch': selected_branch,
        'branches': branches,
        'from_date': from_date.strftime('%Y-%m-%d') if from_date else '',
        'to_date': to_date.strftime('%Y-%m-%d') if to_date else '',
    }
    return render(request, "asset_mgt_app/space_utilization_report.html", context)

@login_required(login_url='login_page')
def stock_value_reports(request):
    print("Inside Stock Value Report")
    first_name = request.session.get('first_name')
    form = DsrForm(request.POST or None)
    customer_name = request.POST.get('ds_customer', '').strip()

    # Warehouse_goods_info.objects.filter(wh_check_in_out=1).update(
    #     wh_storage_time=Cast(
    #         Extract(ExpressionWrapper(
    #             datetime.date.today() - F('wh_checkin_time'),
    #             output_field=DurationField()
    #         ), 'days'),
    #         output_field=fields.FloatField()
    #     )
    # )
    # SKIP MASSIVE DB UPDATE IF IT IS AN AJAX CALL
    is_ajax = bool(request.GET.get('draw') or request.POST.get('draw'))

    if not is_ajax:
        # Only do this on initial full page load, not during DataTables sorting/filtering!
        goods_list_update = Warehouse_goods_info.objects.filter(wh_check_in_out=1)
        # Using bulk_update would be faster, but for now we just prevent it on AJAX
        # (This is still slow on first load, so we comment it out entirely
        # and let it calculate dynamically in the template/JSON instead of writing to DB)
        # for stock in goods_list_update:
        #     if stock.wh_checkin_time:
        #         delta = datetime.now() - stock.wh_checkin_time.replace(tzinfo=None)
        #         stock.wh_storage_time = delta.days
        #         stock.save(update_fields=['wh_storage_time'])

    checkin_goods_list = []


    goods_list = Warehouse_goods_info.objects.select_related('wh_dispatch_id', 'wh_lb_job_no_id', 'wh_gate_injob_no_id', 'wh_customer_name', 'wh_branch', 'wh_unit', 'wh_bay', 'wh_truck_type', 'wh_uom', 'wh_goods_package_type', 'wh_check_in_out', 'wh_fumigation_process').all().order_by('-id')

    # Support both GET and POST for AJAX DataTables compatibility
    if not customer_name:
        customer_name = request.GET.get('ds_customer', '').strip()

    if customer_name:
        goods_list = goods_list.filter(wh_customer_name=customer_name)
        print(f"Filtering by customer name: {customer_name}")

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        try:
            from django.utils import timezone
            dt_from = timezone.make_aware(datetime.strptime(from_date, "%Y-%m-%d"))
            dt_to = timezone.make_aware(datetime.strptime(to_date, "%Y-%m-%d")).replace(hour=23, minute=59, second=59)
            from django.db.models import Q
            # Filter by either checkin or checkout being in the range, or currently in warehouse (similar to export logic)
            goods_list = goods_list.filter(
                Q(wh_check_in_out__in=[1, 4], wh_checkout_time__isnull=True) |
                Q(wh_check_in_out=2, wh_checkout_time__range=(dt_from, dt_to))
            )
        except Exception as e:
            print("Error parsing dates in stock_value_reports:", e)

    if request.GET.get('draw') or request.POST.get('draw'):
        from django.http import JsonResponse
        draw = int(request.GET.get('draw') or request.POST.get('draw') or 1)
        start = int(request.GET.get('start') or request.POST.get('start') or 0)
        length = int(request.GET.get('length') or request.POST.get('length') or 10)

        total_records = goods_list.count()
        paginated_list = goods_list[start:start+length]

        data = []
        for stock_value in paginated_list:
            dispatch = stock_value.wh_dispatch_id
            lb = stock_value.wh_lb_job_no_id

            dispatch = stock_value.wh_dispatch_id
            lb = stock_value.wh_lb_job_no_id
            gatein = stock_value.wh_gate_injob_no_id

            data.append([
                str(stock_value.wh_job_no if stock_value.wh_job_no else ''), # 0 Job Number
                str(stock_value.wh_qr_rand_num if stock_value.wh_qr_rand_num else ''), # 1 Stock Number
                str(stock_value.wh_customer_name if stock_value.wh_customer_name else ''), # 2 Customer
                gatein.gatein_arrival_date.strftime('%b %d, %Y, %I:%M %p') if gatein and gatein.gatein_arrival_date else '', # 3
                gatein.gatein_created_at.strftime('%b %d, %Y, %I:%M %p') if gatein and gatein.gatein_created_at else '', # 4
                lb.lb_stock_unloading_start_time.strftime('%b %d, %Y, %I:%M %p') if lb and lb.lb_stock_unloading_start_time else '', # 5
                lb.lb_stock_unloading_end_time.strftime('%b %d, %Y, %I:%M %p') if lb and lb.lb_stock_unloading_end_time else '', # 6
                str(gatein.gatein_transporter if gatein else ''), # 7 Transporter
                str(gatein.gatein_truck_number if gatein else ''), # 8 Truck Number
                str(gatein.gatein_truck_type if gatein else ''), # 9 Truck Type(In)
                str(stock_value.wh_truck_type if stock_value.wh_truck_type else ''), # 10 Truck Type Placed
                str(gatein.gatein_shipper if gatein else ''), # 11 Consignor
                str(gatein.gatein_consignee if gatein else ''), # 12 Consignee
                str(gatein.gatein_email_count if gatein else ''), # 13 Docs Received
                str(gatein.gatein_hawb if gatein else ''), # 14 HAWB
                str(gatein.gatein_destination if gatein else ''), # 15 Destination
                str(stock_value.wh_goods_invoice if stock_value.wh_goods_invoice else ''), # 16 Invoice Number
                str(stock_value.wh_voucher_num if stock_value.wh_voucher_num else ''), # 17 Case Number / Voucher Num
                str(stock_value.wh_invoice_qty if stock_value.wh_invoice_qty else ''), # 18 Invoice Qty
                str(stock_value.wh_invoice_weight_unit if stock_value.wh_invoice_weight_unit else ''), # 19 Invoice Weight
                str(stock_value.wh_gross_weight if stock_value.wh_gross_weight else ''), # 20 Checkin Weight
                str(stock_value.wh_uom if stock_value.wh_uom else ''), # 21 UOM
                str(stock_value.wh_goods_length if stock_value.wh_goods_length else ''), # 22 Length
                str(stock_value.wh_goods_width if stock_value.wh_goods_width else ''), # 23 Width
                str(stock_value.wh_goods_height if stock_value.wh_goods_height else ''), # 24 Height
                str(stock_value.wh_goods_pieces if stock_value.wh_goods_pieces else ''), # 25 Dims Qty
                str(stock_value.wh_goods_package_type if stock_value.wh_goods_package_type else ''), # 26 Package Type
                str(stock_value.wh_goods_volume_weight if stock_value.wh_goods_volume_weight else ''), # 27 Volume Weight
                str(stock_value.wh_cbm if stock_value.wh_cbm else ''), # 28 CBM
                str(stock_value.wh_invoice_value if stock_value.wh_invoice_value else ''), # 29 Invoice Value
                str(lb.lb_stock_invoice_currency if lb else ''), # 30 Invoice Currency
                str(stock_value.wh_invoice_amount_inr if stock_value.wh_invoice_amount_inr else ''), # 31 Invoice (INR)
                str(lb.lb_eway_bill if lb else ''), # 32 E-Way Bill
                lb.lb_validity_date.strftime('%b %d, %Y, %I:%M %p') if lb and lb.lb_validity_date else '', # 33
                str(stock_value.wh_fumigation_process if stock_value.wh_fumigation_process else ''), # 34 Fumigation
                str(stock_value.wh_check_in_out if stock_value.wh_check_in_out else ''), # 35 Check In-Out
                str(stock_value.wh_branch if stock_value.wh_branch else ''), # 36 Branch
                str(stock_value.wh_unit if stock_value.wh_unit else ''), # 37 Unit
                str(stock_value.wh_bay if stock_value.wh_bay else ''), # 38 Bay
                str((datetime.now() - stock_value.wh_checkin_time.replace(tzinfo=None)).days) if stock_value.wh_checkin_time else str(stock_value.wh_storage_time if stock_value.wh_storage_time else ''), # 39 Storage Days
                str(dispatch.dispatch_truck_number if dispatch else ''), # 40 Truck Number Out
                str(dispatch.dispatch_truck_type if dispatch else ''), # 41 Truck Type Out
                dispatch.dispatch_gatein_time.strftime('%d-%b-%Y %H:%M:%S') if dispatch and dispatch.dispatch_gatein_time else '', # 42
                dispatch.dispatch_dockin_time.strftime('%d-%b-%Y %H:%M:%S') if dispatch and dispatch.dispatch_dockin_time else '', # 43
                dispatch.dispatch_dockout_time.strftime('%d-%b-%Y %H:%M:%S') if dispatch and dispatch.dispatch_dockout_time else '', # 44
                dispatch.dispatch_depature_date.strftime('%d-%b-%Y %H:%M:%S') if dispatch and dispatch.dispatch_depature_date else '', # 45
                str(dispatch.dispatch_sticker_pasted_bvm if dispatch else ''), # 46 Labels
                str(dispatch.dispatch_mawb if dispatch else ''), # 47 MAWB
                str(dispatch.dispatch_num if dispatch else ''), # 48 Dispatch Number
                str(dispatch.dispatch_total_goods if dispatch else ''), # 49 Dispatch Qty
                str(stock_value.wh_total_qty if stock_value.wh_total_qty else '') # 50 Stock on hand
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })

    page_number = request.GET.get('page')
    paginator = Paginator(goods_list, 50)
    page_obj = paginator.get_page(page_number)


    current_date = datetime.today().date()
    # Base query for summaries
    base_qs = Warehouse_goods_info.objects.filter(wh_checkin_time__lte=current_date)
    if customer_name:
        base_qs = base_qs.filter(wh_customer_name=customer_name)

    # Optimized 1-query aggregation
    from django.db.models import Sum, Q
    aggs = base_qs.aggregate(
        maa_in=Sum('wh_invoice_amount_inr', filter=Q(wh_branch=2, wh_check_in_out=1)),
        blr_in=Sum('wh_invoice_amount_inr', filter=Q(wh_branch=1, wh_check_in_out=1))
    )
    maa_in_stock_value_cud_val = aggs.get('maa_in') or 0
    blr_in_stock_value_cud_val = aggs.get('blr_in') or 0

    maa_out_stock_value_cud_val = 0
    maa_total_cud_val = 0
    blr_out_stock_value_cud_val = 0
    blr_total_cud_val = 0
    hyd_in_stock_value_cud_val = 0
    hyd_out_stock_value_cud_val = 0
    hyd_total_cud_val = 0
    pny_in_stock_value_cud_val = 0
    pny_out_stock_value_cud_val = 0
    pny_total_cud_val = 0
    context = {
                'stock_value_list': Loadingbay_Info.objects.all(),
                'first_name': first_name,
                'form': form,
                'customer_name': customer_name,
                'checkin_goods_list': checkin_goods_list,
                'page_obj': page_obj,
                'maa_in_stock_value_cud': round(maa_in_stock_value_cud_val,0),
                'maa_out_stock_value_cud': round(maa_out_stock_value_cud_val,0),
                'maa_total_cud': round(maa_total_cud_val,0),
                'blr_in_stock_value_cud': round(blr_in_stock_value_cud_val, 0),
                'blr_out_stock_value_cud': round(blr_out_stock_value_cud_val, 0),
                'blr_total_cud': round(blr_total_cud_val, 0),
                'hyd_in_stock_value_cud': round(hyd_in_stock_value_cud_val, 0),
                'hyd_out_stock_value_cud': round(hyd_out_stock_value_cud_val, 0),
                'hyd_total_cud': round(hyd_total_cud_val, 0),
                'pny_in_stock_value_cud': round(pny_in_stock_value_cud_val, 0),
                'pny_out_stock_value_cud': round(pny_out_stock_value_cud_val, 0),
                'pny_total_cud': round(pny_total_cud_val, 0),
                 }
    return render(request,"asset_mgt_app/stock_values_report.html",context)
@login_required(login_url='login_page')
def damage_reports_list(request):
    first_name = request.session.get('first_name')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    selected_branch = request.GET.get('branch')
    branches = Location_info.objects.all()

    damage_reports_qs = DamagereportInfo.objects.exclude(dam_damage_type=6).order_by('-id')
    
    # Find matching job numbers
    if selected_branch:
        # Warehouse_goods_info has the branch link
        branch_jobs = Warehouse_goods_info.objects.filter(wh_branch__loc_name=selected_branch).values_list('wh_job_no', flat=True).distinct()
        damage_reports_qs = damage_reports_qs.filter(dam_wh_job_num__in=list(branch_jobs))
        
    if from_date or to_date:
        # Gatein_info is used for arrival dates
        date_filter = {}
        if from_date:
            date_filter['gatein_arrival_date__date__gte'] = from_date
        if to_date:
            date_filter['gatein_arrival_date__date__lte'] = to_date
        matching_date_jobs = Gatein_info.objects.filter(**date_filter).values_list('gatein_job_no', flat=True)
        damage_reports_qs = damage_reports_qs.filter(dam_wh_job_num__in=list(matching_date_jobs))

    if request.GET.get('draw'):
        from django.http import JsonResponse
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))

        total_records = damage_reports_qs.count()
        paginated_qs = damage_reports_qs.select_related('dam_damage_type')[start:start+length]

        data = []
        # We only execute the complex lookups for the 10 items currently visible on the page
        for damage in paginated_qs:
            job_no = damage.dam_wh_job_num.strip() if damage.dam_wh_job_num else ""

            gatein = Gatein_info.objects.filter(gatein_job_no=job_no).select_related('gatein_customer').first()
            if not gatein:
                gatein = Gatein_info.objects.filter(gatein_job_no__icontains=job_no).select_related('gatein_customer').first()

            goods = Warehouse_goods_info.objects.filter(wh_job_no=job_no).select_related('wh_customer_name', 'wh_gate_injob_no_id', 'wh_branch').first()
            if not goods:
                goods = Warehouse_goods_info.objects.filter(wh_job_no__icontains=job_no).select_related('wh_customer_name', 'wh_gate_injob_no_id', 'wh_branch').first()

            if goods and not gatein:
                gatein = goods.wh_gate_injob_no_id
            if gatein and not goods:
                 goods = Warehouse_goods_info.objects.filter(wh_gate_injob_no_id=gatein).first()

            checkin_date = None
            if gatein and gatein.gatein_arrival_date:
                checkin_date = gatein.gatein_arrival_date
            elif goods and goods.wh_checkin_time:
                checkin_date = goods.wh_checkin_time

            customer_name = "-"
            if gatein and gatein.gatein_customer:
                customer_name = gatein.gatein_customer.cu_name
            elif goods and goods.wh_customer_name:
                customer_name = goods.wh_customer_name.cu_name

            invoice_no = "-"
            if gatein and gatein.gatein_invoice:
                invoice_no = gatein.gatein_invoice
            elif goods and goods.wh_goods_invoice:
                invoice_no = goods.wh_goods_invoice

            total_pcs = "-"
            if goods and goods.wh_total_qty:
                total_pcs = goods.wh_total_qty
            elif gatein:
                total_pcs = gatein.gatein_actual_count or gatein.gatein_no_of_pkg or "-"

            data.append([
                checkin_date.strftime('%b %d, %Y, %I:%M %p') if checkin_date else '',
                str(damage.dam_wh_job_num if damage.dam_wh_job_num else ''),
                str(customer_name),
                str(invoice_no),
                str(total_pcs),
                str(damage.dam_damage_type if damage.dam_damage_type else ''),
                str(damage.dam_no_of_pcs_damaged if damage.dam_no_of_pcs_damaged is not None else '0')
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })

    context = {
        'first_name': first_name,
        'branches': branches,
        'selected_branch': selected_branch,
        'from_date': from_date,
        'to_date': to_date,
    }
    return render(request, "asset_mgt_app/damage_report.html", context)

@login_required(login_url='login_page')
def deviation_report(request):
    first_name = request.session.get('first_name')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    selected_branch = request.GET.get('branch')
    branches = Location_info.objects.all()

    deviation_qs = Warehouse_goods_info.objects.filter(
        wh_damage_check=1
    ).filter(
        Q(wh_weights_deviation=1) | 
        Q(wh_dimension_deviation=1) | 
        Q(wh_no_of_units_deviation=1) | 
        Q(wh_mismatches=1)
    ).select_related('wh_customer_name', 'wh_Dam_rep_job_num_id', 'wh_branch').order_by('-id')

    if selected_branch:
        deviation_qs = deviation_qs.filter(wh_branch__loc_name=selected_branch)
    if from_date:
        deviation_qs = deviation_qs.filter(wh_checkin_time__date__gte=from_date)
    if to_date:
        deviation_qs = deviation_qs.filter(wh_checkin_time__date__lte=to_date)
    
    deviation_list = []
    for item in deviation_qs:
        job_no = item.wh_job_no
        
        # 1. Prioritize data from the linked Damage Report if it exists
        damage = item.wh_Dam_rep_job_num_id
        if not damage:
            # Fallback search by job number
            damage = DamagereportInfo.objects.filter(dam_wh_job_num=job_no).first()
            
        if damage:
            invoice_qty = damage.dam_invoice_qty or 0
            checkin_qty = damage.dam_checkin_qty or 0
            invoice_wgt = damage.dam_invoice_weight or 0.0
            checkin_wgt = damage.dam_checkin_weight or 0.0
        else:
            # 2. Fallback to Gate-In for Invoice values and aggregate Warehouse_goods for Check-in values
            try:
                gatein = Gatein_info.objects.filter(gatein_job_no=job_no).first()
                if gatein:
                    invoice_qty = gatein.gatein_no_of_pkg or 0
                    invoice_wgt = gatein.gatein_weight or 0.0
                else:
                    invoice_qty = item.wh_invoice_qty or 0
                    invoice_wgt = item.wh_invoice_weight_unit or 0.0
            except Exception:
                invoice_qty = item.wh_invoice_qty or 0
                invoice_wgt = item.wh_invoice_weight_unit or 0.0
            
            # Aggregate check-in values from ALL records belonging to this job
            checkin_stats = Warehouse_goods_info.objects.filter(wh_job_no=job_no).aggregate(
                total_pcs=Sum('wh_goods_pieces'),
                total_wgt=Sum('wh_goods_weight')
            )
            checkin_qty = checkin_stats['total_pcs'] or 0
            checkin_wgt = checkin_stats['total_wgt'] or 0.0
        
        diff_qty = invoice_qty - checkin_qty
        diff_wgt = invoice_wgt - checkin_wgt
        
        deviation_list.append({
            'wh_checkin_time': item.wh_checkin_time,
            'wh_job_no': item.wh_job_no,
            'customer_name': item.wh_customer_name.cu_name if item.wh_customer_name else "-",
            'wh_goods_invoice': item.wh_goods_invoice,
            'invoice_qty': invoice_qty,
            'checkin_qty': checkin_qty,
            'diff_qty': diff_qty,
            'invoice_wgt': invoice_wgt,
            'checkin_wgt': checkin_wgt,
            'diff_wgt': round(diff_wgt, 2)
        })

    if request.GET.get('draw'):
        from django.http import JsonResponse
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))

        total_records = len(deviation_list)
        paginated_list = deviation_list[start:start+length]

        data = []
        for item in paginated_list:
            data.append([
                str(item.get('checkin_date', '')),
                str(item.get('job_no', '')),
                str(item.get('customer_name', '')),
                str(item.get('invoice_no', '')),
                str(item.get('invoice_qty', '')),
                str(item.get('checkin_qty', '')),
                str(item.get('diff_qty', '')),
                str(item.get('invoice_wgt', '')),
                str(item.get('checkin_wgt', '')),
                str(item.get('diff_wgt', ''))
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })

    context = {
        'first_name': first_name,
        'branches': branches,
        'selected_branch': selected_branch,
        'from_date': from_date,
        'to_date': to_date,
    }
    return render(request, "asset_mgt_app/deviation_report.html", context)


@login_required(login_url='login_page')
def revenue_report(request):
    from ..sub_models.billing_mod import BilingInfo

    first_name = request.session.get('first_name')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    selected_branch = request.GET.get('branch')
    branches = Location_info.objects.all()

    # --- ORIGINAL FLOW (unchanged) ---
    qs = Warehouse_goods_info.objects.exclude(wh_voucher_num__isnull=True)

    if from_date:
        qs = qs.filter(wh_checkout_time__date__gte=from_date)
    if to_date:
        qs = qs.filter(wh_checkout_time__date__lte=to_date)
    if selected_branch:
        qs = qs.filter(wh_branch__loc_name=selected_branch)

    revenue_summary = (
        qs.annotate(month=TruncMonth('wh_checkout_time')).values("wh_customer_name__cu_name", "wh_branch__loc_name", "wh_unit__unit_name","month")
        .annotate(
            storage_total=Sum("wh_storage_cost_total"),
            loading_total=Sum("wh_total_loading_cost"),
            forklift_total=Sum("wh_forklift_cost"),
            crane_total=Sum("wh_crane_cost"),
            unloading_total=Sum("wh_unloading_cost"),
            handling_total=Sum("wh_handling_cost"),
            fumigation_total=Sum("wh_fumigation_cost"),
            packing_total=Sum("wh_packing_cost"),
            total_invoice=Sum("wh_total_invoice_cost"),
        ).order_by("wh_customer_name__cu_name", "wh_branch__loc_name", "wh_unit__unit_name","month"))

    final_summary = []
    for row in revenue_summary:
        row_dict = dict(row)
        
        # Original costs
        row_dict['storage_total']    = float(row_dict.get('storage_total') or 0.0)
        row_dict['loading_total']    = float(row_dict.get('loading_total') or 0.0)
        row_dict['forklift_total']   = float(row_dict.get('forklift_total') or 0.0)
        row_dict['crane_total']      = float(row_dict.get('crane_total') or 0.0)

        # Aggregated charges (Mapped to bill_* for template compatibility)
        row_dict['bill_unloading']  = float(row_dict.get('unloading_total') or 0.0)
        row_dict['bill_handling']   = float(row_dict.get('handling_total') or 0.0)
        row_dict['bill_fumigation'] = float(row_dict.get('fumigation_total') or 0.0)
        row_dict['bill_packing']    = float(row_dict.get('packing_total') or 0.0)

        # Revenue = total of exactly 8 charges aggregated from goods level
        row_dict['revenue'] = (
            row_dict['storage_total'] +
            row_dict['loading_total'] +
            row_dict['forklift_total'] +
            row_dict['crane_total'] +
            row_dict['bill_unloading'] +
            row_dict['bill_handling'] +
            row_dict['bill_fumigation'] +
            row_dict['bill_packing']
        )
        final_summary.append(row_dict)

    # Chart: customer-wise total revenue (sum of row revenue)
    customer_revenue_dict = {}
    for row in final_summary:
        cust = row['wh_customer_name__cu_name']
        customer_revenue_dict[cust] = customer_revenue_dict.get(cust, 0) + row['revenue']

    chart_labels = list(customer_revenue_dict.keys())
    chart_data = [float(v) / 100000 for v in customer_revenue_dict.values()]  # in Lakhs

    context = {
        "revenue_summary": final_summary,
        "first_name": first_name,
        "from_date": from_date,
        "to_date": to_date,
        "chart_labels": json.dumps(chart_labels),
        "chart_data": json.dumps(chart_data),
        "branches":branches,
        "selected_branch":selected_branch
    }
    return render(request, "asset_mgt_app/revenue_report.html", context)


@login_required(login_url='login_page')
def profit_loss_report(request):
    first_name = request.session.get('first_name')
    warehouse_data = Warehouse_goods_info.objects.exclude(wh_voucher_num__isnull=True)
    expense_data = ExpenseExtinfo.objects.all()

    branch_unit_results = {}

    for warehouse in warehouse_data:
        branch_unit_key = (warehouse.wh_branch, warehouse.wh_unit)
        revenue = float(warehouse.wh_total_invoice_cost or 0)
        total_expense = 0.0
        for expense in expense_data:
            if expense.exp_ext_branch == warehouse.wh_branch and expense.exp_ext_unit == warehouse.wh_unit:
                total_expense += float(expense.exp_ext_amount or 0)
        if branch_unit_key not in branch_unit_results:
            branch_unit_results[branch_unit_key] = {'revenue': 0.0, 'expense': 0.0}
        branch_unit_results[branch_unit_key]['revenue'] += revenue
        branch_unit_results[branch_unit_key]['expense'] += total_expense

    result_list = []
    for (branch, unit), totals in branch_unit_results.items():
        revenue = totals['revenue']
        expense = totals['expense']
        profit_loss = revenue - expense
        profit_loss_percentage = (profit_loss / expense) * 100 if expense > 0 else 0.0

        result_list.append({
            'branch': branch,
            'unit': unit,
            'date': warehouse.wh_checkin_time,
            'revenue': round(revenue, 2),
            'expense': round(expense, 2),
            'profit_loss': round(profit_loss, 2),
            'profit_loss_percentage': round(profit_loss_percentage, 2),
        })

    context = {
        'result_list': result_list,
        'first_name': first_name,
    }

    return render(request, "asset_mgt_app/profit_loss_report.html", context)


@login_required(login_url='login_page')
def expense_report(request):
    first_name = request.session.get('first_name')
    expense_list=ExpenseInfo.objects.all()
    context = {
                'expense_list': expense_list,
                'first_name': first_name,
                }
    return render(request,"asset_mgt_app/expense_report.html",context)
@login_required(login_url='login_page')
def damage_report_pdf(request):
    wh_job_id = request.session.get('ses_gatein_id_nam')
    damage_list = DamagereportInfo.objects.filter(dam_wh_job_num=wh_job_id).first()

    damage_names = ", ".join(damage_list.dam_damages1.values_list('damage_name', flat=True))
    deviation_names = ", ".join(damage_list.dam_deviation1.values_list('deviation_name', flat=True))

    # First, try to find goods linked by FK
    warehouse_goods_list = Warehouse_goods_info.objects.filter(
        wh_Dam_rep_job_num_id=damage_list.id
    )

    # If none found, fall back to job number match
    if not warehouse_goods_list.exists():
        warehouse_goods_list = Warehouse_goods_info.objects.filter(
            wh_job_no=damage_list.dam_wh_job_num
        )

    context = {
        'damage_list': damage_list,
        'damage_names': damage_names,
        'deviation_names': deviation_names,
        'warehouse_goods_list': warehouse_goods_list,
    }

    file_name = f"Damage_Report_{wh_job_id}.pdf"
    template_path = 'asset_mgt_app/damage_report_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={file_name}'

    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF: <pre>' + html + '</pre>')
    return response




def export_stockreport_to_csv(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    customer_id = request.GET.get('ds_customer')
    if from_date and to_date:
        from_date = timezone.make_aware(datetime.strptime(from_date, "%Y-%m-%d"))
        to_date = timezone.make_aware(datetime.strptime(to_date, "%Y-%m-%d"))

    else:
        to_date = timezone.now()
        from_date = to_date - timedelta(days=7)

    # Base Query data
    base_qs = Warehouse_goods_info.objects.filter(
        Q(wh_check_in_out__in=[1, 4], wh_checkout_time__isnull=True) |
        Q(wh_check_in_out=2, wh_checkout_time__range=(from_date, to_date))
    )

    if customer_id:
        base_qs = base_qs.filter(wh_customer_name=customer_id)

    data = base_qs.annotate(
        arrival_date=ExpressionWrapper(F('wh_gate_injob_no_id__gatein_arrival_date'),
                                       output_field=fields.DateTimeField()),
        unloading_start_time=ExpressionWrapper(F('wh_lb_job_no_id__lb_stock_unloading_start_time'),
                                               output_field=fields.DateTimeField()),
        unloading_end_time=ExpressionWrapper(F('wh_lb_job_no_id__lb_stock_unloading_end_time'),
                                             output_field=fields.DateTimeField()),
        eway_bill_validity=ExpressionWrapper(F('wh_lb_job_no_id__lb_validity_date'),
                                             output_field=fields.DateTimeField()),
        departure_time=ExpressionWrapper(F('wh_dispatch_id__dispatch_depature_date'),
                                         output_field=fields.DateTimeField()),
    ).order_by('-arrival_date').values_list(
        'id','wh_job_no', 'wh_qr_rand_num', 'wh_customer_name__cu_name',
        'arrival_date', 'unloading_start_time', 'unloading_end_time',
        'wh_gate_injob_no_id__gatein_transporter',
        'wh_gate_injob_no_id__gatein_truck_number',
        'wh_consigner', 'wh_consignee', 'wh_lb_job_no_id__lb_packing_list__ge_gstexcepmtion',
        'wh_gate_injob_no_id__gatein_hawb', 'wh_gate_injob_no_id__gatein_destination',
        'wh_gate_injob_no_id__gatein_invoice', 'wh_po_num', 'wh_total_qty',
        'wh_invoice_weight_unit','wh_gross_weight', 'wh_uom__uom_name', 'wh_goods_length',
        'wh_goods_width', 'wh_goods_height', 'wh_goods_pieces',
        'wh_goods_package_type__package_type', 'wh_chargeable_weight', 'wh_cbm', 'wh_invoice_value',
        'wh_lb_job_no_id__lb_stock_invoice_currency__currency_type', 'wh_invoice_amount_inr',
        'wh_lb_job_no_id__lb_eway_bill', 'eway_bill_validity',
        'wh_fumigation_process__ge_gstexcepmtion', 'wh_check_in_out__check_in_out_name', 'wh_branch__loc_name',
        'wh_unit__unit_name', 'wh_bay__bay_bayname', 'wh_storage_time','wh_comments','wh_damages__damage_name','wh_weights_deviation_id','wh_dimension_deviation_id','wh_no_of_units_deviation_id','wh_damages_id',
        'wh_Dam_rep_job_num_id__dam_GRN_num','wh_gate_injob_no_id__gatein_truck_number_n__pregatein_truck_type__vt_vehicletype',

        # 'wh_dispatch_id__dispatch_truck_number',
        # str('wh_dispatch_id__dispatch_truck_type__vt_vehicletype'),'departure_time',
        # 'wh_dispatch_id__dispatch_sticker_pasted_bvm__lp_name', 'wh_dispatch_id__dispatch_mawb',
        # 'wh_dispatch_id__dispatch_num',
    )

    # Header row
    headers = [
         'Job Number', 'Stock Number', 'Customer', 'Date Of Arrival',
            'Unloading Start Time', 'Unloading End Time', 'Transporter',
            'Truck Number', 'Truck Type(In)', 'Truck Type Placed', 'Consignor', 'Consignee', 'Docs Received', 'HAWB',
            'Destination', 'Invoice Number', 'Case Number', 'Invoice Qty',
            'Invoice Weight (kg)', 'Checkin Weight (kg)', 'UOM', 'Length',
            'Width', 'Height', 'Dims Qty', 'Package Type', 'Volume Weight',
            'CBM', 'Invoice Value', 'Invoice Currency', 'Invoice (INR)',
            'E-Way Bill#', 'E-Way Bill Validity', 'Fumigation Status',
            'Check In-Out?', 'Branch', 'Unit', 'Bay', 'Storage Days','Remarks','Damage Type','GRN Number',
            'Truck_Number(Out)', 'Truck_Type(Out)', 'Gatein Time(Out)', 'Dockin Time(Out)', 'Dockout Time(Out)', 'Truck_Depature_Time(Out)',
            'Labels_Pasted_By', 'MAWB', 'Dispatch Number(s)', 'Total Dispatch Qty'
    ]

    def generate_streamed_excel(red_font=None):
        # Create an in-memory buffer
        output = BytesIO()
        from openpyxl.cell import WriteOnlyCell
        workbook = openpyxl.Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Stock Report")

        # Write headers
        header_row = []
        for header in headers:
            cell = WriteOnlyCell(sheet, value=header)
            cell.font = Font(name='Bookman Old Style', size=10, bold=True, color="000000")
            cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            header_row.append(cell)
        sheet.append(header_row)

        # PREFETCH ALL PARTIALS IN 1 QUERY TO PREVENT 50,000 DB QUERIES!
        from collections import defaultdict

        # Evaluate data (tuple queryset) into a list to prevent multiple executions
        evaluated_data = list(data)
        goods_ids = [row[0] for row in evaluated_data]

        partials_by_goods = defaultdict(list)

        # Fetch all partials for these goods in ONE massive query
        # Split into chunks of 5000 if needed, but SQL IN clause can easily handle 50k on Postgres
        # We will chunk it safely to avoid SQLite max variables limit (999) or generic DB limits
        chunk_size = 1000
        for i in range(0, len(goods_ids), chunk_size):
            chunk_ids = goods_ids[i:i+chunk_size]
            all_partials = GoodsPartialDispatchInfo.objects.filter(pd_goods_id__in=chunk_ids).select_related(
                'pd_dispatch_info__dispatch_truck_type',
                'pd_dispatch_info__dispatch_sticker_pasted_bvm'
            )
            for partial in all_partials:
                partials_by_goods[partial.pd_goods_id].append(partial)

        # Write data rows
        for row_num, row_data in enumerate(evaluated_data, 2):
            goods_id = row_data[0]  # first element is ID
            truck_type_in = row_data[-1]
            row_data = list(row_data[:-1]) # remove truck_type_in for now to restore later or insert correctly
            try:
                partials = partials_by_goods.get(goods_id, [])

                dispatch_nums = []
                truck_numbers = []
                truck_types = []
                truck_types_placed = []
                departure_times = []
                gatein_times_out = []
                dockin_times_out = []
                dockout_times_out = []
                sticker_pasted_bys = []
                mawb_list = []

                total_qty = 0

                for partial in partials:
                    dispatch = partial.pd_dispatch_info
                    if dispatch:
                        dispatch_nums.append(dispatch.dispatch_num or "")
                        truck_numbers.append(dispatch.dispatch_truck_number or "")
                        truck_types.append(getattr(dispatch.dispatch_truck_type, 'vt_vehicletype', "") or "")
                        truck_types_placed.append(dispatch.dispatch_truck_type_placed or "")
                        gatein_times_out.append(dispatch.dispatch_gatein_time.strftime("%d-%b-%y %H:%M:%S") if dispatch.dispatch_gatein_time else "")
                        dockin_times_out.append(dispatch.dispatch_dockin_time.strftime("%d-%b-%y %H:%M:%S") if dispatch.dispatch_dockin_time else "")
                        dockout_times_out.append(dispatch.dispatch_dockout_time.strftime("%d-%b-%y %H:%M:%S") if dispatch.dispatch_dockout_time else "")
                        departure_times.append(
                            dispatch.dispatch_depature_date.strftime(
                                "%d-%b-%y") if dispatch.dispatch_depature_date else ""
                        )
                        sticker_pasted_bys.append(getattr(dispatch.dispatch_sticker_pasted_bvm, 'lp_name', "") or "")
                        mawb_list.append(dispatch.dispatch_mawb or "")

                    total_qty += partial.pd_dispatch_qty or 0

                row_data = list(row_data[1:])  # remove ID
                row_data.insert(8, truck_type_in)
                row_data.insert(9, ", ".join(truck_types_placed))

                weights_dev_id = row_data[-4]
                dim_dev_id = row_data[-3]
                units_dev_id = row_data[-2]
                damage_id = row_data[-1]
                remarks = row_data[headers.index("Remarks") - 1] or "" # Approximate index based on previous logic
                remarks_index = headers.index("Remarks")

                # Blank remarks by default
                row_data[remarks_index] = ""
                damage_type = row_data[-6]
                grn_number = row_data[-5]

                # We need to slice off the 4 extra fields we added
                row_data = list(row_data[:-4])

                if (
                        weights_dev_id == 1 or
                        dim_dev_id == 1 or
                        units_dev_id == 1 or
                        (damage_id and damage_id != 6)
                ):
                    is_damaged = True
                    row_data[remarks_index] = remarks  # show remarks only when flagged
                else:
                    is_damaged = False

                if damage_id == 6:  # Nil Damage
                    damage_type = ""
                    grn_number = ""

                # overwrite cleaned values
                row_data[-2] = damage_type
                row_data[-1] = grn_number

                row_data += [
                    ", ".join(truck_numbers),
                    ", ".join(truck_types),
                    ", ".join(gatein_times_out),
                    ", ".join(dockin_times_out),
                    ", ".join(dockout_times_out),
                    ", ".join(departure_times),
                    ", ".join(sticker_pasted_bys),
                    ", ".join(mawb_list),
                    ", ".join(dispatch_nums),
                    total_qty
                ]

            except Exception as e:
                row_data = list(row_data[1:])
                row_data += ["", "", "", "", "", "", 0]

            red_font = Font(name='Bookman Old Style', size=9, color="FF0000")
            row_cells = []
            for value in row_data:
                from datetime import date
                if isinstance(value, (date, datetime)) and hasattr(value, 'tzinfo') and value.tzinfo:
                    value = make_naive(value)
                cell = WriteOnlyCell(sheet, value=value)
                cell.font = red_font if is_damaged else Font(name='Bookman Old Style', size=9)
                row_cells.append(cell)
            sheet.append(row_cells)



        # Save workbook to buffer
        workbook.save(output)
        output.seek(0)  # Reset buffer pointer
        yield output.read()  # Yield content of buffer
        output.close()

    # Return streaming response
    response = StreamingHttpResponse(
        generate_streamed_excel(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Get the current time in UTC and then convert to IST (UTC + 5:30)
    ist_time = timezone.now() + timedelta(hours=5, minutes=30)

    # Format the time with an underscore between the date and time
    filename = f'Stock_Report_{ist_time.strftime("%Y%m%d_%H%M")}.xlsx'

    # Set the Content-Disposition header with the formatted filename
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required(login_url='login_page')
def goods_in_out_reports_list(request):
    first_name = request.session.get('first_name')

    # Grouping by branch, unit, and date for check-in
    in_statistics = Warehouse_goods_info.objects.filter(wh_check_in_out=1).values(
        'wh_branch__loc_name',
        'wh_unit__unit_name',
        'wh_gate_injob_no_id__gatein_created_at__date'  # Truncate to date for date-wise grouping
    ).annotate(
        total_invoices=Count('wh_goods_invoice', distinct=True),
        total_trucks=Count('wh_gate_injob_no_id__gatein_pre_id', distinct=True),
        total_weights=(Sum('wh_goods_weight')/1000)
    )

    # Grouping by branch, unit, and date for check-out
    out_statistics = Warehouse_goods_info.objects.filter(wh_check_in_out=2).values(
        'wh_branch__loc_name',
        'wh_unit__unit_name',
        'wh_dispatch_id__dispatch_depature_date__date',  # Truncate to date for date-wise grouping
    ).annotate(
        total_invoices=Count('wh_goods_invoice', distinct=True),
        total_trucks=Count('wh_dispatch_num', distinct=True),
        total_weights=(Sum('wh_goods_weight') / 1000)
    )

    context = {
        'first_name': first_name,
        'in_statistics': in_statistics.order_by('wh_gate_injob_no_id__gatein_created_at__date', 'id'),
        'out_statistics': out_statistics.order_by('wh_dispatch_id__dispatch_depature_date__date', 'id'),
    }
    return render(request, "asset_mgt_app/goods_in_out_reports_list.html", context)

@login_required(login_url='login_page')
def stock_value_send_email_view(request,pre_gatein_id=None,customer_name=None,subject=None):
    from_date_str = request.POST.get('from_date') or request.GET.get('from_date')
    to_date_str = request.POST.get('to_date') or request.GET.get('to_date')

    if from_date_str and to_date_str:
        # Parse directly into naive datetime objects
        from_date = make_aware(datetime.strptime(from_date_str, "%Y-%m-%d"))
        to_date = make_aware(datetime.strptime(to_date_str, "%Y-%m-%d")) + timedelta(days=1)
    else:
        # Default to last 120 days
        to_date = now()
        from_date = to_date - timedelta(days=7)

    print('Entering stcokvalue_send_email_view')
    if request.method == 'POST':
        recipient = request.POST.get('recipient')
        # subject = request.POST.get('subject')
        message = request.POST.get('message')
        customer_name_1=customer_name
        if customer_name_1==None:
            customer_name = request.POST.get('ds_customer')
            customer_id=False
        else:
            customer_name=customer_name
            customer_id = True
        recipient_list = [email.strip() for email in recipient.split(',')]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Goods Movement Report"

        # Write the headers
        headers = [
            'Job Number', 'Stock Number', 'Customer', 'Date Of Arrival', 'Dock In Time','Unloading Start Time',
            'Unloading End Time', 'Transporter', 'Truck Number', 'Truck Type(In)', 'Consignor', 'Consignee',
            'Docs Received', 'HAWB', 'Destination', 'Invoice Number', 'Case Number',
            'Invoice Qty', 'Invoice Weight (kg)', 'Checkin Weight (kg)', 'UOM', 'Length',
            'Width', 'Height', 'Dims Qty', 'Package Type', 'Volume Weight', 'CBM',
            'Invoice Value', 'Invoice Currency', 'Invoice (INR)', 'E-Way Bill#', 'E-Way Bill Validity',
            'Fumigation Status', 'Check In-Out?', 'Branch', 'Unit', 'Bay', 'Storage Days','Damage/Deviation?','GRN Number','Damages','Deviations','Remarks',
            'Truck_Number(Out)','Truck_Type(Out)','Truck Type Placed','Gatein Time(Out)','Dockin Time(Out)','Dockout Time(Out)','Truck_Depature_Time(Out)','Labels_Pasted_By',
            'MAWB','Dispatch_Number','Dispatch quantity','Stock On Hand'
        ]
        ws.append(headers)
        header_font = Font(name="Arial", bold=True, size=11, color="000000")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = yellow_fill
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Fetch the IDs from Gatein_info
        gate_in_ids = Gatein_info.objects.filter(gatein_pre_id=pre_gatein_id).values_list('id', flat=True)

        # three_months_ago = timezone.now() - timedelta(days=90)

        # Initialize query
        stock_values = Warehouse_goods_info.objects.all()

        # Build query conditions dynamically
        if customer_name and gate_in_ids.exists():
            print ("Inside first loop")
            stock_values = stock_values.filter(
                wh_customer_name=customer_name,
                wh_gate_injob_no_id__in=list(gate_in_ids)
            )
        elif customer_name:
            print ("Inside second loop")
            stock_values = stock_values.filter(
                wh_customer_name=customer_name
            )
        elif gate_in_ids.exists():
            print ("Inside third loop")
            stock_values = stock_values.filter(
                wh_gate_injob_no_id__in=list(gate_in_ids)
            )
        # If no customer selected and no gate_in_ids, it takes all customers by default
        
        from collections import defaultdict

        # Step 1: Filter stock_values and ensure the date fields are timezone-free
        stock_values = stock_values.filter(
            Q(wh_check_in_out__in=[1, 4]) |
            Q(wh_check_in_out=2, wh_checkout_time__isnull=False,wh_checkout_time__range=(from_date, to_date))  #wh_checkout_time__gte=three_months_ago
        ).order_by('-wh_gate_injob_no_id__gatein_arrival_date').select_related(
            'wh_gate_injob_no_id',
            'wh_gate_injob_no_id__gatein_truck_number_n',
            'wh_gate_injob_no_id__gatein_truck_number_n__pregatein_truck_type',
            'wh_lb_job_no_id',
            'wh_lb_job_no_id__lb_stock_invoice_currency',
            'wh_customer_name',
            'wh_uom',
            'wh_goods_package_type',
            'wh_fumigation_process',
            'wh_dispatch_id',
            'wh_check_in_out',
            'wh_branch',
            'wh_unit',
            'wh_bay'
        )

        job_numbers = set()
        stock_ids = set()
        for sv in stock_values:
            job_numbers.add(sv.wh_job_no)
            stock_ids.add(sv.id)

        damage_reports_qs = DamagereportInfo.objects.filter(
            dam_wh_job_num__in=job_numbers
        ).prefetch_related('dam_damages1', 'dam_deviation1')
        
        damage_reports_dict = {}
        for dr in damage_reports_qs:
            if dr.dam_wh_job_num not in damage_reports_dict:
                damage_reports_dict[dr.dam_wh_job_num] = dr

        partials_qs = GoodsPartialDispatchInfo.objects.filter(
            pd_goods_id__in=stock_ids
        ).select_related(
            'pd_dispatch_info',
            'pd_dispatch_info__dispatch_truck_type',
            'pd_dispatch_info__dispatch_sticker_pasted_bvm'
        )
        partials_dict = defaultdict(list)
        for partial in partials_qs:
            partials_dict[partial.pd_goods_id].append(partial)

        # Step 2: In the loop, replace the date with timezone-free date objects
        for stock_value in stock_values:
            date_of_arrival = None  # Default value

            if stock_value.wh_gate_injob_no_id:  # Check if exists
                date_of_arrival = getattr(stock_value.wh_gate_injob_no_id, 'gatein_arrival_date', None)
                if date_of_arrival:
                    date_of_arrival = date_of_arrival.replace(tzinfo=None)
                else:
                    date_of_arrival = ""

            checkin_qty = stock_value.wh_goods_pieces if stock_value.wh_goods_pieces else 0
            # Fetch partial dispatches for this stock
            partials = partials_dict.get(stock_value.id, [])
            dispatch_nums = []
            truck_numbers = []
            truck_types = []
            truck_types_placed = []
            gatein_times_out = []
            dockin_times_out = []
            dockout_times_out = []
            departure_times = []
            sticker_pasted_bys = []
            mawb_list = []

            dispatch_qty = 0  # initialize
            damage_check_flag = stock_value.wh_damage_check_id == 1
            damage_report = damage_reports_dict.get(stock_value.wh_job_no)
            damage_names = ", ".join(
                [d.damage_name for d in damage_report.dam_damages1.all()]
            ) if damage_report else ""
            deviation_names = ", ".join(
                [d.deviation_name for d in damage_report.dam_deviation1.all()]
            ) if damage_report else ""
            grn_number = damage_report.dam_GRN_num if damage_report else ""
            remarks = damage_report.dam_comments if damage_report else ""
            for partial in partials:
                dispatch = partial.pd_dispatch_info
                if dispatch:
                    dispatch_nums.append(dispatch.dispatch_num or "")
                    truck_numbers.append(dispatch.dispatch_truck_number or "")

                    truck_type = getattr(dispatch.dispatch_truck_type, 'vt_vehicletype', "")
                    if truck_type:
                        truck_types.append(truck_type)

                    truck_type_pl = dispatch.dispatch_truck_type_placed or ""
                    if truck_type_pl:
                        truck_types_placed.append(truck_type_pl)

                    sticker = getattr(dispatch.dispatch_sticker_pasted_bvm, 'lp_name', "")
                    if sticker:
                        sticker_pasted_bys.append(sticker)

                    mawb_list.append(dispatch.dispatch_mawb or "")

                    if dispatch.dispatch_gatein_time:
                        gatein_times_out.append(dispatch.dispatch_gatein_time.strftime('%d-%b-%Y %H:%M:%S'))
                    if dispatch.dispatch_dockin_time:
                        dockin_times_out.append(dispatch.dispatch_dockin_time.strftime('%d-%b-%Y %H:%M:%S'))
                    if dispatch.dispatch_dockout_time:
                        dockout_times_out.append(dispatch.dispatch_dockout_time.strftime('%d-%b-%Y %H:%M:%S'))

                    if dispatch.dispatch_depature_date:
                        departure_times.append(dispatch.dispatch_depature_date.strftime('%d-%b-%Y %H:%M:%S'))

                dispatch_qty += partial.pd_dispatch_qty or 0  # Sum total qty

            stock_on_hand = checkin_qty - dispatch_qty  # Subtract dispatch quantity
            try:
                gatein_obj = stock_value.wh_gate_injob_no_id
                dock_in_time = None
                truck_type_in = ""
                if gatein_obj:
                    pregatein_truck_obj = gatein_obj.gatein_truck_number_n
                    if pregatein_truck_obj:
                        dock_in_time = pregatein_truck_obj.pregatein_dock_in_date_time
                        if dock_in_time:
                            dock_in_time = dock_in_time.replace(tzinfo=None)
                        
                        truck_type_obj = pregatein_truck_obj.pregatein_truck_type
                        if truck_type_obj:
                            truck_type_in = truck_type_obj.vt_vehicletype

                if stock_value.wh_dispatch_id and stock_value.wh_dispatch_id.dispatch_depature_date:
                    dispatch_depature_time = stock_value.wh_dispatch_id.dispatch_depature_date.replace(tzinfo=None)
                else:
                    dispatch_depature_time = None
            except AttributeError:
                dispatch_depature_time = None
                dock_in_time = None
                truck_type_in = ""


            row = [
                stock_value.wh_job_no,  # Index 0
                stock_value.wh_qr_rand_num,  # Index 1
                str(stock_value.wh_customer_name),  # Index 2
                date_of_arrival if date_of_arrival else '',  # Index 3: Only Date, no time
                dock_in_time if dock_in_time else '', # NEW: Dock In Time

                stock_value.wh_lb_job_no_id.lb_stock_unloading_start_time.replace(tzinfo=None)
                if stock_value.wh_lb_job_no_id and stock_value.wh_lb_job_no_id.lb_stock_unloading_start_time else '',

                stock_value.wh_lb_job_no_id.lb_stock_unloading_end_time.replace(tzinfo=None)
                if stock_value.wh_lb_job_no_id and stock_value.wh_lb_job_no_id.lb_stock_unloading_end_time else '',
                # Index 6: gatein_transporter
                getattr(stock_value.wh_gate_injob_no_id, 'gatein_transporter', ''),

                # Index 7: gatein_truck_number
                getattr(stock_value.wh_gate_injob_no_id, 'gatein_truck_number', ''),
                truck_type_in, # NEW: Truck Type (In)

                stock_value.wh_consigner,  # Index 8
                stock_value.wh_consignee,  # Index 9

                # Index 10: lb_packing_list
                str(getattr(stock_value.wh_lb_job_no_id, 'lb_packing_list', '')),

                # Index 11-13: gatein_hawb, gatein_destination, gatein_invoice
                getattr(stock_value.wh_gate_injob_no_id, 'gatein_hawb', ''),
                getattr(stock_value.wh_gate_injob_no_id, 'gatein_destination', ''),
                getattr(stock_value.wh_gate_injob_no_id, 'gatein_invoice', ''),

                stock_value.wh_po_num,  # Index 14
                stock_value.wh_total_qty,  # Index 15

                stock_value.wh_invoice_weight_unit,  # Index 16
                stock_value.wh_gross_weight,  # Index 17

                # Index 18: wh_uom
                str(stock_value.wh_uom),

                stock_value.wh_goods_length,  # Index 19
                stock_value.wh_goods_width,  # Index 20
                stock_value.wh_goods_height,  # Index 21
                stock_value.wh_goods_pieces,  # Index 22

                # Index 23: wh_goods_package_type
                str(stock_value.wh_goods_package_type),

                stock_value.wh_chargeable_weight,  # Index 24
                stock_value.wh_cbm,  # Index 25
                stock_value.wh_invoice_value,  # Index 26

                # Index 27: lb_stock_invoice_currency
                str(getattr(stock_value.wh_lb_job_no_id, 'lb_stock_invoice_currency', '')),

                stock_value.wh_invoice_amount_inr,  # Index 28

                # Index 29: lb_eway_bill
                getattr(stock_value.wh_lb_job_no_id, 'lb_eway_bill', ''),

                # Index 30: lb_validity_date (remove tzinfo)
                getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', None).replace(tzinfo=None)
                if getattr(stock_value.wh_lb_job_no_id, 'lb_validity_date', None) else None,

                # Index 31: wh_fumigation_process
                str(stock_value.wh_fumigation_process or ''),

                "Stock on Hand" if str(stock_value.wh_check_in_out) == "Checked-In" else "Checked-Out", # Index # Index 32
                str(stock_value.wh_branch),  # Index 33
                str(stock_value.wh_unit),  # Index 34
                str(stock_value.wh_bay),  # Index 35
                stock_value.wh_storage_time,# Index 36
                # Damage Info
                "Yes" if damage_check_flag else "No",  # 37: Damage/Deviation?
                grn_number,
                damage_names,# 38
                deviation_names,  # 39 merged damages+deviations
                remarks,  # 40
                # getattr(stock_value.wh_dispatch_id, 'dispatch_truck_number', ''),# Index 37
                # str(getattr(stock_value.wh_dispatch_id, 'dispatch_truck_type', '')),# Index 38
                # # getattr(stock_value.wh_dispatch_id, 'dispatch_depature_date', ''),# Index 39
                # dispatch_depature_time,
                # str(getattr(stock_value.wh_dispatch_id, 'dispatch_sticker_pasted_bvm', '')),# Index 40
                # getattr(stock_value.wh_dispatch_id, 'dispatch_mawb', ''),# Index 41
                # getattr(stock_value.wh_dispatch_id, 'dispatch_num', ''),# Index 42
                # getattr(stock_value.wh_dispatch_id, 'dispatch_total_goods', ''),# Index 43
                ", ".join(truck_numbers),
                ", ".join(truck_types),
                ", ".join(truck_types_placed), # Truck Type Placed (next to Truck_Type(Out))
                ", ".join(gatein_times_out),
                ", ".join(dockin_times_out),
                ", ".join(dockout_times_out),
                ", ".join(departure_times),
                ", ".join(sticker_pasted_bys),
                ", ".join(mawb_list),
                ", ".join(dispatch_nums),
                dispatch_qty,

                stock_on_hand,# Index 44
            ]

            # # Debugging the row values
            # for idx, value in enumerate(row):
            #     print(f"Index {idx}: Value={value}, Type={type(value)}")

            ws.append(row)  # Append the row to the worksheet

            # Check if this row has damage/deviation
            damage_flag = (
                    (stock_value.wh_damages_id and stock_value.wh_damages_id != 6) or
                    stock_value.wh_weights_deviation_id == 1 or
                    stock_value.wh_dimension_deviation_id == 1 or
                    stock_value.wh_no_of_units_deviation_id == 1
            )

            # Apply borders + font immediately for this row
            for cell in ws[ws.max_row]:
                cell.border = border_style
                if damage_check_flag:
                    cell.font = Font(name="Arial", bold=False, size=10, color="FF0000")  # Red font
                else:
                    cell.font = Font(name="Arial", bold=False, size=10, color="000000")

        sheet = wb.active

        # Format the first row (Header)

        if str(customer_name).isdigit():
            customer_name = CustomerInfo.objects.get(pk=int(customer_name)).cu_name
        else:
            customer_name = CustomerInfo.objects.filter(cu_name=customer_name).first()
        file_name = str(customer_name)+'_Goods_Movement_report.xlsx'  # Set your desired file name
        # Apply formatting to the first row
        # for cell in sheet[1]:
        #     cell.font = header_font
        #     cell.fill = yellow_fill
        #     cell.border = border_style
        #     cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply borders to the rest of the cells in the sheet, skipping the first row


        # Set column width for all columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name (e.g., 'A')
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width =adjusted_width  # Set column width to 20
        # Save the workbook to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        attachment = excel_file
        attachment_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        if subject==None:
            subject = f"{customer_name}Goods Movement Report"
        else:
            subject = subject
        pre_gatein_id = request.session.get('ses_pre_gatein_id')
        send_department_email('warehouse', subject, message, recipient_list,attachment,attachment_type,file_name)
        # Redirect back to the previous page
        messages.success(request, f"E-mail sent successfully")
        return redirect(request.META['HTTP_REFERER'])
    else:
        messages.error(request, 'Invalid input in the email form.')
    return redirect(request.META['HTTP_REFERER'])
    # return render(request, "asset_mgt_app/dsr_send_email.html", context)


@login_required(login_url='login_page')
def transport_reports(request):
    first_name = request.session.get('first_name')
    context = {
               'first_name': first_name
               }
    return render(request,"asset_mgt_app/trans_report.html",context)