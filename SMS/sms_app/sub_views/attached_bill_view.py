from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Sum, Q
import calendar
from datetime import date as dt_date, timedelta, datetime as dt_datetime
from django.utils import timezone as dj_timezone

from ..sub_forms.attached_bill_form import AttachedBillForm
from ..sub_models.attached_bill_mod import AttachedBillInfo
from ..sub_models.vehiclemaster_mod import VehiclemasterInfo
from ..sub_models.tripdetail_mod import TripdetailInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


# ==================================================
# ADD ATTACHED BILL
# ==================================================
@login_required(login_url='login_page')
def attached_bill_add(request):
    if request.method == "POST":
        form = AttachedBillForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ab_created_by = request.user
            obj.save()

            # Save per-trip costs from table
            selected_trips = request.POST.get('ab_selected_trips', '')
            if selected_trips:
                trip_numbers = [t.strip() for t in selected_trips.split(',') if t.strip()]
                # Find trips to get their IDs for input name matching
                trips_to_update = TripdetailInfo.objects.filter(tr_tripnumber__in=trip_numbers)
                for trip in trips_to_update:
                    tid = trip.id
                    t_buy = request.POST.get(f'trip_buy_cost_{tid}')
                    t_parking = request.POST.get(f'trip_parking_cost_{tid}')
                    t_toll = request.POST.get(f'trip_toll_cost_{tid}')

                    if t_parking is not None: trip.tc_parkingcost = float(t_parking or 0)
                    if t_toll is not None: trip.tc_tollcost = float(t_toll or 0)
                    trip.save()

            messages.success(request, "Attached Bill saved successfully.")
            return redirect('attached_bill_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AttachedBillForm()

    return render(
        request,
        "asset_mgt_app/attached_bill_add.html",
        {
            "form": form,
            "title": "Add Attached Bill"
        }
    )


# ==================================================
# LIST ATTACHED BILLS
# ==================================================
@login_required(login_url='login_page')
def attached_bill_list(request):
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    vendor_id = request.GET.get('vendor_id', '')

    bills = AttachedBillInfo.objects.all().order_by('-ab_created_at')
    if from_date:
        bills = bills.filter(ab_from_date__gte=from_date)
    if to_date:
        bills = bills.filter(ab_to_date__lte=to_date)
    if vendor_id:
        bills = bills.filter(ab_vendor_id=vendor_id)

    # Fetch only vendors who have at least one Attached vehicle (ownership_id=2)
    from sms_app.sub_models.vendor_info_mod import Vendor_info
    from sms_app.sub_models.vehiclemaster_mod import VehiclemasterInfo
    attached_vendor_ids = VehiclemasterInfo.objects.filter(vm_ownership_id=2).exclude(vm_vendor=None).values_list('vm_vendor_id', flat=True).distinct()
    vendors = Vendor_info.objects.filter(id__in=attached_vendor_ids).order_by('vend_name')

    return render(
        request,
        "asset_mgt_app/attached_bill_list.html",
        {
            "bills": bills,
            "from_date": from_date,
            "to_date": to_date,
            "vendor_id": vendor_id,
            "vendors": vendors,
        }
    )


# ==================================================
# EDIT ATTACHED BILL
# ==================================================
@login_required(login_url='login_page')
def attached_bill_edit(request, id):
    record = get_object_or_404(AttachedBillInfo, id=id)
    if request.method == "POST":
        form = AttachedBillForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ab_updated_by = request.user
            obj.save()

            # Save per-trip costs from table
            selected_trips = request.POST.get('ab_selected_trips', '')
            if selected_trips:
                trip_numbers = [t.strip() for t in selected_trips.split(',') if t.strip()]
                trips_to_update = TripdetailInfo.objects.filter(tr_tripnumber__in=trip_numbers)
                for trip in trips_to_update:
                    tid = trip.id
                    t_buy = request.POST.get(f'trip_buy_cost_{tid}')
                    t_parking = request.POST.get(f'trip_parking_cost_{tid}')
                    t_toll = request.POST.get(f'trip_toll_cost_{tid}')

                    if t_parking is not None: trip.tc_parkingcost = float(t_parking or 0)
                    if t_toll is not None: trip.tc_tollcost = float(t_toll or 0)
                    trip.save()

            messages.success(request, "Attached Bill updated successfully.")
            return redirect('attached_bill_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AttachedBillForm(instance=record)

    return render(
        request,
        "asset_mgt_app/attached_bill_add.html",
        {
            "form": form,
            "record": record,
            "title": "Edit Attached Bill"
        }
    )


# ==================================================
# DELETE ATTACHED BILL
# ==================================================
@login_required(login_url='login_page')
def attached_bill_delete(request, id):
    record = get_object_or_404(AttachedBillInfo, id=id)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Attached Bill deleted successfully.")
    return redirect('attached_bill_list')


# ==================================================
# QUICK MEDIA UPLOAD
# ==================================================
@login_required(login_url='login_page')
def attached_bill_upload(request, id):
    record = get_object_or_404(AttachedBillInfo, id=id)
    if request.method == "POST" and request.FILES.get('ab_bill_upload'):
        record.ab_bill_upload = request.FILES.get('ab_bill_upload')
        record.save()
        messages.success(request, "File uploaded successfully.")
    return redirect('attached_bill_list')


# ==================================================
# AJAX: GET VEHICLE DETAILS & KM RUN
# ==================================================
@login_required(login_url='login_page')
def get_attached_vehicle_details(request):
    vendor_id = request.GET.get('vendor_id')
    vehicle_id = request.GET.get('vehicle_id')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    data = {'trips': [], 'total_km_run': 0, 'leave_days': 0, 'vehicle_type': '', 'buy_cost': 0, 'agreed_km': 0, 'extra_km_rate': 0}

    def parse_dt(d_str):
        if not d_str: return None
        try: return dt_date.fromisoformat(d_str)
        except:
            try:
                p = d_str.split('-')
                if len(p) == 3: return dt_date(int(p[2]), int(p[1]), int(p[0]))
            except: pass
        return None

    f_dt_obj = parse_dt(from_date)
    t_dt_obj = parse_dt(to_date)

    try:
        if vehicle_id:
            try:
                vehicle = VehiclemasterInfo.objects.get(id=vehicle_id)
                data.update({
                    'vehicle_type': str(vehicle.vm_vehicletype) if vehicle.vm_vehicletype else "",
                    'buy_cost': float(vehicle.vm_buycost) if vehicle.vm_buycost and str(vehicle.vm_buycost).replace('.','',1).isdigit() else 0.0,
                    'agreed_km': float(vehicle.vm_agreedkm) if vehicle.vm_agreedkm and str(vehicle.vm_agreedkm).replace('.','',1).isdigit() else 0.0,
                    'extra_km_rate': float(vehicle.vm_extrakm) if vehicle.vm_extrakm and str(vehicle.vm_extrakm).replace('.','',1).isdigit() else 0.0,
                })
            except VehiclemasterInfo.DoesNotExist:
                pass
        elif vendor_id:
            pass

        filters = Q()
        if vehicle_id:
            vehicle = VehiclemasterInfo.objects.get(id=vehicle_id)
            filters &= Q(tr_vehiclenumber=vehicle.vm_registrationnumber)
        elif vendor_id:
            vehicle_reg_nos = VehiclemasterInfo.objects.filter(vm_vendor_id=vendor_id).values_list('vm_registrationnumber', flat=True)
            filters &= Q(tr_vehiclenumber__in=vehicle_reg_nos)
        
        if f_dt_obj and t_dt_obj:
            # Use IST-aware datetime boundaries to avoid UTC offset issues on server
            # expanded_start = day before from_date, expanded_end = day after to_date
            expanded_start = f_dt_obj - timedelta(days=1)
            expanded_end = t_dt_obj + timedelta(days=1)
            # Convert to timezone-aware datetimes (start of day in IST = 00:00 IST, end = 23:59:59 IST)
            ist_start = dj_timezone.make_aware(dt_datetime(expanded_start.year, expanded_start.month, expanded_start.day, 0, 0, 0))
            ist_end = dj_timezone.make_aware(dt_datetime(expanded_end.year, expanded_end.month, expanded_end.day, 23, 59, 59))
            filters &= Q(tr_departeddate__gte=ist_start, tr_departeddate__lte=ist_end)

        if filters:
            bill_id = request.GET.get('bill_id')
            billed_trips_query = AttachedBillInfo.objects.all()
            if bill_id:
                billed_trips_query = billed_trips_query.exclude(id=bill_id)
            
            already_billed = []
            for b_trips in billed_trips_query.values_list('ab_selected_trips', flat=True):
                if b_trips:
                    already_billed.extend([t.strip() for t in b_trips.split(',') if t.strip()])
            
            if already_billed:
                filters &= ~Q(tr_tripnumber__in=already_billed)

            # Compute Total KM Run using the exact same filters as the bill table
            # so that it perfectly matches the sum of the trips and avoids the "KM Difference Adjustment"
            all_period_trips = TripdetailInfo.objects.filter(filters).filter(tr_vehiclesource_id__in=[1, 2])
            total_km_run_sum = 0
            for t in all_period_trips:
                s_km = t.tr_reportedkm_pickup if t.tr_reportedkm_pickup else (t.tr_departedkm or 0)
                c_km = t.tr_reportedkm_delivery if t.tr_reportedkm_delivery else (t.tr_reportedkm or 0)
                if c_km and s_km:
                    diff = c_km - s_km
                    if 0 < diff < 15000:
                        total_km_run_sum += diff

            trips = TripdetailInfo.objects.filter(filters).filter(tr_category_id=1).order_by('tr_departeddate')
            trip_dates = set()

            for trip in trips:
                # Per-trip KM logic updated to match Vehicle Log Report
                start_km = trip.tr_reportedkm_pickup if trip.tr_reportedkm_pickup else (trip.tr_departedkm or 0)
                closing_km = trip.tr_reportedkm_delivery if trip.tr_reportedkm_delivery else (trip.tr_reportedkm or 0)
                km = 0
                if closing_km and start_km:
                    diff = closing_km - start_km
                    if 0 < diff < 15000:
                        km = diff

                from django.utils import timezone
                trip_dte = timezone.localtime(trip.tr_departeddate).date() if trip.tr_departeddate else None

                if trip.tr_departeddate:
                    from django.utils import timezone
                    start_date = timezone.localtime(trip.tr_departeddate).date()
                    end_dt_val = (trip.tr_reporteddate_delivery or trip.tr_reporteddate or trip.tr_departeddate)
                    end_date_val = timezone.localtime(end_dt_val).date()
                    temp_date = start_date
                    while temp_date <= end_date_val:
                        trip_dates.add(temp_date)
                        temp_date += timedelta(days=1)

                # Filter trips for the table display to just the current period
                if trip_dte and f_dt_obj <= trip_dte <= t_dt_obj:
                    customer_name = trip.tr_enquirynumber.en_customername.cu_name if trip.tr_enquirynumber and trip.tr_enquirynumber.en_customername else "N/A"
                    data['trips'].append({
                        'id': trip.id,
                        'trip_date': timezone.localtime(trip.tr_departeddate).strftime('%d-%m-%Y') if trip.tr_departeddate else "",
                        'trip_no': trip.tr_tripnumber or "",
                        'cnote': trip.tr_consignmentnumber.co_consignmentnumber if trip.tr_consignmentnumber else "",
                        'customer': customer_name,
                        'from': str(trip.tr_departedlocation) if trip.tr_departedlocation else "",
                        'to': str(trip.tr_reportedlocation) if trip.tr_reportedlocation else "",
                        'trip_km': km,
                        'selling_cost': float(trip.tc_tripcost or 0),
                        'buy_cost': 0.0, # Will be calculated on frontend from header distribution
                        'parking_cost': float(trip.tc_parkingcost or 0),
                        'toll_cost': float(trip.tc_tollcost or 0),
                        'total_buy_cost': 0.0 # Will be calculated on frontend
                    })

            # Total KM = sum of Used KM for ALL period trips (matches Vehicle Log Report)
            data['total_km_run'] = total_km_run_sum

            if f_dt_obj and t_dt_obj:
                # Build the full set of leave dates first
                leave_dates_set = set()
                curr = f_dt_obj
                while curr <= t_dt_obj:
                    if curr not in trip_dates:
                        leave_dates_set.add(curr)
                    curr += timedelta(days=1)

                # Sunday rule: Sunday only counts as leave if BOTH Saturday and Monday are also leave
                leave_days_count = 0
                for d in sorted(leave_dates_set):
                    if d.weekday() == 6:  # Sunday
                        saturday = d - timedelta(days=1)
                        monday = d + timedelta(days=1)
                        # Only count Sunday as leave if both Saturday and Monday are also leave
                        sat_is_leave = saturday < f_dt_obj or saturday in leave_dates_set
                        mon_is_leave = monday > t_dt_obj or monday in leave_dates_set
                        if sat_is_leave and mon_is_leave:
                            leave_days_count += 1
                    else:
                        leave_days_count += 1
                data['leave_days'] = leave_days_count
        else:
            data['total_km_run'] = 0
            data['leave_days'] = 0

    except Exception as e:
        data['error'] = str(e)
        import traceback
        data['traceback'] = traceback.format_exc()

    return JsonResponse(data)


@login_required(login_url='login_page')
def attached_bill_export_tally(request):
    """Export Attached Bills in Tally-friendly Excel format.
    Supports optional GET params: from_date and to_date (YYYY-MM-DD).
    """
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    vendor_id = request.GET.get('vendor_id', '')

    bills = AttachedBillInfo.objects.all().order_by('ab_bill_date')
    if from_date:
        bills = bills.filter(ab_from_date__gte=from_date)
    if to_date:
        bills = bills.filter(ab_to_date__lte=to_date)
    if vendor_id:
        bills = bills.filter(ab_vendor_id=vendor_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Export"

    headers = [
        "VOUCHER NUMBER", "DATE", "REF NO.", "SUNDRY CREDITORS", "TOTAL AMT",
        "EXPENSES LEDGER", "AMOUNT", "Primary Cost Category", "Job No",
        "VEH. NO.", "Customer", "TDS LEDGER", "TDS AMOUNT", "NARRATION"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="B2FFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    expense_fill = PatternFill("solid", fgColor="FFE5CC")

    for bill in bills:
        selected_trip_numbers = []
        if bill.ab_selected_trips:
            selected_trip_numbers = [t.strip() for t in bill.ab_selected_trips.split(',') if t.strip()]

        trips = TripdetailInfo.objects.filter(tr_tripnumber__in=selected_trip_numbers).select_related('tr_enquirynumber', 'tr_consignmentnumber')

        empty_trips = []
        if bill.ab_vehicle_number and bill.ab_from_date and bill.ab_to_date:
            km_filters = Q(tr_vehiclenumber=bill.ab_vehicle_number.vm_registrationnumber)
            km_filters &= (
                Q(tr_departeddate__date__range=[bill.ab_from_date, bill.ab_to_date]) |
                Q(tr_departeddate_pickup__date__range=[bill.ab_from_date, bill.ab_to_date]) |
                Q(tr_reporteddate__date__range=[bill.ab_from_date, bill.ab_to_date]) |
                Q(tr_reporteddate_pickup__date__range=[bill.ab_from_date, bill.ab_to_date]) |
                Q(tr_loading_time__date__range=[bill.ab_from_date, bill.ab_to_date]) |
                Q(tr_unloading_time__date__range=[bill.ab_from_date, bill.ab_to_date])
            )
            empty_trips = TripdetailInfo.objects.filter(km_filters).filter(tr_category_id__in=[2, 3]).filter(tr_vehiclesource_id__in=[1, 2]).select_related('tr_enquirynumber', 'tr_consignmentnumber')

        all_trips = list(trips) + list(empty_trips)
        unique_trips = []
        seen_trip_ids = set()
        for t in all_trips:
            if t.id not in seen_trip_ids:
                seen_trip_ids.add(t.id)
                unique_trips.append(t)

        # Financial year and month in voucher
        if bill.ab_bill_date:
            year = bill.ab_bill_date.year
            month = bill.ab_bill_date.month
            if month >= 4:
                fy_str = f"{str(year)[-2:]}-{str(year+1)[-2:]}"
            else:
                fy_str = f"{str(year-1)[-2:]}-{str(year)[-2:]}"
            month_str_num = f"{month:02d}"
        else:
            fy_str = "00-00"
            month_str_num = "00"

        voucher_number = f"MAA_ATT_{fy_str}_{month_str_num}_{bill.id:03d}"
        bill_date = bill.ab_bill_date.strftime("%d-%m-%Y") if bill.ab_bill_date else ""
        ref_no = bill.ab_bill_no or ""
        vendor_name = bill.ab_vendor.vend_name if bill.ab_vendor else ""
        # Show payable amount in the TOTAL AMT column; fall back to bill amount if payable not set
        total_amt = float(bill.ab_payable_amount if (bill.ab_payable_amount is not None and bill.ab_payable_amount != 0) else (bill.ab_bill_amount or 0.0))

        tds_amount = float(bill.ab_tds_amount or 0.0)
        # Determine TDS ledger label based on saved TDS type
        tds_ledger = ""
        try:
            tds_type = (bill.ab_tds_type or '').strip()
            if tds_amount > 0:
                if tds_type == 'Company':
                    tds_ledger = "TDS Payable 194C (Company)"
                else:
                    tds_ledger = "TDS Payable 194C (Non Company)"
        except Exception:
            tds_ledger = "TDS Payable 194C (Non Company)" if tds_amount > 0 else ""

        month_short = bill.ab_from_date.strftime('%b%y') if bill.ab_from_date else ""
        rec_date_str = bill.ab_created_at.strftime('%d-%b-%y') if bill.ab_created_at else ""
        narration = f"Being ATT Vehicle expenses for the month of {month_short} (Bill Received on {rec_date_str})"

        is_first_row = True

        if not unique_trips:
            vehicle_disp = ""
            if bill.ab_vehicle_number and getattr(bill.ab_vehicle_number, 'vm_registrationnumber', None):
                vehicle_disp = f"{bill.ab_vehicle_number.vm_registrationnumber}(A)"
            row = [
                voucher_number, bill_date, ref_no, vendor_name, total_amt,
                "Transportation", total_amt, "Maa-Att", "",
                vehicle_disp, "", (tds_ledger if tds_amount > 0 else ""), (tds_amount if tds_amount > 0 else ""), narration
            ]
            ws.append(row)
            continue

        # Separate regular trips from empty/business empty
        regular_trips = [t for t in unique_trips if getattr(t, 'tr_category_id', None) not in [2, 3]]
        empty_only_trips = [t for t in unique_trips if getattr(t, 'tr_category_id', None) == 2]
        biz_empty_trips = [t for t in unique_trips if getattr(t, 'tr_category_id', None) == 3]

        total_toll = sum(float(t.tc_tollcost or 0) for t in regular_trips)
        bill_buy_cost = float(bill.ab_bill_amount or 0) - total_toll
        bill_total_km = float(bill.ab_total_km_run or 0)

        def calc_trip_km(trip):
            start_km = trip.tr_reportedkm_pickup if trip.tr_reportedkm_pickup else (trip.tr_departedkm or 0)
            closing_km = trip.tr_reportedkm_delivery if trip.tr_reportedkm_delivery else (trip.tr_reportedkm or 0)
            if closing_km and start_km:
                diff = closing_km - start_km
                if 0 < diff < 15000:
                    return diff
            return 0

        def calc_buy_cost(trip_km):
            if bill_total_km > 0 and trip_km > 0:
                return round((bill_buy_cost / bill_total_km) * trip_km, 2)
            return 0.0

        # Pre-calculate proportional amounts for ALL rows (regular + empty + biz empty)
        # so we can adjust the last row to make the sum exactly equal bill_buy_cost
        all_rows = []  # each entry: (type, trip_or_None, km, amount, job_no, customer_name, vehicle_disp, narration_suffix)

        for trip in regular_trips:
            cnote = trip.tr_consignmentnumber.co_consignmentnumber if trip.tr_consignmentnumber else ""
            job_no = cnote or (trip.tr_enquirynumber.en_enquirynumber if trip.tr_enquirynumber else (trip.tr_tripnumber or ""))
            customer_name = ""
            if trip.tr_enquirynumber and getattr(trip.tr_enquirynumber, 'en_customername', None):
                try:
                    customer_name = trip.tr_enquirynumber.en_customername.cu_name
                except:
                    customer_name = str(trip.tr_enquirynumber.en_customername)
            trip_km = calc_trip_km(trip)
            transport_cost = calc_buy_cost(trip_km)
            vehicle_disp = f"{trip.tr_vehiclenumber}(A)" if trip.tr_vehiclenumber else ""
            all_rows.append(("Transportation", transport_cost, job_no, vehicle_disp, customer_name, ""))
            if trip.tc_tollcost and float(trip.tc_tollcost) > 0:
                all_rows.append(("Toll", float(trip.tc_tollcost), job_no, vehicle_disp, customer_name, ""))

        # Consolidated empty row
        if empty_only_trips:
            total_empty_km = sum(calc_trip_km(t) for t in empty_only_trips)
            total_empty_amount = calc_buy_cost(total_empty_km)
            vehicle_disp = f"{bill.ab_vehicle_number.vm_registrationnumber}(A)" if bill.ab_vehicle_number and getattr(bill.ab_vehicle_number, 'vm_registrationnumber', None) else ""
            all_rows.append(("Transportation", total_empty_amount, "NA(J)", vehicle_disp, "NA(C)", " (Empty)"))

        # Consolidated business empty row
        if biz_empty_trips:
            total_biz_km = sum(calc_trip_km(t) for t in biz_empty_trips)
            total_biz_amount = calc_buy_cost(total_biz_km)
            vehicle_disp = f"{bill.ab_vehicle_number.vm_registrationnumber}(A)" if bill.ab_vehicle_number and getattr(bill.ab_vehicle_number, 'vm_registrationnumber', None) else ""
            all_rows.append(("Transportation", total_biz_amount, "NA(J)", vehicle_disp, "NA(C)", " (Business Empty)"))

        # Adjust last Transportation row so sum of all Transportation amounts = bill_buy_cost exactly
        transport_rows_idx = [i for i, r in enumerate(all_rows) if r[0] == "Transportation"]
        if transport_rows_idx:
            current_transport_sum = sum(all_rows[i][1] for i in transport_rows_idx)
            diff = round(bill_buy_cost - current_transport_sum, 2)
            if diff != 0:
                # Append difference as a separate adjustment row to keep empty/business empty amounts perfectly matched with summary
                vehicle_disp = f"{bill.ab_vehicle_number.vm_registrationnumber}(A)" if bill.ab_vehicle_number and getattr(bill.ab_vehicle_number, 'vm_registrationnumber', None) else ""
                all_rows.append(("Transportation", diff, "NA(J)", vehicle_disp, "NA(C)", " (KM Difference Adjustment)"))

        # Write all rows to worksheet
        for exp_name, amt, job_no, vehicle_disp, customer_name, narr_suffix in all_rows:
            row = [
                voucher_number, bill_date, ref_no,
                vendor_name if is_first_row else "",
                total_amt if is_first_row else "",
                exp_name, amt,
                "Maa-Att", job_no, vehicle_disp, customer_name,
                tds_ledger if is_first_row else "",
                tds_amount if is_first_row and tds_amount > 0 else "",
                narration + narr_suffix
            ]
            ws.append(row)
            ws.cell(row=ws.max_row, column=6).fill = expense_fill
            is_first_row = False

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Attached_Bill_Tally_Export.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login_page')
def attached_bill_summary(request, id):
    """Return structured summary data for the bill summary modal."""
    bill = get_object_or_404(AttachedBillInfo, id=id)
    vehicle = bill.ab_vehicle_number
    vendor = bill.ab_vendor

    # --- Period info ---
    from_date = bill.ab_from_date
    to_date = bill.ab_to_date
    days_in_month = 0
    working_days = 0
    if from_date and to_date:
        days_in_month = (to_date - from_date).days + 1
    # --- Use SAVED bill fields for reliable values ---
    contract_amount = float(bill.ab_buy_cost or 0)
    
    # Recalculate working_days and leave_days based on the new 100% days rule
    working_days = days_in_month  # Denominator is now all days in the month
    leave_per_day = (contract_amount / working_days) if working_days > 0 else 0

    # Fetch ALL trips for this vehicle in the expanded range to calculate leave days correctly
    all_trip_dates = set()
    actual_empty_km = 0.0
    actual_business_empty_km = 0.0
    if vehicle:
        expanded_start = from_date - timedelta(days=1)
        expanded_end = to_date + timedelta(days=1)
        all_trips_in_period = TripdetailInfo.objects.filter(
            tr_vehiclenumber=vehicle.vm_registrationnumber,
            tr_departeddate__date__range=[expanded_start, expanded_end]
        )
        for t in all_trips_in_period:
            if t.tr_departeddate:
                sd = t.tr_departeddate.date()
                ed = (t.tr_reporteddate_delivery or t.tr_reporteddate or t.tr_departeddate).date()
                curr_t = sd
                while curr_t <= ed:
                    all_trip_dates.add(curr_t)
                    curr_t += timedelta(days=1)

        # Compute actual empty KM (category 2 or 3) matching Total KM run period
        km_filters = Q(tr_vehiclenumber=vehicle.vm_registrationnumber)
        km_filters &= (
            Q(tr_departeddate__date__range=[from_date, to_date]) |
            Q(tr_departeddate_pickup__date__range=[from_date, to_date]) |
            Q(tr_reporteddate__date__range=[from_date, to_date]) |
            Q(tr_reporteddate_pickup__date__range=[from_date, to_date]) |
            Q(tr_loading_time__date__range=[from_date, to_date]) |
            Q(tr_unloading_time__date__range=[from_date, to_date])
        )
        all_km_trips = TripdetailInfo.objects.filter(km_filters).filter(tr_vehiclesource_id__in=[1, 2])
        for t in all_km_trips:
            if t.tr_category_id in [2, 3]:
                s_km = t.tr_reportedkm_pickup if t.tr_reportedkm_pickup else (t.tr_departedkm or 0)
                c_km = t.tr_reportedkm_delivery if t.tr_reportedkm_delivery else (t.tr_reportedkm or 0)
                if c_km and s_km:
                    diff = c_km - s_km
                    if 0 < diff < 15000:
                        if t.tr_category_id == 2:
                            actual_empty_km += diff
                        elif t.tr_category_id == 3:
                            actual_business_empty_km += diff

    # Use the reliably saved leave days and amount
    leave_days = bill.ab_leave_days or 0
    leave_amount = float(bill.ab_leave_amount or 0)
    agreed_km      = float(bill.ab_agreed_km or 0)
    total_km_saved = float(bill.ab_total_km_run or 0)   # saved total KM from ADD/EDIT
    extra_km       = float(bill.ab_extra_km_run or 0)
    extra_km_amount = float(bill.ab_extra_km_amount or 0)
    toll_cost      = float(bill.ab_toll_cost or 0)
    actual_amount  = float(bill.ab_bill_amount or 0)

    # --- Fetch selected trips ONLY ---
    selected_trip_numbers = [t.strip() for t in (bill.ab_selected_trips or '').split(',') if t.strip()]
    total_trips = len(selected_trip_numbers)

    trips = []
    if selected_trip_numbers:
        trip_filters = Q(tr_tripnumber__in=selected_trip_numbers)
        # Optional: verify vehicle and date range for data integrity
        if vehicle:
            trip_filters &= Q(tr_vehiclenumber=vehicle.vm_registrationnumber)
        
        trips = list(TripdetailInfo.objects.filter(trip_filters).select_related(
            'tr_departedlocation', 'tr_reportedlocation', 'tr_enquirynumber'
        ))

    # --- Fetch selected trips ONLY (for KM/Selling) ---
    selected_trip_numbers = [t.strip() for t in (bill.ab_selected_trips or '').split(',') if t.strip()]
    total_trips = len(selected_trip_numbers)
    trips = []
    if selected_trip_numbers:
        trip_filters = Q(tr_tripnumber__in=selected_trip_numbers)
        if vehicle:
            trip_filters &= Q(tr_vehiclenumber=vehicle.vm_registrationnumber)
        trips = list(TripdetailInfo.objects.filter(trip_filters).select_related(
            'tr_departedlocation', 'tr_reportedlocation', 'tr_enquirynumber'
        ))

    # trip_days for display in summary
    summary_billed_trip_dates = set()
    trip_km_run = 0
    for t in trips:
        start_km = t.tr_reportedkm_pickup if t.tr_reportedkm_pickup else (t.tr_departedkm or 0)
        closing_km = t.tr_reportedkm_delivery if t.tr_reportedkm_delivery else (t.tr_reportedkm or 0)
        if closing_km and start_km:
            diff = closing_km - start_km
            if 0 < diff < 15000:
                trip_km_run += diff
        if t.tr_departeddate:
            summary_billed_trip_dates.add(t.tr_departeddate.date())

    # Use saved total_km_run
    display_total_km = total_km_saved

    days_run   = len(summary_billed_trip_dates)
    trip_index = f"{total_trips} TRIPS / {days_run} DAYS RUN" if days_run else f"{total_trips} TRIPS"

    # --- Empty KM = from DB trips with category 2 or 3 ---
    empty_km           = actual_empty_km
    business_empty_km  = actual_business_empty_km

    per_km_amount      = ((actual_amount - toll_cost) / total_km_saved) if total_km_saved > 0 else 0
    empty_km_buy_cost  = per_km_amount * empty_km
    business_empty_km_buy_cost = per_km_amount * business_empty_km

    # --- Selling = sum of all trip charges billed to customer ---
    selling = 0.0
    for t in trips:
        selling += (float(t.tc_tripcost or 0) + float(t.tc_tollcost or 0) +
                    float(t.tc_supervisorcost or 0) + float(t.tc_loadingcost or 0) +
                    float(t.tc_unloadingcost or 0) + float(t.tc_weighmentcost or 0) +
                    float(t.tc_haltingcost or 0) + float(t.tc_handlingcost or 0))

    # --- Buying = actual bill amount paid to vendor ---
    buying = actual_amount

    # --- Profit / % ---
    profit   = selling - buying
    sell_pct = round((profit / selling) * 100, 2) if selling > 0 else 0
    buy_pct  = round((profit / buying)  * 100, 2) if buying  > 0 else 0

    # --- Title ---
    vendor_name  = vendor.vend_name if vendor else "N/A"
    period_label = from_date.strftime("%b %Y").upper() if from_date else ""
    title        = f"{vendor_name} - {period_label}"

    return JsonResponse({
        'title': title,
        'vehicle_no':   vehicle.vm_registrationnumber if vehicle else '',
        'vehicle_type': str(vehicle.vm_vehicletype) if vehicle and vehicle.vm_vehicletype else '',
        'days_in_month': days_in_month,
        'working_days':  working_days,
        'leave_days':    leave_days,
        'contract_amount': round(contract_amount, 2),
        'toll_cost':       round(toll_cost, 2),
        'leave_per_day':   round(leave_per_day, 2),
        'leave_amount':    round(leave_amount, 2),
        'extra_km':        round(extra_km, 2),
        'extra_km_amount': round(extra_km_amount, 2),
        'actual_amount':   round(actual_amount, 2),
        'total_trips': total_trips,
        'trip_index':  trip_index,
        'total_km':    round(display_total_km, 2),
        'agreed_km':   round(agreed_km, 2),
        'empty_km':              round(empty_km, 2),
        'per_km_amount':         round(per_km_amount, 2),
        'empty_km_buy_cost':     round(empty_km_buy_cost, 2),
        'business_empty_km':          round(business_empty_km, 2),
        'business_empty_km_buy_cost': round(business_empty_km_buy_cost, 2),
        'selling': round(selling, 2),
        'buying':  round(buying, 2),
        'profit':  round(profit, 2),
        'sell_pct': sell_pct,
        'buy_pct':  buy_pct,
    })


@login_required(login_url='login_page')
def get_vehicles_by_vendor(request):
    vendor_id = request.GET.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'vehicles': []})
    
    # Ownership ID 2 = ATTACHED
    vehicles = VehiclemasterInfo.objects.filter(vm_vendor_id=vendor_id, vm_ownership_id=2).order_by('vm_registrationnumber')
    
    veh_list = []
    for v in vehicles:
        veh_list.append({
            'id': v.id,
            'reg_no': v.vm_registrationnumber
        })
    
    return JsonResponse({'vehicles': veh_list})
