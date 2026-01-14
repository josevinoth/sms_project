import calendar
from datetime import datetime, date
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook
from ..models import TripdetailInfo, EnquirynoteInfo, ConsignmentdetailInfo, MyUser, Places, Emailmaster
from .send_department_email import send_department_email
from ..sub_forms.dmr_report_form import DmrForm
from ..sub_models.consignmentgoods_mod import ConsignmentgoodsInfo
from ..sub_models.customer_mod import CustomerInfo
from ..sub_models.customerdepartment_mod import CustomerdepartmentInfo
from ..sub_models.vehicle_allotment_mod import Vehicle_allotmentInfo


# -------------------------------------------------------------------------
# CONSTANTS & TEMPLATES (Consolidated)
# -------------------------------------------------------------------------

DMR_TEMPLATES = {
    "Air Export": [
        "S.NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "CONSIGNEE NAME", "DELIVERY LOCATION", "CS NAME", "PLANNING RECEIVED DATE",
        "PLANNING RECEIVED TIME", "VEHICLE PLACED TIME", "HAWB #", "BOE NUMBER # / EWAYBILL #",
        "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )", "TRUCK NO", "TRUCK TYPE", "VENDOR",
        "CHA NAME", "LLR NO", "DRIVER NAME", "DRIVER MOBILE", "DRIVER DL #",
        "PICKUP POINT IN DATE", "PICKUP POINT IN TIME", "PICKUP POINT OUT DATE",
        "PICKUP POINT OUT TIME", "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT",
        "CBM", "SHIPPER SEAL #", "UNLOADING POINT IN DATE", "UNLOADING POINT IN TIME",
        "UNLOADING POINT", "UNLOADING POINT OUT DATE", "UNLOADING POINT OUT TIME",
        "NO OF DAYS HALTING", "ADDITIONAL CHARGES", "CANCELLATION CHARGES",
        "HALTING CHARGES", "CHARGES", "WEIGHTMENT CHARGES", "PARKING / UNLOADING CHARGES",
        "TOTAL CHARGES", "REMARKS"
    ],
    "Air Import": [
        "S.NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "CONSIGNEE NAME", "DELIVERY LOCATION", "CS NAME", "PLANNING RECEIVED DATE",
        "PLANNING RECEIVED TIME", "VEHICLE PLACED TIME", "HAWB #", "BOE NUMBER # / EWAYBILL #",
        "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )", "TRUCK NO", "TRUCK TYPE", "VENDOR",
        "CHA NAME", "LLR NO", "DRIVER NAME", "DRIVER MOBILE", "DRIVER DL #",
        "PICKUP POINT IN DATE", "PICKUP POINT IN TIME", "PICKUP POINT OUT DATE",
        "PICKUP POINT OUT TIME", "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT",
        "CBM", "SHIPPER SEAL #", "UNLOADING POINT IN DATE", "UNLOADING POINT IN TIME",
        "UNLOADING POINT", "UNLOADING POINT OUT DATE", "UNLOADING POINT OUT TIME",
        "NO OF DAYS HALTING", "ADDITIONAL CHARGES", "CANCELLATION CHARGES",
        "HALTING CHARGES", "CHARGES", "WEIGHTMENT CHARGES", "PARKING CHARGES",
        "UNLOADING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "CHB": [
        "S.NO", "JOB NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "CONSIGNEE NAME", "DELIVERY LOCATION", "CS NAME", "PLANNING RECEIVED DATE",
        "PLANNING RECEIVED TIME", "HAWB #", "BOE NUMBER # / EWAYBILL #",
        "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )", "TRUCK NO", "TRUCK TYPE", "VENDOR",
        "CHA NAME", "LLR NO", "DRIVER NAME", "DRIVER MOBILE", "DRIVER DL #",
        "PICKUP POINT IN DATE", "PICKUP POINT IN TIME", "PICKUP POINT OUT DATE",
        "PICKUP POINT OUT TIME", "NO OF PIECES", "ACTUAL WEIGHT (KGS)",
        "CHARGEABLE WEIGHT", "CAPACITY", "SHIPPER SEAL #", "UNLOADING POINT IN DATE",
        "UNLOADING POINT IN TIME", "UNLOADING POINT", "UNLOADING POINT OUT DATE",
        "UNLOADING POINT OUT TIME", "HALTING STATUS (YES / NO)", "NO OF DAYS HALTING",
        "ADDITIONAL CHARGES", "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES",
        "PARKING / UNLOADING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "Hub Movement": [
        "S.NO", "OFD DATE", "VENDOR NAME", "VEH NO", "COC NO", "ORGIN", "DESTINATION",
        "CONSIGNOR", "CONSIGNEE NAME", "NO PKG", "WEIGHT", "DELIVERY STATUS", "REMARKS",
        "DELAY&ONTIME", "POD REMARKS", "RETURN BOX"
    ],
    "Order Management": [
        "S.NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "CONSIGNEE NAME", "DELIVERY LOCATION", "CS NAME", "PLANNING RECEIVED DATE",
        "PLANNING RECEIVED TIME", "VEHICLE PLACED TIME", "HAWB #", "BOE NUMBER # / EWAYBILL #",
        "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )", "TRUCK NO", "TRUCK TYPE", "VENDOR",
        "DRIVER NAME", "DRIVER MOBILE", "DRIVER DL #", "PICKUP POINT IN DATE",
        "PICKUP POINT IN TIME", "PICKUP POINT OUT DATE", "PICKUP POINT OUT TIME",
        "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT", "CBM", "SHIPPER SEAL #",
        "UNLOADING POINT IN DATE", "UNLOADING POINT IN TIME", "UNLOADING POINT",
        "UNLOADING POINT OUT DATE", "UNLOADING POINT OUT TIME", "NO OF DAYS HALTING",
        "ADDITIONAL CHARGES", "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES",
        "WEIGHTMENT CHARGES", "UNLOADING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "Sea Export": [
        "S.NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "CONSIGNEE NAME", "DELIVERY LOCATION", "CS NAME", "PLANNING RECEIVED DATE",
        "PLANNING RECEIVED TIME", "VEHICLE PLACED TIME", "HAWB #", "BOE NUMBER # / EWAYBILL #",
        "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )", "TRUCK NO", "TRUCK TYPE", "VENDOR",
        "DRIVER NAME", "DRIVER MOBILE", "DRIVER DL #", "PICKUP POINT IN DATE",
        "PICKUP POINT IN TIME", "PICKUP POINT OUT DATE", "PICKUP POINT OUT TIME",
        "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT", "CBM", "SHIPPER SEAL #",
        "UNLOADING POINT IN DATE", "UNLOADING POINT IN TIME", "UNLOADING POINT",
        "UNLOADING POINT OUT DATE", "UNLOADING POINT OUT TIME", "NO OF DAYS HALTING",
        "ADDITIONAL CHARGES", "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES",
        "WEIGHTMENT CHARGES", "UNLOADING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "Sea Import": [
        "S.NO", "JOB NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "CONSIGNEE NAME", "DELIVERY LOCATION", "CS NAME", "PLANNING RECEIVED DATE",
        "PLANNING RECEIVED TIME", "HAWB #", "BOE NUMBER # / EWAYBILL #",
        "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )", "TRUCK NO", "TRUCK TYPE", "VENDOR",
        "CHA NAME", "LLR NO", "DRIVER NAME", "DRIVER MOBILE", "DRIVER DL #",
        "PICKUP POINT IN DATE", "PICKUP POINT IN TIME", "PICKUP POINT OUT DATE",
        "PICKUP POINT OUT TIME", "NO OF PIECES", "ACTUAL WEIGHT (KGS)", "CHARGEABLE WEIGHT",
        "CAPACITY", "SHIPPER SEAL #", "UNLOADING POINT IN DATE", "UNLOADING POINT IN TIME",
        "UNLOADING POINT", "UNLOADING POINT OUT DATE", "UNLOADING POINT OUT TIME",
        "HALTING STATUS (YES / NO)", "NO OF DAYS HALTING", "ADDITIONAL CHARGES",
        "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES", "PARKING / UNLOADING CHARGES",
        "TOTAL CHARGES", "REMARKS"
    ],
    "TCS Local": [
        "SO NO", "JOB NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "DELIVERY LOCATION", "BOE NUMBER # / EWAYBILL #", "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )",
        "TRUCK NO", "TRUCK TYPE", "VENDOR", "DRIVER MOBILE", "PICKUP POINT IN DATE",
        "PICKUP POINT OUT DATE", "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT", "CBM",
        "SHIPPER SEAL #", "UNLOADING POINT IN DATE", "UNLOADING POINT OUT DATE", "NO OF DAYS HALTING",
        "ADDITIONAL CHARGES", "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES", "PARKING CHARGES",
        "UNLOADING CHARGES & LASHING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "TCS Outstation": [
        "S.NO", "JOB NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "DELIVERY LOCATION", "BOE NUMBER # / EWAYBILL #", "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )",
        "TRUCK NO", "TRUCK TYPE", "VENDOR", "DRIVER MOBILE", "PICKUP POINT IN DATE", "PICKUP POINT OUT DATE",
        "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT", "CBM", "SHIPPER SEAL #",
        "UNLOADING POINT IN DATE", "UNLOADING POINT OUT DATE", "NO OF DAYS HALTING", "ADDITIONAL CHARGES",
        "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES", "PARKING CHARGES",
        "UNLOADING CHARGES AND PARKING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "TCS Reefer": [
        "S.NO", "JOB NO", "DEPARTMENT NAME", "PICKUP DATE", "SHIPPER NAME", "PICKUP LOCATION",
        "DELIVERY LOCATION", "BOE NUMBER # / EWAYBILL #", "REFERENCE NUMBER ( MASTER INVOICE # / FILE # )",
        "TRUCK NO", "TRUCK TYPE", "VENDOR", "DRIVER MOBILE", "PICKUP POINT IN DATE", "PICKUP POINT OUT DATE",
        "NO OF PIECES", "ACTUAL WEIGHT", "CHARGEABLE WEIGHT", "CBM", "SHIPPER SEAL #",
        "UNLOADING POINT IN DATE", "UNLOADING POINT OUT DATE", "NO OF DAYS HALTING", "ADDITIONAL CHARGES",
        "CANCELLATION CHARGES", "HALTING CHARGES", "CHARGES", "PARKING CHARGES",
        "UNLOADING CHARGES AND PARKING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "WH TO AIRPORT": [
        "S.NO", "DEPT", "DATE", "SHIPPER NAME", "FROM", "TO", "CUTOMER SERVICE", "HBL #",
        "TRUCK NO", "TRUCK TYPE", "DRIVER MOBILE", "IN DATE", "IN TIME", "OUT DATE",
        "OUT TIME", "NO OF PIECES", "CARGO WEIGHT", "AIRPORT/BVM GATE IN DATE",
        "AIRPORT/BVM   GATE IN TIME", "UNLOADING  POINT", "IN TIME @ UNLOADING POINT",
        "UNLOADING  TIME", "DLV OUT  DATE", "DLV OUT TIME", "HALTING STATUS   (YES / NO)",
        "NO OF DAYS  HALTING", "ADDITIONAL CHARGES", "CANCELLING CHARGES", "HALTING CHARGES",
        "CHARGES", "WEIGHMENT PASS", "PARKING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "APMT": [
        "MONTH", "REQUESTOR", "SHIPPER", "TRANSPORTER", "CONTAINER SIZE", "BOOKING NO", "FROM",
        "TO", "BVM JOB", "VEHICLE NO.", "DRIVER NO.", "PLACEMENT & VEHICLE PLACED DATE",
        "VEHICLE PLACED TIME", "VEHICLE RELEASED DATE", "VEHICLE RELEASED TIME", "CFS REACHED DATE",
        "CFS REACHED TIME", "UNLOADING CHARGES", "HALTING CHARGES", "TRIP COST", "TOTAL COST",
        "LR NO", "COMMENTS", "DETENTION DAYS", "POD STATUS"
    ],
    "CEVA Air Import": [
        "DATE", "CONSIGNEE NAME", "HBL NUMBER", "BOE NO", "PKGS", "GROSS WEIGHT", "FROM",
        "DELIVERY PLACE", "CEVA JOB NO", "TRUCK NO", "TRUCK TYPE", "DELIVERY DATE",
        "HALTING CHARGES", "UNLOADING CHARGES", "AIRPORT PASS", "TRIP COST", "TOTAL COST",
        "BVM JOB NO", "POD STATUS"
    ],
    "CEVA Export": [
        "DATE", "BVM JOB NO", "CONSIGNEE NAME", "HBL NO", "PKGS", "G WEIGHT", "CEVA JOB NO",
        "FROM", "DELIVERY PLACE", "TRUCK NO", "TRUCK TYPE", "REACHED CFS", "DELIVERY DATE",
        "UNLOADING CHARGES", "TRIP COST", "TOTAL COST", "POD STATUS"
    ],
    "DHL Sea Import": [
        "DATE", "CONSIGNEE NAME", "HBL NO", "BE #", "PKGS", "G WEIGHT", "CBM", "FROM",
        "DELIVERY PLACE", "TRUCK NO", "TRUCK TYPE", "REACHED PLANT", "DELIVERY DATE",
        "HBL WISE SPLIT COST", "LOADING/UNLOADING CHARGES", "HALTING CHARGES",
        "TRANSPORT COST", "TOTAL COST", "VENDOR CODE", "REMARK", "BVM JOB NO", "POD STATUS"
    ],
    "DSV": [
        "DATE", "BVM JOB NO", "BVM LR NO", "USER NAME", "SHIPPER NAME", "HBL NO/REFERENCE NO",
        "TRANSPORT BILL TO", "VEHICLE NO", "VEHICLE TYPE", "DRIVER NAME", "DRIVER NAMBER", "FROM",
        "TO", "DIVISION", "SUM OF PIECES", "WEIGHT", "HALTING CHARGES", "LOADING CHARGES",
        "UNLOADING CHARGES", "WEIGHMENT CHARGES", "AAI S.NO.", "AAI CHARGES", "TRIP COST",
        "TOTAL COST", "E-WAY BILL", "BVM INVOICE", "SEEL NO", "REMARK", "POD", "C NOTE"
    ],
    "DSV DD REPORT": [
        "DATE", "TRIP SHEET NO", "VEHICLE NO", "STARTING TIME", "CLOSING TIME", "STARTING KM",
        "CLOSING KM", "USED KM", "STARTING PLACE", "CLOSING PLACE", "HBL NO/REFERENCE NO",
        "DETENTION HOURS", "RATE PER KM", "TRIP CHARGE", "PARKING CHARGES", "HALTING CHARGES",
        "DETENTION CHARGES", "TOTAL COST"
    ],
    "EIPL Nagalkeni To Airport": [
        "S.NO", "DEPT", "DATE", "SHIPPER NAME", "FROM", "TO", "CUTOMER SERVICE", "HBL #",
        "TRUCK NO", "TRUCK TYPE", "DRIVER MOBILE", "IN DATE", "IN TIME", "OUT DATE",
        "OUT TIME", "NO OF PIECES", "CARGO WEIGHT", "AIRPORT/BVM GATE IN DATE",
        "AIRPORT/BVM   GATE IN TIME", "UNLOADING  POINT", "IN TIME @ UNLOADING POINT",
        "UNLOADING  TIME", "DLV OUT  DATE", "DLV OUT TIME", "HALTING STATUS   (YES / NO)",
        "NO OF DAYS  HALTING", "ADDITIONAL CHARGES", "CANCELLING CHARGES", "HALTING CHARGES",
        "CHARGES", "WEIGHMENT PASS", "PARKING CHARGES", "TOTAL CHARGES", "REMARKS"
    ],
    "DHL Other": [
        "DATE", "BVM JOB NO", "CONSIGNEE NAME", "HBL NO", "PKGS", "G WEIGHT", "CBM",
        "VENDOR CODE", "FROM", "DELIVERY PLACE", "TRUCK NO", "TRUCK TYPE", "REACHED",
        "DELIVERY DATE", "HBL WISE SPLIT COST", "TRANSPORT COST", "TOTAL COST",
        "REMARK", "POD STATUS"
    ],
    "DHL BLR Import": [
        "DATE", "CONSIGNEE NAME", "HBL No", "BE #", "PKGS", "G WEIGHT", "CBM", "FROM",
        "DELIVERY PLACE", "TRUCK NO", "TRUCK TYPE", "REACHED PLANT", "DELIVERY DATE",
        "HBL WISE SPLIT COST", "LOADING/UNLOADING CHARGES", "TOLL CHARGES", "PARKING CHARGES",
        "HALTING CHARGES", "TRANSPORT COST", "TOTAL COST", "VENDOR CODE", "REMARK", "Cnote No"
    ],
    "DHL BLR Export": [
        "DATE", "CONSIGNEE NAME", "HBL No", "PKGS", "G WEIGHT", "CBM", "FROM",
        "DELIVERY PLACE", "TRUCK NO", "TRUCK TYPE", "REACHED PLANT", "DELIVERY DATE",
        "HBL WISE SPLIT COST", "LOADING/UNLOADING CHARGES", "TOLL CHARGES", "PARKING CHARGES",
        "HALTING CHARGES", "TRANSPORT COST", "TOTAL COST", "VENDOR CODE", "REMARK", "BVM JOB NO"
    ],
    "UPS BLR": [
        "SE.No", "Indent Date", "BVM JOB NUMBER", "Customer Name",
        "Air AIR IMPORT &Air AIR EXPORT  &  Ocean AIR EXPORT & DOM",
        "Shipper", "From", "To", "PCS", "Weight", "INVOICE NO", "PO #;",
        "MAWB", "HAWB", "Vehicle Number", "VEHILE  SIZE", "TRIPCOST",
        "TOLL PASS", "AAI charges", "UnLoading charges", "Loading charges",
        "Halting charge", "Handling charges", "Total charges", "BVM IN", "UPDATE STATUS"
    ],
    "FEDEX BLR": [
        "SE.No", "Indent Date", "BVM JOB NUMBER", "Customer Name",
        "Air AIR IMPORT &Air AIR EXPORT  &  Ocean AIR EXPORT & DOM",
        "Shipper", "From", "To", "PCS", "Weight", "INVOICE NO",
        "Vehicle Number", "VEHILE  SIZE", "TRIPCOST", "TOLL PASS",
        "AAI charges", "UnLoading charges", "Loading charges", "Halting charge",
        "Handling charges", "Total charges", "BVM IN", "UPDATE STATUS"
    ],
    "GEODIS BLR": [
        "SE.No", "Indent Date", "BVM JOB NUMBER", "Customer Name",
        "Air AIR IMPORT &Air AIR EXPORT  &  Ocean AIR EXPORT & DOM",
        "Shipper", "From", "To", "PCS", "Weight", "INVOICE NO", "HAWB",
        "Vehicle Number", "VEHILE  SIZE", "TRIPCOST", "TOLL PASS",
        "AAI charges", "UnLoading charges", "Loading charges", "Halting charge",
        "Handling charges", "Total charges", "BVM IN"
    ]
}

DEFAULT_HEADERS = [
    "SR. NO.", "TRIP DATE", "CONSIGNMENT NOTE NO", "CUSTOMER NAME", "CUSTOMER DEPT",
    "SHIPPER", "FROM", "TO", "VEH NO", "VEH TYPE", "TRIPCOST", "AAI CHARGES",
    "UNLOADING CHARGES", "LOADING CHARGES", "HALTING CHARGE", "HANDLING CHARGES",
    "SUPERVISOR CHARGES", "TOTAL CHARGES", "REFERENCE # (JOB ID/HAWB)",
    "VEH REPORTED KM @ LOADING POINT", "VEH REPORTED TIME @ LOADING POINT",
    "LOADING DATE", "LOADING TIME", "VEH REPORTED KM @ UNLOADING POINT",
    "VEH REPORTED TIME @ UNLOADING POINT", "UNLOADING DATE", "UNLOADING TIME",
    "NO OF HALTING DAYS"
]

CUSTOMER_DMR_TEMPLATES = {
    "EIPL": {
        "Air Export": DMR_TEMPLATES["Air Export"],
        "Air Import": DMR_TEMPLATES["Air Import"],
        "Sea Import": DMR_TEMPLATES["Sea Import"],
        "Sea Export": DMR_TEMPLATES["Sea Export"],
        "Order Management": DMR_TEMPLATES["Order Management"],
        "CHB": DMR_TEMPLATES["CHB"],
        "TCS Local": DMR_TEMPLATES["TCS Local"],
        "TCS Outstation": DMR_TEMPLATES["TCS Outstation"],
        "TCS Reefer": DMR_TEMPLATES["TCS Reefer"],
        "Hub Movement": DMR_TEMPLATES["Hub Movement"],  # Added for EIPL
    },
    "TVS": {
        "Hub Movement": DMR_TEMPLATES["Hub Movement"],
    },
    "FORD": {
        "Order Management": DMR_TEMPLATES["Order Management"],
    },
    "CEVA": {
        "Air Import": DMR_TEMPLATES["CEVA Air Import"],
        "Air Export": DMR_TEMPLATES["CEVA Export"], # Note: Original code mapped Air Export to "CEVA Export".
        "Sea Import": DMR_TEMPLATES["CEVA Air Import"], # Note: Original code mapped Sea Import to "CEVA Air Import".
        "Sea Export": DMR_TEMPLATES["CEVA Export"],
    },
    "DHL BLR": {
        "Air Import": DMR_TEMPLATES["DHL BLR Import"],
        "Sea Import": DMR_TEMPLATES["DHL BLR Import"],
        "Air Export": DMR_TEMPLATES["DHL BLR Export"],
        "Sea Export": DMR_TEMPLATES["DHL BLR Export"],
    },
    "DHL": {
        "Air Import": DMR_TEMPLATES["DHL Other"],
        "Air Export": DMR_TEMPLATES["DHL Other"],
        "Sea Import": DMR_TEMPLATES["DHL Sea Import"],
        "Sea Export": DMR_TEMPLATES["DHL Other"],
    },
    "DSVDD": {
        # Catch-all: if customer is DSVDD, use DSV DD REPORT for any department
        "DSVDD": DMR_TEMPLATES["DSV DD REPORT"],
        "Air Export": DMR_TEMPLATES["DSV DD REPORT"],
        "Air Import": DMR_TEMPLATES["DSV DD REPORT"],
        "Sea Export": DMR_TEMPLATES["DSV DD REPORT"],
        "Sea Import": DMR_TEMPLATES["DSV DD REPORT"],
        "CHB": DMR_TEMPLATES["DSV DD REPORT"],
        "Transcon": DMR_TEMPLATES["DSV DD REPORT"],
        "Order Management": DMR_TEMPLATES["DSV DD REPORT"],
        "Transport": DMR_TEMPLATES["DSV DD REPORT"],
    },
    "APMT": {
        # APMT uses APMT template for all departments
        "APMT": DMR_TEMPLATES["APMT"],
        "Air Export": DMR_TEMPLATES["APMT"],
        "Air Import": DMR_TEMPLATES["APMT"],
        "Sea Export": DMR_TEMPLATES["APMT"],
        "Sea Import": DMR_TEMPLATES["APMT"],
        "CHB": DMR_TEMPLATES["APMT"],
        "Transcon": DMR_TEMPLATES["APMT"],
        "Order Management": DMR_TEMPLATES["APMT"],
        "Transport": DMR_TEMPLATES["APMT"],
    },
    "UPS BLR": {
        "UPS BLR": DMR_TEMPLATES["UPS BLR"],
        "Air Export": DMR_TEMPLATES["UPS BLR"],
        "Air Import": DMR_TEMPLATES["UPS BLR"],
        "Sea Export": DMR_TEMPLATES["UPS BLR"],
        "Sea Import": DMR_TEMPLATES["UPS BLR"],
        "Transport": DMR_TEMPLATES["UPS BLR"],
    },
    "FEDEX BLR": {
        "FEDEX BLR": DMR_TEMPLATES["FEDEX BLR"],
        "Air Export": DMR_TEMPLATES["FEDEX BLR"],
        "Air Import": DMR_TEMPLATES["FEDEX BLR"],
        "Sea Export": DMR_TEMPLATES["FEDEX BLR"],
        "Sea Import": DMR_TEMPLATES["FEDEX BLR"],
        "Transport": DMR_TEMPLATES["FEDEX BLR"],
    },
    "GEODIS BLR": {
        "GEODIS BLR": DMR_TEMPLATES["GEODIS BLR"],
        "Air Export": DMR_TEMPLATES["GEODIS BLR"],
        "Air Import": DMR_TEMPLATES["GEODIS BLR"],
        "Sea Export": DMR_TEMPLATES["GEODIS BLR"],
        "Sea Import": DMR_TEMPLATES["GEODIS BLR"],
        "Transport": DMR_TEMPLATES["GEODIS BLR"],
    },
}

COMMON_DMR_MAP = {
    "Air Export": DMR_TEMPLATES.get("Air Export", DEFAULT_HEADERS),
    "Air Import": DMR_TEMPLATES.get("Air Import", DEFAULT_HEADERS),
    "Sea Export": DMR_TEMPLATES.get("Sea Export", DEFAULT_HEADERS),
    "Sea Import": DMR_TEMPLATES.get("Sea Import", DEFAULT_HEADERS),
    "CHB": DMR_TEMPLATES.get("CHB", DEFAULT_HEADERS),
    "Transcon": DEFAULT_HEADERS,
    "Order Management": DMR_TEMPLATES.get("Order Management", DEFAULT_HEADERS),
    "APMT": DMR_TEMPLATES.get("APMT", DEFAULT_HEADERS),
}

ADDITIONAL_CUSTOMERS = [
    "APMT",
    "ABRECO-MAA-WH", "AERTRANS-M", "AIR CARGO-", "AJITH-MAA", "ANOOP-MAA", "BVMPACK-MA",
    "CHROB-MAA", "CARGO-MAA", "CONT-MAA", "DACH-MAA", "DEURGO-MAA", "DHL-MAA", "EIPL-MAA",
    "FREI-MAA", "HAIKO-MAA", "HERPO-MAA", "INSIG-MAA", "JEENA-MAA", "KASA-MAA", "KWE(W)MAA",
    "KRISH-MAA", "LF-MAA", "MARINE-MAA", "MOVE-MAA", "NAVEEN-MAA", "PSKT-MAA", "ROHLIG-MAA",
    "DBS-MAA", "SHIFT-MAA", "TRANSY-MAA", "TVS-MAA", "VILLA-MAA", "VRRDI-MAA", "BOLLO-BLR",
    "CEVA-BLR", "DART-BLR", "DHL-BLR", "DM-BLR", "DSV-BLR", "EIPL-BLR", "FEDEX-BLR",
    "GEODIS-BLR", "HARI-BLR", "HELL-BLR", "HNC-BLR", "KSR-BLR", "KWE-BLR", "KRISH-BLR",
    "POSH-BLR", "RINGO-BLR", "DBS-BLR", "VTL-BLR", "SEAMAN-BLR", "SHIFT-BLR", "SUN-BLR",
    "TVS-BLR", "UPS-BLR", "EIPL-HYD", "EIPL-PNY", "LAU-MAA-PKG", "DBS-MAA-PKG", "GE-MAA-PKG",
    "WIT-MAA-PKG", "POW-MAA-PKG", "WALLA-MAA-PKG", "TULSI-MAA-PKG", "BENZ-MAA-PKG",
    "NORDEX-MAA-PKG", "GEA-MAA-PKG", "METAL-MAA-PKG", "GEODIS-MAA-PKG", "JUSDA-MAA-PKG",
    "PSKT-MAA-PKG", "TECRA-MAA-PKG", "EICK-MAA-PKG", "FLEX-MAA-PKG", "WHEELS-MAA-PKG",
    "VRDHI-MAA-PKG", "BVM-MAA-PKG", "SWAS-MAA-PKG", "EXP-MAA-PKG", "SHIFCO-MAA-PKG",
    "SWE-MAA-PKG", "FEDEX-MAA-PKG", "DIM-MAA-PKG", "ABRECO-MAA-PKG", "MCWNE-MAA-PKG",
    "CRAFT-MAA-PKG", "COOP-MAA-PKG", "MANGAL-MAA-PKG", "FLENDER-MAA-PKG", "TOTAL-MAA-PKG",
    "AARGUS-MAA-TRANS", "CONTINENTAL(T)MAA", "BROEKCLP(T)MAA", "BONFICLP(T)MAA",
    "DHL LCLP(T)MAA", "DSVAIRSEA(T)MAA", "HAIKO(T)MAA", "INSIGHT(T)MAA", "JUSDA(T)MAA",
    "ROBIN(T)MAA", "KWE(T)MAA", "KRISHKO(T)MAA", "DBS(T)MAA", "TRANSYS(T)MAA", "UCS)(T)MAA",
    "FLOMIC - MAA", "FFAF-MAA", "SM-MAA", "Leadking-MAA", "Dimerco-MAA", "MAERSK-MAA",
    "RCS - MAA", "R.M ENT - MAA", "SAR TRANSPORT - MAA", "VKL-MAA", "HIGHLIGHT-MAA",
    "GOODLUCK-MAA", "GE-MAA", "Wen-parker-MAA", "BROEKMAN-MAA", "FCSG-MAA", "DSV-SEA",
    "Freight Bridge -MAA", "APEX -MAA", "Reliance - MAA", "Penta-MAA", "Capricorn-MAA",
    "Gonsai-BLR", "KENSHO-MAA", "AVK-MAA", "MJOSE-MAA", "FLOMICLOG-MAA", "LUCAS-MAA",
    "KERRY-MAA", "AARGUS-MAA", "TOLL - BLR", "GEODIS-MAA", "TRIO-MAA", "JAGJA-MAA",
    "MRLOG-MAA", "APEX -BLR", "NAUT-MAA", "TRANSSAFE-MAA", "GAERISH-MAA", "VINAYAK-MAA",
    "OLS-MAA", "EAGLE-BLR", "LEAAP-MAA", "FEDEX(DD)-BLR", "ALONSO-BLR", "LIGI-MAA",
    "20CUBE-MAA", "ROBINSON-MAA", "KSR-MAA", "SUDHARSHAN-MAA", "EVO-MAA", "CARGOTRANS-MAA",
    "RENISHAW-BLR", "ROHLIG-BLR", "ASW-MAA", "QUICK-MAA", "SIJA-MAA", "SARVAM-MAA",
    "SHREE-MAA", "DAHNAY-MAA", "SA-MAA", "VNAI-MAA", "Asian-Maa - PKG", "DAMCO-MAA",
    "NNR-MAA", "DSV DD-MAA", "MARIANA-MAA", "RYAN-MAA", "TAG-MAA", "LA-MAA",
    "VRRDIc2c-MAA", "FLYCON-MAA", "SCAN-MAA", "India Impex - PKG", "Dukane-PKG",
    "GEODIS(H&M)-BLR", "DAMCO-BLR", "ORBIS-MAA", "Krishya(W)MAA", "SOL(W)MAA", "APM(W)MAA",
    "NILPETER-PKG", "KRR-PKG", "New Customer", "PACKDD-MAA", "APM(W)BLR", "BLUELION(W)BLR",
    "Eshwa- MAA - PKG", "BVM Trans - MAA - PKG", "DHL - MAA - PKG", "DPack - MAA - PKG",
    "Dsv - MAA - PKG", "EIPL-MAA-PKG", "KCP-MAA-PKG", "KWE-MAA-PKG", "Mogli-MAA-PKG",
    "NNR-MAA-PKG", "Nuventura-MAA-PKG", "TASE-MAA-PKG", "Willfred-MAA-PKG", "Bvmtrans(W)MAA",
    "PALANIAPPA(W)MAA", "UNIWORLD(W)MAA", "Scrab-pkg", "IR INDIA-MAA-PKG", "Sri Balaji(W)Blr",
    "AIR CONNECTION(W)MAA", "DBS-MAA(APPLE)", "DBS-MAA(LAND)", "JTB(W)BLR", "ROUND(W)BLR",
    "ANVAL-MAA-PKG", "CARRLANE-MAA-PKG", "Insaplex-MAA-PKG", "Techno Products-MAA-PKG",
    "VFG-MAA-PKG", "Excel-MAA-PKG", "Fretlog-MAA-PKG", "Aircargo-MAA-PKG",
    "Broekman-MAA-PKG", "Capricorn-MAA-PKG", "JEENA-MAA-PKG", "KCP-MAA-PKG", "KRISHKO-MAA-PKG",
    "SA-MAA-PKG", "SS ROADLINE-MAA-PKG", "UPS-MAA-PKG", "Sks-MAA-PKG", "LIMRA-MAA",
    "BRISK-MAA", "SHIFTCO(W)MAA", "NETWORK(W)MAA", "KAYVEE(W)MAA", "BVMTRANS(W)BLR",
    "KGL(T)MAA", "SRISAI(W)BLR", "COMPASS(W)BLR", "INBA(W)BLR", "BNR(W)BLR", "JUSDA(S)BLR",
    "SCANWELL(W)BLR", "INDEV(W)BLR", "BVMPACK(D)-MAA", "SF(E)MAA", "EIPL(T)MAA",
    "DBS(T)MAA", "DSVINTER(T)MAA", "SA GROUP(T)MAA", "GEODIS(T)MAA", "CONTI(T)MAA",
    "UPS(T)MAA", "SUDHARSHAN(T)MAA", "SHIFTCO(T)MAA", "APMT(T)MAA", "4G(T)MAA",
    "AARGUS(T)MAA", "AVA(T)MAA", "AIRCARGO(T)MAA", "BRISK(T)MAA", "BROKEMAN(T)MAA",
    "BROKEMAN CLE(T)MAA", "CEVA(T)MAA", "COMEN(T)MAA", "COMPASS(T)MAA", "TVS(T)MAA",
    "TRIPADAM(T)MAA", "HANKYU(T)MAA", "PROLOGIS(T)MAA", "DEUGRO(T)MAA", "MAGNUM(T)MAA",
    "DACHSER(T)MAA", "FEDEX(T)MAA", "KWE(T)MAA", "FCSG(T)MAA", "BVM PACK(T)MAA",
    "EXECUTIVE(T)MAA", "DUKANE(T)MAA", "RIKUN(T)MAA", "Udhaya Kumar D"
]

# Helper to normalize for matching
def _norm(s):
    return (s or "").lower().replace(" ", "").replace("-", "")

# INITIALIZE DYNAMIC MAPPINGS ON IMPORT
for cust_str in ADDITIONAL_CUSTOMERS:
    found_existing = False
    n_cust = _norm(cust_str)
    for existing_key in CUSTOMER_DMR_TEMPLATES:
        if _norm(existing_key) in n_cust:
            for k, v in COMMON_DMR_MAP.items():
                CUSTOMER_DMR_TEMPLATES[existing_key].setdefault(k, v)
            found_existing = True
            break
    if not found_existing:
        CUSTOMER_DMR_TEMPLATES[cust_str] = COMMON_DMR_MAP.copy()

    if "apmt" in n_cust:
        if cust_str in CUSTOMER_DMR_TEMPLATES:
            CUSTOMER_DMR_TEMPLATES[cust_str]["Transport"] = DMR_TEMPLATES.get("APMT", DEFAULT_HEADERS)
        else:
            for existing_key in CUSTOMER_DMR_TEMPLATES:
                if _norm(existing_key) in n_cust:
                    CUSTOMER_DMR_TEMPLATES[existing_key]["Transport"] = DMR_TEMPLATES.get("APMT", DEFAULT_HEADERS)

    if "dsv" in n_cust:
        if cust_str in CUSTOMER_DMR_TEMPLATES:
            if "dsv" not in CUSTOMER_DMR_TEMPLATES[cust_str]:
                CUSTOMER_DMR_TEMPLATES[cust_str]["DSV"] = DMR_TEMPLATES.get("DSV", DEFAULT_HEADERS)
        if "dsvdd" in n_cust:
            if cust_str in CUSTOMER_DMR_TEMPLATES:
                CUSTOMER_DMR_TEMPLATES[cust_str]["DSV DD"] = DMR_TEMPLATES.get("DSV DD REPORT", DEFAULT_HEADERS)


# -------------------------------------------------------------------------
# VIEW HELPERS
# -------------------------------------------------------------------------

def get_dmr_headers(customer_name, dept_name, from_loc_id, to_loc_id):
    # 1. Resolve Location Names (for Route matching)
    from_loc_name = ""
    to_loc_name = ""
    if from_loc_id:
        try:
            from_obj = Places.objects.get(id=from_loc_id)
            from_loc_name = from_obj.place_name
        except:
            pass
    if to_loc_id:
        try:
            to_obj = Places.objects.get(id=to_loc_id)
            to_loc_name = to_obj.place_name
        except:
            pass

    # Clean inputs
    cust_value = _norm(customer_name)
    n_dept = _norm(dept_name)

    # 3. Construct Search Strings
    n_route = _norm(f"{from_loc_name}-{to_loc_name}") if from_loc_name or to_loc_name else ""
    n_combined = ""
    if n_dept and n_route:
        n_combined = _norm(f"{dept_name}_{from_loc_name}-{to_loc_name}")
    elif n_route:
        n_combined = n_route
    elif n_dept:
        n_combined = n_dept

    template_key = None
    headers = DEFAULT_HEADERS

    # Search Logic
    for cust_key, dept_map in CUSTOMER_DMR_TEMPLATES.items():
        if _norm(cust_key) in cust_value:
            matches = []
            for key, template in dept_map.items():
                n_key = _norm(key)
                prio = 0
                if n_combined and n_key in n_combined:
                    is_simple_dept = n_dept and (n_key in n_dept)
                    is_simple_route = n_route and (n_key in n_route)
                    if not is_simple_dept and not is_simple_route:
                        prio = 3
                    elif is_simple_route:
                        prio = 2
                    elif is_simple_dept:
                        prio = 1
                if prio == 0 and n_key in cust_value:
                    prio = 1
                if prio > 0:
                    matches.append((prio, len(n_key), key, template))

            if matches:
                matches.sort(key=lambda x: (x[0], x[1]), reverse=True)
                template_key = matches[0][2]
                headers = matches[0][3]
            elif dept_map:
                # If no specific department match found, fallback to the first template defined for this customer
                template_key = next(iter(dept_map))
                headers = dept_map[template_key]
            break

    return headers, template_key

def safe(value):
    if value is None: return ""
    return str(value)

def safe_str(v):
    return "" if v is None else str(v)

def safe_num(v):
    try: return float(v) if v not in ("", None, "None") else 0
    except: return 0

def get_dmr_rows(trips, headers, template_key, customer_name):
    rows = []
    # Fetch all ConsignmentgoodsInfo and Vehicle_allotmentInfo for the current page in bulk
    trip_cons_nums = [t.tr_consignmentnumber_id for t in trips if t.tr_consignmentnumber_id]
    trip_enq_nums = [t.tr_enquirynumber_id for t in trips]
    
    goods_map = {g.cg_consignmentnumber_id: g for g in ConsignmentgoodsInfo.objects.filter(cg_consignmentnumber_id__in=trip_cons_nums)}
    va_map = {v.va_enquirynumber_id: v for v in Vehicle_allotmentInfo.objects.filter(va_enquirynumber_id__in=trip_enq_nums).select_related('va_vendor', 'va_driver')}
    cons_detail_map = {c.id: c for c in ConsignmentdetailInfo.objects.filter(id__in=trip_cons_nums)}

    for idx, trip in enumerate(trips, start=1):
        cons_detail = cons_detail_map.get(trip.tr_consignmentnumber_id)
        cons_goods = goods_map.get(trip.tr_consignmentnumber_id)
        va = va_map.get(trip.tr_enquirynumber_id)

        row = []
        for h in headers:
            hh = " ".join(h.strip().lower().split())
            if hh in ("s.no", "sr. no.", "so no", "se.no"):
                row.append(idx); continue
            
            # --- DATES & TIMES ---
            # 1. Generic Trip Dates (Start of Trip)
            if hh in ("trip date", "indent date", "date", "ofd date"):
                # Anchor 'Date' to the arrival at shipper (start of the operation)
                row.append(trip.tr_loading_time.strftime("%d-%m-%Y") if trip.tr_loading_time else (trip.tr_departeddate.strftime("%d-%m-%Y") if trip.tr_departeddate else "")); continue
            
            # 2. Pickup Point (Arrival at Shipper - IN)
            if hh in ("pickup date", "pickup point in date", "in date", "airport/bvm gate in date", "loading date"):
                row.append(trip.tr_loading_time.strftime("%d-%m-%Y") if trip.tr_loading_time else ""); continue
            if hh in ("pickup point in time", "in time", "airport/bvm gate in time", "loading time", "veh reported time @ loading point"):
                row.append(trip.tr_loading_time.strftime("%H:%M") if trip.tr_loading_time else ""); continue

            # 3. Pickup Point (Departure from Shipper - OUT)
            if hh in ("pickup point out date", "out date"):
                row.append(trip.tr_departeddate.strftime("%d-%m-%Y") if trip.tr_departeddate else ""); continue
            if hh in ("pickup point out time", "out time", "starting time", "closing time"):
                row.append(trip.tr_departeddate.strftime("%H:%M") if trip.tr_departeddate else ""); continue

            # 4. Unloading Point (Arrival at Destination - IN)
            if hh in ("unloading point in date", "gate in date", "reached plant", "reached cfs", "reached", "cfs reached date", "closing place", "closing time"):
                # Handling 'Today > Tomorrow' logic: Always pull from the specific event timestamp
                if "time" in hh:
                    row.append(trip.tr_reporteddate.strftime("%H:%M") if trip.tr_reporteddate else ""); continue
                row.append(trip.tr_reporteddate.strftime("%d-%m-%Y") if trip.tr_reporteddate else ""); continue
            
            if hh in ("unloading point in time", "in time @ unloading point", "gate in time", "cfs reached time", "closing time"):
                row.append(trip.tr_reporteddate.strftime("%H:%M") if trip.tr_reporteddate else ""); continue

            # 5. Unloading Point (Departure from Destination - OUT)
            if hh in ("unloading point out date", "dlv out date", "delivery date", "released date"):
                row.append(trip.tr_unloading_time.strftime("%d-%m-%Y") if trip.tr_unloading_time else ""); continue
            if hh in ("dlv out time", "unloading time", "unloading point out time", "released time"):
                row.append(trip.tr_unloading_time.strftime("%H:%M") if trip.tr_unloading_time else ""); continue

            if hh == "bvm in":
                # Special handler for 'BVM IN' which refers to arrival at hub/warehouse
                v = trip.tr_reporteddate # BVM Hub Arrival is modeled as tr_reporteddate in most templates
                if not v: row.append(""); continue
                fmt = "%d-%m-%Y" if "date" in hh else "%H:%M" if "time" in hh else "%d-%m-%Y %H:%M"
                row.append(v.strftime(fmt)); continue

            if "consignment note" in hh or "cnote" in hh:
                row.append(safe_str(cons_detail.co_consignmentnumber) if cons_detail else ""); continue
            if hh == "trip sheet no":
                row.append(safe_str(trip.tr_tripnumber)); continue
            if hh == "customer name":
                row.append(safe_str(trip.tr_enquirynumber.en_customername)); continue
            if hh == "customer dept":
                row.append(safe_str(trip.tr_enquirynumber.en_customerdepartment)); continue
            if hh == "tripcost":
                row.append(safe_num(trip.tc_tripcost)); continue
            if hh in ("job no", "bvm job no", "bvm job number", "ceva job no", "bvm job"):
                row.append(safe_str(trip.tr_tripnumber)); continue
            if "department name" in hh or hh == "dept" or hh == "air air import &air air export & ocean air export & dom" or hh == "division":
                row.append(safe_str(trip.tr_enquirynumber.en_customerdepartment)); continue

            if hh in ("veh reported km @ loading point", "starting km"):
                row.append(safe_str(trip.tr_departedkm)); continue
            if hh in ("veh reported km @ unloading point", "closing km"):
                row.append(safe_str(trip.tr_reportedkm)); continue
            if hh == "used km":
                diff = (trip.tr_reportedkm or 0) - (trip.tr_departedkm or 0)
                row.append(safe_num(max(0, diff))); continue

            if hh in ("from", "origin", "starting place", "pickup location", "orgin"):
                row.append(safe_str(trip.tr_departedlocation)); continue
            if hh in ("unloading point", "unloading point", "unloading  point"):
                row.append(safe_str(trip.tr_reportedlocation)); continue
            if hh in ("delivery location", "to", "destination", "delivery place", "delivery point", "closing place"):
                if cons_goods and getattr(cons_goods, "cg_deliverylocation", None):
                    row.append(safe(cons_goods.cg_deliverylocation))
                elif trip.tr_reportedlocation:
                    row.append(safe(trip.tr_reportedlocation))
                else:
                    row.append(safe(trip.tr_enquirynumber.en_tolocation))
                continue
            if hh == "planning received date":
                en_date = getattr(trip.tr_enquirynumber, "en_created_at", None)
                row.append(en_date.strftime("%d-%m-%Y") if en_date else ""); continue
            if hh == "planning received time":
                en_date = getattr(trip.tr_enquirynumber, "en_created_at", None)
                row.append(en_date.strftime("%H:%M") if en_date else ""); continue

            # --- SHIPPER / CONSIGNEE ---
            if "shipper seal #" in hh or hh == "seel no" or hh == "seal no":
                val = (cons_detail.co_seal_number or cons_detail.co_smart_lock_number) if cons_detail else ""
                row.append(safe_str(val)); continue
            if "shipper" in hh or hh == "consignor":
                row.append(safe_str(cons_goods.cg_consigner) if cons_goods else ""); continue
            if "llr no" in hh or "lr no" in hh:
                row.append(safe_str(cons_detail.co_consignmentnumber) if cons_detail else ""); continue
            if "consignee" in hh:
                row.append(safe_str(cons_goods.cg_consignee) if cons_goods else ""); continue
            if "cs name" in hh or hh == "cutomer service":
                row.append(safe_str(trip.tr_enquirynumber.en_assignedto)); continue
            if hh in ("vehicle placed time", "placement & vehicle placed date"):
                 fmt = "%d-%m-%Y" if "date" in hh else "%H:%M"
                 row.append(va.va_created_at.strftime(fmt) if va and va.va_created_at else ""); continue
            if "hawb" in hh or "hbl" in hh:
                row.append(safe_str(cons_goods.cg_hawbno) if cons_goods else ""); continue
            if "boe" in hh or "ewaybill" in hh or hh == "e-way bill":
                row.append(safe_str(cons_goods.cg_ebillno) if cons_goods else ""); continue
            if "reference" in hh or hh == "hbl no/reference no":
                row.append(safe_str(cons_detail.co_cusrefnum if cons_detail else "")); continue
            if "truck no" in hh or "veh no" in hh or hh in ("vehicle no.", "vehicle no", "vehicle number", "veh.no", "veh no"):
                row.append(safe_str(trip.tr_vehiclenumber)); continue
            if "truck type" in hh or "veh type" in hh or hh in ("container size", "vehicle type", "vehiile size", "vehicle type"):
                row.append(safe_str(trip.tr_vehicletype)); continue
            if hh in ("booking no", "po #;", "bvm invoice"):
                row.append(safe_str(trip.tr_customerref or (cons_detail.co_cusrefnum if cons_detail else ""))); continue

            # --- UPS / FEDEX / GEODIS BLR SPECIFIC ---
            if hh == "pcs":
                row.append(safe_str(cons_goods.cg_qty) if cons_goods else ""); continue
            if hh == "weight":
                row.append(safe_str(cons_goods.cg_weight) if cons_goods else ""); continue
            if hh == "invoice no":
                row.append(safe_str(cons_detail.co_cusrefnum) if cons_detail else ""); continue
            if hh == "mawb":
                row.append(safe_str(cons_goods.cg_mawbno) if cons_goods else ""); continue
            if hh == "vehile size":
                row.append(safe_str(trip.tr_vehicletype_placed or trip.tr_vehicletype)); continue
            if hh == "toll pass":
                row.append(safe_num(trip.tc_tollcost)); continue
            if hh == "aai charges":
                row.append(safe_num(trip.tc_supervisorcost)); continue
            if hh == "supervisor charges":
                row.append(safe_num(trip.tc_supervisorcost)); continue
            if hh == "unloading charges":
                row.append(safe_num(trip.tc_unloadingcost)); continue
            if hh == "loading charges":
                row.append(safe_num(trip.tc_loadingcost)); continue
            if hh == "halting charge":
                row.append(safe_num(trip.tc_haltingcost)); continue
            if hh == "handling charges":
                row.append(safe_num(trip.tc_handlingcost)); continue
            if hh == "bvm in":
                row.append(trip.tr_reporteddate.strftime("%d-%m-%Y") if "date" in hh else trip.tr_reporteddate.strftime("%H:%M") if "time" in hh else trip.tr_reporteddate.strftime("%d-%m-%Y %H:%M") if trip.tr_reporteddate else ""); continue
            if hh == "update status":
                row.append(safe_str(trip.tc_financestatus)); continue

            # --- GENERIC FIELDS ---
            if "vendor" in hh or "transporter" in hh:
                vendor = safe_str(va.va_vendor) if (va and trip.tr_vehiclesource_id in (2, 3)) else "OWN VEHICLE"
                row.append(vendor); continue
            if "driver name" in hh:
                val = trip.tr_drivername or (va.va_drivername if va else "") or (va.va_driver.dm_name if va and va.va_driver else "")
                row.append(safe_str(val)); continue
            if "driver mobile" in hh or "driver number" in hh or hh == "driver no.":
                val = trip.tr_drivernumber or (va.va_drivernumber if va else "") or (va.va_driver.dm_drivernumber if va and va.va_driver else "")
                row.append(safe_str(val)); continue
            if "driver dl" in hh:
                val = trip.tr_driver_lic or (va.va_driver_lic if va else "") or (va.va_driver.dm_driver_lic if va and va.va_driver else "")
                row.append(safe_str(val)); continue
            if hh in ("trip number", "bvm job", "bvm job no", "bvm job number"):
                row.append(safe_str(trip.tr_tripnumber)); continue
            if "no of pieces" in hh or hh in ("pcs", "no pkg", "pkgs", "sum of pieces"):
                qty = cons_goods.cg_loaded_qty if cons_goods and cons_goods.cg_loaded_qty else (cons_goods.cg_qty if cons_goods else "")
                row.append(safe_str(qty)); continue
            if "actual weight" in hh or "invoice weight" in hh or "gross weight" in hh or "cargo weight" in hh or hh in ("weight", "g weight", "actual weight (kgs)"):
                row.append(safe_str(cons_goods.cg_weight) if cons_goods else ""); continue
            if "chargeable weight" in hh:
                row.append(""); continue
            if "cbm" in hh or "volume" in hh:
                row.append(""); continue

            # --- UNLOADING ---
            if hh in ("unloading point in date", "gate in date", "reached plant", "reached cfs", "reached", "cfs reached date"):
                row.append(trip.tr_reporteddate.strftime("%d-%m-%Y") if trip.tr_reporteddate else ""); continue
            if hh in ("unloading point in time", "in time @ unloading point", "gate in time", "cfs reached time"):
                row.append(trip.tr_reporteddate.strftime("%H:%M") if trip.tr_reporteddate else ""); continue
            if hh in ("unloading point out date", "dlv out date", "delivery date"):
                row.append(trip.tr_unloading_time.strftime("%d-%m-%Y") if trip.tr_unloading_time else ""); continue
            if hh in ("dlv out time", "unloading time", "unloading time", "unloading point out time") or (hh == "unloading time" and "time" in hh):
                row.append(trip.tr_unloading_time.strftime("%H:%M") if trip.tr_unloading_time else ""); continue

            # --- CHARGES ---
            if "no of days halting" in hh:
                row.append(safe_num(trip.tc_no_of_days_halting)); continue
            if "additional charges" in hh:
                row.append(safe_num(trip.tc_handlingcost)); continue
            if "cancellation charges" in hh or "cancelling charges" in hh:
                row.append(safe_num(trip.tc_cancellation)); continue
            if "halting charges" in hh:
                row.append(safe_num(trip.tc_haltingcost)); continue
            if hh == "charges":
                row.append(safe_num(trip.tc_tripcost)); continue
            if "weightment charges" in hh or hh == "weighment pass":
                row.append(safe_num(trip.tc_weighmentcost)); continue
            if hh == "parking charges":
                row.append(safe_num(trip.tc_parkingcost)); continue
            if hh == "unloading charges":
                row.append(safe_num(trip.tc_unloadingcost)); continue
            if "parking / unloading charges" in hh or "unloading charges and parking charges" in hh:
                row.append(safe_num(trip.tc_parkingcost) + safe_num(trip.tc_unloadingcost)); continue
            if "unloading charges & lashing charges" in hh:
                row.append(safe_num(trip.tc_unloadingcost)); continue
            if "total charges" in hh:
                row.append(safe_num(trip.tc_tripcost) + safe_num(trip.tc_parkingcost) + safe_num(trip.tc_unloadingcost) + safe_num(trip.tc_loadingcost) + safe_num(trip.tc_weighmentcost) + safe_num(trip.tc_handlingcost) + safe_num(trip.tc_supervisorcost) + safe_num(trip.tc_haltingcost) + safe_num(trip.tc_tollcost)); continue

            # --- REMARKS & STATUS ---
            if hh in ("remarks", "remark", "comments", "pod remarks", "return box"):
                row.append(safe_str(cons_detail.co_remarks) if cons_detail else safe_str(trip.tr_remarks)); continue
            if hh in ("delivery status", "update status", "pod status", "pod"):
                status_val = str(trip.tc_financestatus.status if trip.tc_financestatus else "PENDING")
                row.append(status_val); continue
            if hh == "delay&ontime":
                row.append(""); continue # No direct field, could calculate later if needed

            row.append("")
        rows.append(row)
    return rows


# -------------------------------------------------------------------------
# VIEWS
# -------------------------------------------------------------------------

@login_required(login_url='login_page')
def trip_report(request):
    first_name = request.session.get('first_name')
    form = DmrForm(request.POST or None)

    # -------------------------
    # GET FILTER VALUES FROM FORM
    # -------------------------
    customer_id = request.POST.get('dmr_customer')  # Correct field name from DmrForm
    dept_id = request.POST.get('customer_department')
    selected_month = request.POST.get('month')
    selected_year = request.POST.get('year')
    from_loc = request.POST.get('from_location')
    to_loc = request.POST.get('to_location')

    # -------------------------
    # GET CUSTOMER + DEPT NAMES
    # -------------------------
    cust_name = ""
    dept_name = ""

    if customer_id:
        try:
            cust_name = CustomerInfo.objects.get(id=customer_id).cu_name
        except:
            cust_name = ""

    if dept_id:
        try:
            dept_name = CustomerdepartmentInfo.objects.get(id=dept_id).ct_customerdepartment
        except:
            dept_name = ""

    cust_value = (cust_name or "").lower()
    dept_value = (dept_name or "").lower()

    # -------------------------
    # BASE QUERY
    # -------------------------
    trips = TripdetailInfo.objects.filter(tr_category_id=1).select_related(
        'tr_enquirynumber', 
        'tr_enquirynumber__en_customername',
        'tr_enquirynumber__en_customerdepartment',
        'tr_enquirynumber__en_fromlocaion',
        'tr_enquirynumber__en_tolocation',
        'tr_consignmentnumber',
        'tr_departedlocation',
        'tr_reportedlocation',
        'tc_financestatus'
    ).order_by('-tr_tripnumber')

    if customer_id:
        trips = trips.filter(tr_enquirynumber__en_customername_id=customer_id)
    if dept_id:
        trips = trips.filter(tr_enquirynumber__en_customerdepartment_id=dept_id)
    if selected_month and selected_year:
        selected_month = int(selected_month)
        selected_year = int(selected_year)
        first_day = date(selected_year, selected_month, 1)
        last_day = date(selected_year, selected_month, calendar.monthrange(selected_year, selected_month)[1])
        trips = trips.filter(
            tr_departeddate__date__gte=first_day,
            tr_departeddate__date__lte=last_day
        )
    if from_loc:
        trips = trips.filter(tr_enquirynumber__en_fromlocaion_id=from_loc)
    if to_loc:
        trips = trips.filter(tr_enquirynumber__en_tolocation_id=to_loc)

    # -------------------------
    # PAGINATION
    # -------------------------
    paginator = Paginator(trips, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # -------------------------
    # ATTACH CONSIGNER + REF NO (ONLY for paginated objects)
    # -------------------------
    page_cons_nums = [t.tr_consignmentnumber_id for t in page_obj if t.tr_consignmentnumber_id]
    goods_map = {g.cg_consignmentnumber_id: g for g in ConsignmentgoodsInfo.objects.filter(cg_consignmentnumber_id__in=page_cons_nums)}
    
    for trip in page_obj:
        cg = goods_map.get(trip.tr_consignmentnumber_id)
        trip.consigner_name = str(cg.cg_consigner) if cg and cg.cg_consigner else ""
        trip.co_cusrefnum = trip.tr_consignmentnumber.co_cusrefnum if trip.tr_consignmentnumber else ""

    current_month = datetime.now().month
    current_year = datetime.now().year

    # -------------------------
    # PREPARE DYNAMIC HEADERS & ROWS (for HTML View)
    # -------------------------
    cust_name = ""
    dept_name = ""
    if customer_id:
        try: cust_name = CustomerInfo.objects.get(id=customer_id).cu_name
        except: pass
    if dept_id:
        try: dept_name = CustomerdepartmentInfo.objects.get(id=dept_id).ct_customerdepartment  # FIXED
        except: pass

    # Fallback to trip data if filter names missing
    if not cust_name and page_obj:
        try: cust_name = str(page_obj[0].tr_enquirynumber.en_customername)
        except: pass
    if not dept_name and page_obj:
        try: dept_name = str(page_obj[0].tr_enquirynumber.en_customerdepartment)
        except: pass

    headers, template_key = get_dmr_headers(cust_name, dept_name, from_loc, to_loc)
    data_rows = get_dmr_rows(page_obj, headers, template_key, cust_name)

    context = {
        'first_name': first_name,
        'form': form,
        'page_obj': page_obj,
        'headers': headers,
        'data_rows': data_rows,
        'customer_id': customer_id or '',
        'dept_id': dept_id or '',
        'selected_month': int(selected_month) if selected_month else current_month,
        'selected_year': int(selected_year) if selected_year else current_year,
        'years': range(current_year - 5, current_year + 1),
        'from_location': int(from_loc) if from_loc else '',
        'to_location': int(to_loc) if to_loc else '',
    }
    return render(request, "asset_mgt_app/dmr_report.html", context)


@login_required(login_url='login_page')
def trip_send_email(request):
    if request.method != 'POST':
        messages.error(request, "Invalid request.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    customer_id = request.POST.get('customer_id')
    dept_id = request.POST.get('customer_department') or request.POST.get('dept_id')
    month = request.POST.get('month')
    year = request.POST.get('year')
    from_loc = request.POST.get('from_location')
    to_loc = request.POST.get('to_location')
    recipient = request.POST.get('recipient', "")
    subject = request.POST.get('subject', "")
    message_body = request.POST.get('message', "")

    if not customer_id:
        messages.error(request, "Please select a customer.")
        return redirect('trip_report')

    try:
        customer_obj = CustomerInfo.objects.get(id=customer_id)
    except CustomerInfo.DoesNotExist:
        messages.error(request, "Customer not found.")
        return redirect('trip_report')

    # Query Trips
    qs = TripdetailInfo.objects.filter(
        tr_enquirynumber__en_customername_id=customer_id,
        tr_category_id=1
    ).order_by('tr_departeddate') # Date sort

    if dept_id:
        try: qs = qs.filter(tr_enquirynumber__en_customerdepartment_id=int(dept_id))
        except: qs = qs.filter(tr_enquirynumber__en_customerdepartment__icontains=str(dept_id))

    if month and year:
        try:
            m, y = int(month), int(year)
            first_day = date(y, m, 1)
            last_day = date(y, m, calendar.monthrange(y, m)[1])
            qs = qs.filter(tr_departeddate__date__gte=first_day, tr_departeddate__date__lte=last_day)
        except: pass

    if from_loc: qs = qs.filter(tr_enquirynumber__en_fromlocaion_id=from_loc)
    if to_loc: qs = qs.filter(tr_enquirynumber__en_tolocation_id=to_loc)

    trips = list(qs)

    # Identify Template
    department = CustomerdepartmentInfo.objects.filter(id=dept_id).first()
    dept_name = str(department) if department else None

    headers, template_key = get_dmr_headers(customer_obj.cu_name, dept_name, from_loc, to_loc)

    # Prepare Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "DMR Report"

    # Styles
    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = yellow_fill
        cell.border = border
        cell.alignment = center_align

    # Write Rows
    rows = get_dmr_rows(trips, headers, template_key, customer_obj.cu_name)
    for row in rows:
        ws.append(row)

    # Auto Width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    recipient_list = [x.strip() for x in recipient.split(",") if x.strip()]
    subject = subject or f"{customer_obj.cu_name} - {template_key or 'DMR'} Report"
    message = message_body.replace("\n", "<br>")

    send_department_email(
        department='itadmin',
        subject=subject,
        message=message,
        recipient_list=recipient_list,
        attachment=excel_file,
        attachment_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        file_name=f"{customer_obj.cu_name}_DMR_Report.xlsx"
    )
    messages.success(request, "DMR Report sent successfully.")
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required(login_url='login_page')
def get_dmr_email_details(request):
    customer_id = request.GET.get("customer_id")
    department_id = request.GET.get("department_id")
    if not customer_id: return JsonResponse({'status': False})

    qs = Emailmaster.objects.filter(em_Customer_name_id=customer_id)
    if department_id:
        qs = qs.filter(em_customerdepartment_id=department_id)

    email_obj = qs.first()
    if not email_obj or not email_obj.em_to_names:
        return JsonResponse({'status': False})

    return JsonResponse({'status': True, 'to_mail': email_obj.em_to_names})


def get_dmr_template(customer_name, dept_name):
    # Backward compatibility wrapper
    h, k = get_dmr_headers(customer_name, dept_name, None, None)
    return h